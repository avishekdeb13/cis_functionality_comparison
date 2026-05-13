"""
Universal Portal Testing Tool v1.5
- GUI login window
- Second GUI: user provides Diary No, Case No for form testing
- Login with CAPTCHA support
- Expand all sidebar menus
- Extract all components per page
- Deep test: click menus, fill forms, submit, screenshot results
- Excel + CSV named after login user
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select as SeleniumSelect
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import tkinter as tk
from tkinter import ttk, messagebox
import time, os, csv, re
from datetime import datetime
from urllib.parse import urlparse


# ── STEP 1A: GUI — Login credentials ─────────────────────────────────────────
def get_config():
    config = {}

    def submit():
        url  = url_var.get().strip()
        user = user_var.get().strip()
        pwd  = pass_var.get().strip()
        if not url or not user or not pwd:
            messagebox.showerror("Error", "All fields are required")
            return
        config["url"]      = url
        config["username"] = user
        config["password"] = pwd
        ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
        folder = user + "_" + ts
        ss_dir = os.path.join(folder, "screenshots")
        os.makedirs(ss_dir, exist_ok=True)
        config["folder"]      = folder
        config["screenshots"] = ss_dir
        config["html"]        = os.path.join(folder, user + "_portal.html")
        config["csv"]         = os.path.join(folder, user + "_menu.csv")
        config["excel"]       = os.path.join(folder, user + "_test_report.xlsx")
        config["base"]        = urlparse(url).scheme + "://" + urlparse(url).netloc
        root.destroy()

    root = tk.Tk()
    root.title("Portal Testing Tool v1.5 - Step 1: Login")
    root.geometry("560x340")
    root.resizable(False, False)

    tk.Label(root, text="Universal Portal Testing Tool v1.5",
             font=("Arial", 15, "bold"), fg="#003366").pack(pady=12)
    tk.Label(root, text="Step 1 of 2 — Enter login credentials",
             font=("Arial", 10), fg="#555").pack()

    frame = ttk.Frame(root, padding=20)
    frame.pack(fill="both", expand=True)

    url_var  = tk.StringVar(value="https://drt.etribunals.gov.in/cis2.0/filing/login")
    user_var = tk.StringVar(value="")
    pass_var = tk.StringVar(value="")

    ttk.Label(frame, text="Portal URL:").grid(row=0, column=0, sticky="w", pady=8)
    ttk.Entry(frame, textvariable=url_var, width=48).grid(row=0, column=1, padx=8)
    ttk.Label(frame, text="Username:").grid(row=1, column=0, sticky="w", pady=8)
    ttk.Entry(frame, textvariable=user_var, width=48).grid(row=1, column=1, padx=8)
    ttk.Label(frame, text="Password:").grid(row=2, column=0, sticky="w", pady=8)
    ttk.Entry(frame, textvariable=pass_var, show="*", width=48).grid(row=2, column=1, padx=8)

    ttk.Button(frame, text="Next  ▶", command=submit).grid(
        row=3, column=1, pady=20, sticky="e")

    root.mainloop()
    return config


# ── STEP 1B: GUI — Test data input ────────────────────────────────────────────
def get_test_data():
    data = {}

    def submit():
        data["diary_no"]   = diary_var.get().strip()
        data["case_no"]    = case_var.get().strip()
        data["case_year"]  = year_var.get().strip()
        data["from_date"]  = from_var.get().strip()
        data["to_date"]    = to_var.get().strip()
        data["party_name"] = party_var.get().strip()
        data["skip_deep"]  = skip_var.get()
        root2.destroy()

    def skip_all():
        data["diary_no"]   = ""
        data["case_no"]    = ""
        data["case_year"]  = ""
        data["from_date"]  = ""
        data["to_date"]    = ""
        data["party_name"] = ""
        data["skip_deep"]  = True
        root2.destroy()

    root2 = tk.Tk()
    root2.title("Portal Testing Tool v1.5 - Step 2: Test Data")
    root2.geometry("560x420")
    root2.resizable(False, False)

    tk.Label(root2, text="Step 2 of 2 — Provide Test Data",
             font=("Arial", 15, "bold"), fg="#003366").pack(pady=12)
    tk.Label(root2,
             text="This data will be used to fill forms during testing.\nLeave blank to skip form submission tests.",
             font=("Arial", 9), fg="#555", justify="center").pack()

    frame = ttk.Frame(root2, padding=20)
    frame.pack(fill="both", expand=True)

    diary_var  = tk.StringVar()
    case_var   = tk.StringVar()
    year_var   = tk.StringVar(value=str(datetime.now().year))
    from_var   = tk.StringVar(value="01/01/2026")
    to_var     = tk.StringVar(value="31/12/2026")
    party_var  = tk.StringVar()
    skip_var   = tk.BooleanVar(value=False)

    fields = [
        ("Diary Number:",    diary_var,  "e.g. 100/2026"),
        ("Case Number:",     case_var,   "e.g. OA/5/2026"),
        ("Case Year:",       year_var,   "e.g. 2026"),
        ("From Date:",       from_var,   "DD/MM/YYYY"),
        ("To Date:",         to_var,     "DD/MM/YYYY"),
        ("Party Name:",      party_var,  "e.g. BANK OF INDIA"),
    ]
    for ri, (lbl, var, hint) in enumerate(fields):
        ttk.Label(frame, text=lbl).grid(row=ri, column=0, sticky="w", pady=5)
        e = ttk.Entry(frame, textvariable=var, width=30)
        e.grid(row=ri, column=1, padx=8, sticky="w")
        ttk.Label(frame, text=hint, foreground="#999",
                  font=("Arial", 8)).grid(row=ri, column=2, sticky="w")

    ttk.Checkbutton(frame, text="Skip form submission tests (component extraction only)",
                    variable=skip_var).grid(row=len(fields), column=0,
                                            columnspan=3, pady=8, sticky="w")

    btn_frame = ttk.Frame(frame)
    btn_frame.grid(row=len(fields)+1, column=0, columnspan=3, pady=10, sticky="e")
    ttk.Button(btn_frame, text="Skip All Tests", command=skip_all).pack(side="left", padx=5)
    ttk.Button(btn_frame, text="Start Testing  ▶", command=submit).pack(side="left")

    root2.mainloop()
    return data


# ── STEP 2: Browser ───────────────────────────────────────────────────────────
def init_driver():
    opts = webdriver.ChromeOptions()
    opts.add_argument("--start-maximized")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=opts)


# ── STEP 3: Login ─────────────────────────────────────────────────────────────
def login(driver, cfg):
    print("\n[1] Opening login page...")
    driver.get(cfg["url"])
    time.sleep(3)

    for fid in ["user_name", "username"]:
        try:
            driver.find_element(By.ID, fid).send_keys(cfg["username"])
            print("    Username filled"); break
        except: pass

    for fid in ["user_pass", "password"]:
        try:
            driver.find_element(By.ID, fid).send_keys(cfg["password"])
            print("    Password filled"); break
        except: pass

    driver.save_screenshot(os.path.join(cfg["screenshots"], "00_login.png"))
    print("\n" + "="*52)
    print("  Solve CAPTCHA then click LOGIN")
    print("  Waiting 120 seconds...")
    print("="*52 + "\n")

    WebDriverWait(driver, 120).until(lambda d: "login" not in d.current_url)
    time.sleep(3)
    driver.save_screenshot(os.path.join(cfg["screenshots"], "01_dashboard.png"))
    print("[OK] Logged in:", driver.current_url)


# ── STEP 4: Expand sidebar ────────────────────────────────────────────────────
def get_html(driver, cfg):
    print("\n[2] Expanding ALL sidebar menus...")

    driver.execute_script("""
        document.querySelectorAll('.collapse').forEach(el => el.classList.add('show'));
        document.querySelectorAll('[aria-expanded]').forEach(el =>
            el.setAttribute('aria-expanded', 'true'));
    """)
    time.sleep(1)

    toggles = driver.find_elements(By.CSS_SELECTOR, "[data-bs-toggle='collapse']")
    print("    Found", len(toggles), "collapsible toggles")
    for toggle in toggles:
        try:
            driver.execute_script("arguments[0].scrollIntoView(true);", toggle)
            time.sleep(0.1)
            target = (toggle.get_attribute("data-bs-target") or "").replace("#", "")
            if target:
                try:
                    el = driver.find_element(By.ID, target)
                    if "show" not in (el.get_attribute("class") or ""):
                        driver.execute_script("arguments[0].click();", toggle)
                        time.sleep(0.3)
                except:
                    driver.execute_script("arguments[0].click();", toggle)
                    time.sleep(0.3)
            print("    [+]", toggle.text.strip()[:40] or target)
        except: pass

    driver.execute_script("""
        document.querySelectorAll('.collapse').forEach(el => el.classList.add('show'));
    """)
    time.sleep(2)

    driver.save_screenshot(os.path.join(cfg["screenshots"], "02_sidebar_expanded.png"))
    html = driver.page_source
    with open(cfg["html"], "w", encoding="utf-8") as f:
        f.write(html)

    soup  = BeautifulSoup(html, "html.parser")
    items = soup.find_all("a", class_="dropdown-item")
    print("    Captured", len(items), "dropdown items")
    for item in items:
        print("      -", item.get_text(strip=True), "->", item.get("href","")[:60])
    return html


# ── STEP 5: URL cleaner ───────────────────────────────────────────────────────
def clean_url(href, base):
    if not href or not href.strip():
        return "", "Empty href"
    href = href.strip()
    if href.lower().startswith("javascript"):
        return "", "JavaScript link"
    if re.search(r"['\",]", href):
        clean = re.split(r"['\",]", href)[0].strip()
        if clean and clean.startswith("/"):
            return base + clean, "PORTAL BUG: broken routerlink. Original: " + href[:60]
        return "", "PORTAL BUG: broken routerlink - " + href[:60]
    if href.startswith("http"):
        return href, ""
    if href.startswith("/"):
        return base + href, ""
    return href, "Relative URL"


# ── STEP 6: Parse menu ────────────────────────────────────────────────────────
def parse_menu(html, cfg):
    print("\n[3] Parsing ALL menu items...")
    soup = BeautifulSoup(html, "html.parser")
    rows = []

    sidebar = soup.find("ul", id="accordionSidebar")
    if not sidebar:
        sidebar = soup.find("ul", class_=lambda c: c and "navbar-nav" in c)

    if not sidebar:
        print("    [!] No sidebar — collecting all anchors")
        for a in soup.find_all("a", href=True):
            txt = a.get_text(strip=True)
            if txt and len(txt) < 80:
                url, issue = clean_url(a.get("href",""), cfg["base"])
                rows.append({"main": txt, "sub": "", "url": url,
                             "issue": issue, "href": a.get("href","")})
        return rows

    for li in sidebar.find_all("li", class_="nav-item"):
        main_a = li.find("a", class_="nav-link")
        if not main_a: continue
        main_name = main_a.get_text(strip=True)
        if not main_name: continue

        collapses = li.find_all("div", class_="collapse-inner")
        if collapses:
            for collapse in collapses:
                sub_links = collapse.find_all("a", class_="dropdown-item")
                if not sub_links:
                    sub_links = collapse.find_all("a", href=True)
                for a in sub_links:
                    sub = a.get_text(strip=True)
                    if not sub: continue
                    href = a.get("href", "")
                    url, issue = clean_url(href, cfg["base"])
                    print("    +", main_name, ">", sub)
                    if issue:
                        print("      ISSUE:", issue)
                    rows.append({"main": main_name, "sub": sub,
                                 "url": url, "issue": issue, "href": href})
        else:
            href = main_a.get("href", "")
            url, issue = clean_url(href, cfg["base"])
            if url:
                rows.append({"main": main_name, "sub": "",
                             "url": url, "issue": issue, "href": href})
                print("    +", main_name, "(direct)")

    print("\n    Total:", len(rows), "items")
    return rows


# ── STEP 7: Save CSV ──────────────────────────────────────────────────────────
def save_csv(rows, cfg):
    with open(cfg["csv"], "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Main Menu", "Sub Menu", "URL", "Original Href", "Issue"])
        for r in rows:
            w.writerow([r["main"], r["sub"], r["url"], r["href"], r["issue"]])
    print("\n[4] CSV saved:", cfg["csv"])


# ── STEP 8: Extract components ────────────────────────────────────────────────
def extract_components(driver, soup):
    comps = {
        "forms": [], "buttons": [], "inputs": [],
        "tables": [], "dropdowns": [], "labels": [],
        "headings": [], "page_sections": [],
    }

    for tag in ["h1","h2","h3","h4","h5"]:
        for h in soup.find_all(tag):
            txt = h.get_text(strip=True)
            if txt and len(txt) > 2:
                comps["headings"].append(tag.upper() + ": " + txt[:50])

    for card in soup.find_all(class_=re.compile(r"card-header|panel-heading|section-title", re.I)):
        txt = card.get_text(strip=True)
        if txt:
            comps["page_sections"].append(txt[:50])

    for i, form in enumerate(soup.find_all("form"), 1):
        method = form.get("method", "GET").upper()
        action = form.get("action", "")[:40]
        fc     = len(form.find_all(["input","select","textarea"]))
        comps["forms"].append("Form" + str(i) + " [" + method + "] fields:" + str(fc) +
                              (" action:" + action if action else ""))

    seen_lbl = set()
    for lbl in soup.find_all("label"):
        txt = lbl.get_text(strip=True)
        if txt and txt not in seen_lbl and len(txt) > 1:
            comps["labels"].append(txt[:40])
            seen_lbl.add(txt)

    seen_inp = set()
    for inp in soup.find_all("input"):
        itype = inp.get("type","text").lower()
        if itype in ["hidden","submit","button","reset"]: continue
        name  = inp.get("name","")
        id_   = inp.get("id","")
        ph    = inp.get("placeholder","")
        lbl_txt = ""
        if id_:
            lbl = soup.find("label", {"for": id_})
            if lbl: lbl_txt = lbl.get_text(strip=True)
        display = lbl_txt or ph or name or id_ or itype
        k = itype + ":" + display
        if k not in seen_inp and display:
            comps["inputs"].append(itype.upper() + ": " + display[:35])
            seen_inp.add(k)

    for ta in soup.find_all("textarea"):
        name = ta.get("name","") or ta.get("id","") or ta.get("placeholder","")
        if name: comps["inputs"].append("TEXTAREA: " + name[:35])

    for sel in soup.find_all("select"):
        name    = sel.get("name","") or sel.get("id","")
        lbl_txt = ""
        if sel.get("id"):
            lbl = soup.find("label", {"for": sel.get("id")})
            if lbl: lbl_txt = lbl.get_text(strip=True)
        display = lbl_txt or name
        opts    = [o.get_text(strip=True) for o in sel.find_all("option") if o.get_text(strip=True)][:5]
        comps["dropdowns"].append("SELECT [" + display[:25] + "]: " + " | ".join(opts))

    seen_btn = set()
    for btn in soup.find_all(["button","input"]):
        btype = btn.get("type","").lower()
        if btype in ["submit","button","reset"] or btn.name == "button":
            label = (btn.get_text(strip=True) or btn.get("value","") or
                     btn.get("name","") or btype)
            if label and label not in seen_btn and len(label) > 1:
                comps["buttons"].append("[" + btype.upper() + "] " + label[:35])
                seen_btn.add(label)

    for i, tbl in enumerate(soup.find_all("table"), 1):
        hdrs = [th.get_text(strip=True) for th in tbl.find_all("th")][:6]
        rc   = len(tbl.find_all("tr")) - 1
        comps["tables"].append("Table" + str(i) + " [" + str(rc) + " rows] " +
                               ("cols: " + " | ".join(hdrs) if hdrs else ""))

    return comps


# ── STEP 9: Deep page testing ─────────────────────────────────────────────────
def deep_test_page(driver, url, main, sub, cfg, index, test_data):
    """
    Intelligently tests page components using user-provided test data.
    - Fills diary/case number in search forms
    - Fills date ranges for reports
    - Clicks radio buttons and dropdowns
    - Takes screenshot after each action
    """
    results  = []
    label    = (main + " > " + sub) if sub else main
    diary_no = test_data.get("diary_no","")
    case_no  = test_data.get("case_no","")
    from_dt  = test_data.get("from_date","01/01/2026")
    to_dt    = test_data.get("to_date","31/12/2026")

    def screenshot(suffix):
        safe    = re.sub(r"[^\w\-]", "_", label)[:30]
        ss_name = str(index).zfill(3) + "_" + safe + "_" + suffix + ".png"
        driver.save_screenshot(os.path.join(cfg["screenshots"], ss_name))
        return ss_name

    def add(component, action, result, notes, ss=""):
        results.append({
            "component": component,
            "action":    action,
            "result":    result,
            "notes":     notes,
            "screenshot": ss
        })
        print("         [" + result + "] " + component + " — " + notes[:60])

    try:
        soup = BeautifulSoup(driver.page_source, "html.parser")

        # ── 1. Radio buttons — click each, check page reacts ─────────────
        radios = driver.find_elements(By.CSS_SELECTOR, "input[type='radio']")
        if radios:
            try:
                for r in radios[:3]:
                    val = r.get_attribute("value") or ""
                    driver.execute_script("arguments[0].click();", r)
                    time.sleep(0.4)
                ss = screenshot("radio")
                add("Radio Buttons",
                    "Clicked " + str(min(3,len(radios))) + " radio(s)",
                    "PASS",
                    str(len(radios)) + " radio(s) found. Values: " +
                    ", ".join([r.get_attribute("value") or "?" for r in radios[:3]]),
                    ss)
            except Exception as e:
                add("Radio Buttons", "Click", "FAIL", str(e)[:50])

        # ── 2. Search with Diary No or Case No ───────────────────────────
        search_val = diary_no or case_no
        if search_val:
            try:
                # Look for radio first (Diary No / Case No option)
                for radio in driver.find_elements(By.CSS_SELECTOR, "input[type='radio']"):
                    val = (radio.get_attribute("value") or "").lower()
                    lbl = ""
                    try:
                        pid = radio.get_attribute("id")
                        if pid:
                            lbl_el = driver.find_element(
                                By.CSS_SELECTOR, "label[for='" + pid + "']")
                            lbl = lbl_el.text.lower()
                    except: pass
                    if diary_no and ("diary" in val or "diary" in lbl):
                        driver.execute_script("arguments[0].click();", radio)
                        time.sleep(0.3)
                        break
                    if case_no and ("case" in val or "case" in lbl):
                        driver.execute_script("arguments[0].click();", radio)
                        time.sleep(0.3)
                        break

                # Fill search input
                search_filled = False
                for selector in [
                    "input[type='text']:not([placeholder*='Search']):not([placeholder*='search'])",
                    "input[name*='diary']", "input[name*='case']",
                    "input[name*='no']",    "input[placeholder*='No']",
                ]:
                    els = driver.find_elements(By.CSS_SELECTOR, selector)
                    for el in els:
                        try:
                            if el.is_displayed() and el.is_enabled():
                                el.clear()
                                el.send_keys(search_val)
                                search_filled = True
                                break
                        except: pass
                    if search_filled: break

                if search_filled:
                    ss = screenshot("search_filled")
                    add("Search Form", "Filled: " + search_val,
                        "PASS", "Search field filled with test data", ss)

                    # Click search/submit button
                    for btn_text in ["Search", "Get", "Find", "Submit", "Go"]:
                        try:
                            btn = driver.find_element(
                                By.XPATH,
                                "//button[contains(translate(text(),'abcdefghijklmnopqrstuvwxyz',"
                                "'ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'" + btn_text.upper() + "')] | "
                                "//input[@type='submit'][contains(translate(@value,'abcdefghijklmnopqrstuvwxyz',"
                                "'ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'" + btn_text.upper() + "')]"
                            )
                            if btn.is_displayed():
                                btn.click()
                                time.sleep(2.5)
                                ss = screenshot("search_result")
                                new_soup = BeautifulSoup(driver.page_source, "html.parser")
                                body     = new_soup.get_text(strip=True).lower()
                                tbls     = new_soup.find_all("table")
                                err_kw   = ["invalid","error","not found","required","no record"]
                                errs     = [k for k in err_kw if k in body]
                                add("Search Submit",
                                    "Clicked: " + btn_text,
                                    "FAIL" if errs else "PASS",
                                    ("Errors: " + ", ".join(errs)) if errs
                                    else (str(len(tbls)) + " table(s) returned"),
                                    ss)
                                # Go back for next test
                                driver.back()
                                time.sleep(1.5)
                                break
                        except: pass
                else:
                    add("Search Form", "Fill", "SKIP",
                        "No suitable input found for diary/case no")
            except Exception as e:
                add("Search Form", "Fill+Submit", "ERROR", str(e)[:60])

        # ── 3. Date range forms (Reports) ────────────────────────────────
        if from_dt and to_dt:
            date_inputs = driver.find_elements(
                By.CSS_SELECTOR,
                "input[placeholder*='Date'], input[name*='date'], "
                "input[name*='Date'], input[name*='from'], input[name*='to']"
            )
            if len(date_inputs) >= 2:
                try:
                    driver.execute_script(
                        "arguments[0].value=arguments[1];", date_inputs[0], from_dt)
                    driver.execute_script(
                        "arguments[0].value=arguments[1];", date_inputs[1], to_dt)
                    # Trigger Angular change detection
                    driver.execute_script(
                        "arguments[0].dispatchEvent(new Event('input'));"
                        "arguments[0].dispatchEvent(new Event('change'));",
                        date_inputs[0])
                    driver.execute_script(
                        "arguments[0].dispatchEvent(new Event('input'));"
                        "arguments[0].dispatchEvent(new Event('change'));",
                        date_inputs[1])
                    time.sleep(0.5)
                    ss = screenshot("dates_filled")
                    add("Date Range", "Filled " + from_dt + " to " + to_dt,
                        "PASS", "Date fields filled", ss)

                    # Click generate/submit
                    for btn_text in ["Generate Report","Generate","Submit","View","Search"]:
                        try:
                            btn = driver.find_element(
                                By.XPATH,
                                "//button[contains(text(),'" + btn_text + "')] | "
                                "//input[@type='submit'][@value='" + btn_text + "']"
                            )
                            if btn.is_displayed():
                                btn.click()
                                time.sleep(3)
                                ss = screenshot("report_result")
                                new_soup = BeautifulSoup(driver.page_source, "html.parser")
                                body     = new_soup.get_text(strip=True).lower()
                                tbls     = new_soup.find_all("table")
                                err_kw   = ["invalid","error","required","no record","please"]
                                errs     = [k for k in err_kw if k in body]
                                add("Report Generate",
                                    "Clicked: " + btn_text,
                                    "FAIL" if errs else "PASS",
                                    ("Errors: " + ", ".join(errs)) if errs
                                    else (str(len(tbls)) + " result table(s) returned"),
                                    ss)
                                driver.back()
                                time.sleep(1.5)
                                break
                        except: pass
                except Exception as e:
                    add("Date Range", "Fill+Submit", "ERROR", str(e)[:60])

        # ── 4. Select dropdowns ───────────────────────────────────────────
        selects = driver.find_elements(By.CSS_SELECTOR, "select")
        for sel in selects[:3]:
            try:
                s    = SeleniumSelect(sel)
                name = sel.get_attribute("name") or sel.get_attribute("id") or "dropdown"
                opts = [o.text.strip() for o in s.options if o.text.strip()]
                if len(opts) > 1:
                    s.select_by_index(1)
                    time.sleep(0.3)
                    add("Dropdown [" + name[:20] + "]",
                        "Selected: " + opts[1][:30],
                        "PASS",
                        str(len(opts)) + " options available")
            except Exception as e:
                add("Dropdown", "Select", "FAIL", str(e)[:50])

        # ── 5. Table data check ───────────────────────────────────────────
        tables = driver.find_elements(By.CSS_SELECTOR, "table")
        for idx, tbl in enumerate(tables[:2], 1):
            try:
                rows_el  = tbl.find_elements(By.TAG_NAME, "tr")
                data_rows = max(0, len(rows_el) - 1)
                add("Table " + str(idx),
                    "Row count",
                    "PASS" if data_rows > 0 else "WARN",
                    str(data_rows) + " data row(s) visible on screen")
            except: pass

    except Exception as e:
        add("Page", "Deep test", "ERROR", str(e)[:80])

    return results


# ── STEP 10: Test every URL ───────────────────────────────────────────────────
def test_urls(driver, rows, cfg, test_data):
    print("\n[5] Testing", len(rows), "URLs...")
    results = []

    for i, row in enumerate(rows):
        main  = row["main"]
        sub   = row["sub"]
        url   = row["url"]
        issue = row["issue"]
        href  = row["href"]
        label = (main + " > " + sub) if sub else main

        print("\n  [" + str(i+1).rjust(3) + "/" + str(len(rows)) + "] " + label)
        print("       URL:", (url or href)[:70])

        if not url:
            print("       STATUS: PORTAL BUG")
            results.append({
                "main": main, "sub": sub, "url": href,
                "status": "PORTAL BUG",
                "issue": issue if issue else "Broken URL: " + href,
                "load": "", "title": "Could not resolve URL",
                "headings": "", "sections": "", "forms": "",
                "labels": "", "inputs": "", "buttons": "",
                "tables": "", "dropdowns": "",
                "deep_results": "", "deep_notes": "",
                "screenshot": ""
            })
            continue

        try:
            t0 = time.time()
            driver.get(url)
            try:
                WebDriverWait(driver, 8).until(
                    lambda d: d.execute_script(
                        "return document.readyState") == "complete")
            except: pass
            time.sleep(2)

            load  = round(time.time() - t0, 2)
            title = driver.title[:55]

            # Screenshot before interaction
            safe    = re.sub(r"[^\w\-]", "_", label)[:45]
            ss_name = str(i+1).zfill(3) + "_" + safe + ".png"
            driver.save_screenshot(os.path.join(cfg["screenshots"], ss_name))

            soup_p = BeautifulSoup(driver.page_source, "html.parser")
            body   = soup_p.get_text(separator=" ", strip=True).lower()
            comps  = extract_components(driver, soup_p)

            err_kw = ["access denied","not found","unauthorized","404","403","500"]
            errors = [k for k in err_kw if k in body]
            if errors:
                status = "FAIL"
            elif "PORTAL BUG" in issue:
                status = "PASS*"
            else:
                status = "PASS"

            # Deep test (only if not skipped)
            deep         = []
            deep_summary = ""
            deep_notes   = ""
            if not test_data.get("skip_deep", False):
                print("       Running deep component tests...")
                deep = deep_test_page(
                    driver, url, main, sub, cfg, i+1, test_data)
                deep_summary = " | ".join(
                    r["component"] + ":" + r["result"] for r in deep)
                deep_notes = " | ".join(
                    r["component"] + "→" + r["notes"] for r in deep)
                # Navigate back to original page
                try:
                    driver.get(url)
                    time.sleep(1.5)
                except: pass

            # Print summary
            if comps["headings"]:
                print("       Headings:", " | ".join(comps["headings"][:2]))
            if comps["inputs"]:
                print("       Fields  :", " | ".join(comps["inputs"][:3]))
            if comps["buttons"]:
                print("       Buttons :", " | ".join(comps["buttons"][:3]))
            if deep_summary:
                print("       DeepTest:", deep_summary[:80])
            print("       STATUS  :", status, "(" + str(load) + "s)")

            results.append({
                "main":         main,
                "sub":          sub,
                "url":          url,
                "status":       status,
                "issue":        issue,
                "load":         load,
                "title":        title,
                "headings":     " | ".join(comps["headings"])[:100],
                "sections":     " | ".join(comps["page_sections"])[:80],
                "forms":        " | ".join(comps["forms"])[:100],
                "labels":       " | ".join(comps["labels"])[:100],
                "inputs":       " | ".join(comps["inputs"])[:120],
                "buttons":      " | ".join(comps["buttons"])[:100],
                "tables":       " | ".join(comps["tables"])[:100],
                "dropdowns":    " | ".join(comps["dropdowns"])[:100],
                "deep_results": deep_summary[:150],
                "deep_notes":   deep_notes[:250],
                "screenshot":   ss_name,
            })

        except Exception as e:
            print("       STATUS: ERROR -", str(e)[:60])
            results.append({
                "main": main, "sub": sub, "url": url,
                "status": "ERROR", "issue": str(e)[:80],
                "load": "", "title": "",
                "headings": "", "sections": "", "forms": "",
                "labels": "", "inputs": "", "buttons": "",
                "tables": "", "dropdowns": "",
                "deep_results": "", "deep_notes": "",
                "screenshot": ""
            })

    return results


# ── STEP 11: Excel report ─────────────────────────────────────────────────────
def write_excel(rows, results, cfg, test_data):
    print("\n[6] Writing Excel report...")

    def fill(c):  return PatternFill("solid", fgColor=c)
    def fnt(bold=False, color="FF000000", sz=10):
        return Font(bold=bold, color=color, size=sz)
    def brd():
        s = Side(style="thin", color="FFD0D0D0")
        return Border(left=s, right=s, top=s, bottom=s)
    def aln(h="left"):
        return Alignment(horizontal=h, vertical="center", wrap_text=True)

    wb = Workbook()

    # ── Sheet 1: Menu Structure ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Menu Structure"
    for ci, (h, w) in enumerate(zip(
        ["#", "Level 1 (Main Menu)", "Level 2 (Sub Menu)", "URL", "Issue"],
        [5, 28, 35, 65, 45]
    ), 1):
        c = ws1.cell(1, ci, h)
        c.fill = fill("FF003366"); c.font = fnt(True,"FFFFFFFF",11)
        c.alignment = aln("center"); c.border = brd()
        ws1.column_dimensions[c.column_letter].width = w

    for ri, row in enumerate(rows, 2):
        is_sub = bool(row["sub"])
        bg = ("FFFFF3CD" if row["issue"] else
              "FFE8F4FD" if not is_sub else
              "FFFAFAFA" if ri%2==0 else "FFF5F5F5")
        for ci, val in enumerate([ri-1, row["main"], row["sub"],
                                   row["url"] or row["href"], row["issue"]], 1):
            c = ws1.cell(ri, ci, val)
            c.fill = fill(bg); c.border = brd(); c.alignment = aln()
            if ci == 4:
                c.font = Font(color="FF0055AA", size=10, underline="single")
            elif ci == 5 and val:
                c.font = fnt(color="FFB71C1C", bold=True)
            else:
                c.font = fnt(bold=not is_sub and ci==2)
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = "A1:E" + str(len(rows)+1)

    # ── Sheet 2: Test Results ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Test Results")
    s2_cols = ["#","Main Menu","Sub Menu","URL","Result","Issue",
               "Load(s)","Page Title","Page Headings","Page Sections",
               "Forms","Field Labels","Input Fields","Buttons",
               "Tables","Dropdowns",
               "Component Test Results","Component Notes",
               "Screenshot"]
    s2_wids = [4,18,22,45,10,35,7,30,40,30,
               25,40,45,35,40,35,
               50,70,
               30]
    for ci, (h, w) in enumerate(zip(s2_cols, s2_wids), 1):
        c = ws2.cell(1, ci, h)
        c.fill = fill("FF005B96"); c.font = fnt(True,"FFFFFFFF",11)
        c.alignment = aln("center"); c.border = brd()
        ws2.column_dimensions[c.column_letter].width = w

    pass_c = fail_c = err_c = bug_c = 0
    for ri, r in enumerate(results, 2):
        st = r["status"]
        if   st=="PASS":       sf=fill("FFE8F5E9"); ff=fnt(True,"FF1B5E20")
        elif st=="PASS*":      sf=fill("FFF1F8E9"); ff=fnt(True,"FF33691E")
        elif st=="FAIL":       sf=fill("FFFFF3E0"); ff=fnt(True,"FFE65100")
        elif st=="PORTAL BUG": sf=fill("FFFCE4EC"); ff=fnt(True,"FFC62828")
        else:                  sf=fill("FFFFEBEE"); ff=fnt(True,"FFB71C1C")
        if   st in ["PASS","PASS*"]: pass_c+=1
        elif st=="FAIL":             fail_c+=1
        elif st=="PORTAL BUG":       bug_c+=1
        else:                        err_c+=1

        bg = "FFFAFAFA" if ri%2==0 else "FFFFFFFF"
        vals = [
            ri-1, r["main"], r["sub"],
            r["url"] or r.get("href",""),
            st, r["issue"], r["load"], r["title"],
            r.get("headings",""), r.get("sections",""),
            r["forms"], r.get("labels",""), r["inputs"],
            r["buttons"], r["tables"], r["dropdowns"],
            r.get("deep_results",""), r.get("deep_notes",""),
            r["screenshot"]
        ]
        for ci, val in enumerate(vals, 1):
            c = ws2.cell(ri, ci, val)
            c.fill = sf if ci==5 else fill(bg)
            c.font = (ff if ci==5 else
                      fnt(color="FFB71C1C",bold=True) if ci==6 and val
                      else fnt())
            c.alignment = aln(); c.border = brd()
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = "A1:S" + str(len(results)+1)

    # ── Sheet 3: Summary ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 35

    summary = [
        ("Report Generated",  datetime.now().strftime("%Y-%m-%d %H:%M:%S"), None),
        ("Portal URL",        cfg["url"],       None),
        ("Login User",        cfg["username"],  None),
        ("Output Folder",     cfg["folder"],    None),
        ("", "", None),
        ("── Test Data Used ──", "", None),
        ("Diary Number",      test_data.get("diary_no","(not provided)"),  None),
        ("Case Number",       test_data.get("case_no","(not provided)"),   None),
        ("Date Range",        test_data.get("from_date","") + " to " +
                              test_data.get("to_date",""),                 None),
        ("Party Name",        test_data.get("party_name","(not provided)"),None),
        ("", "", None),
        ("── Results ──", "", None),
        ("Total Menu Items",  len(rows),     None),
        ("Total URLs Tested", len(results),  None),
        ("PASS",              pass_c,        "FF1B5E20"),
        ("FAIL",              fail_c,        "FFE65100"),
        ("PORTAL BUG",        bug_c,         "FFC62828"),
        ("ERROR",             err_c,         "FFB71C1C"),
    ]
    for ri, (lbl, val, color) in enumerate(summary, 1):
        l = ws3.cell(ri, 1, lbl)
        v = ws3.cell(ri, 2, str(val))
        l.font = fnt(bold=True, sz=11)
        if color:
            v.font = fnt(color=color, sz=12, bold=True)
        else:
            v.font = fnt(sz=11)

    wb.save(cfg["excel"])
    print("    Saved:", os.path.abspath(cfg["excel"]))
    print("\n    PASS:", pass_c, "| FAIL:", fail_c,
          "| PORTAL BUG:", bug_c, "| ERROR:", err_c)


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    # Step 1: Get login credentials
    cfg = get_config()
    if not cfg:
        print("Cancelled."); return

    # Step 2: Get test data
    test_data = get_test_data()

    print("\n" + "="*52)
    print("  User     :", cfg["username"])
    print("  Folder   :", cfg["folder"])
    print("  Diary No :", test_data.get("diary_no","(skip)"))
    print("  Case No  :", test_data.get("case_no","(skip)"))
    print("  Dates    :", test_data.get("from_date",""), "to", test_data.get("to_date",""))
    print("  Deep Test:", "NO" if test_data.get("skip_deep") else "YES")
    print("="*52)

    driver = init_driver()
    try:
        login(driver, cfg)
        html      = get_html(driver, cfg)
        rows      = parse_menu(html, cfg)
        save_csv(rows, cfg)
        results   = test_urls(driver, rows, cfg, test_data)
        write_excel(rows, results, cfg, test_data)

        print("\n" + "="*52)
        print("  DONE!")
        print("  Folder :", cfg["folder"])
        print("  CSV    :", cfg["csv"])
        print("  Excel  :", cfg["excel"])
        print("="*52)

    except Exception as e:
        print("\n[ERROR]", e)
        import traceback; traceback.print_exc()
    finally:
        input("\nPress ENTER to close browser...")
        driver.quit()


if __name__ == "__main__":
    main()