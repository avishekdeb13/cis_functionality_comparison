"""
Universal Portal Testing Tool v2.0
====================================
TWO PORTAL MODE:
  - LIVE portal   → READ ONLY (extract menu structure only, zero interaction)
  - TEST portal   → FULL testing (fill forms, click buttons, screenshots)

Workflow:
  Step 1 → GUI: Enter LIVE portal credentials (read-only reference)
  Step 2 → GUI: Enter TEST portal credentials (full testing)
  Step 3 → GUI: Enter test data (diary no, case no, dates)
  Step 4 → Login to LIVE portal, extract menu structure only
  Step 5 → Login to TEST portal, compare menus, run full deep testing
  Step 6 → Excel report with 5 sheets:
            - Live Menu Structure
            - Test Menu Structure
            - Menu Comparison (missing/extra items)
            - Test Results (with field details)
            - Summary
"""

from selenium import webdriver
# from selenium.webdriver.by import By
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select as SeleniumSelect
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, WebDriverException)
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import tkinter as tk
from tkinter import ttk, messagebox
import time, os, csv, re
from datetime import datetime
from urllib.parse import urlparse, urljoin


# ─────────────────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def fill(c):   return PatternFill("solid", fgColor=c)
def fnt(bold=False, color="FF000000", sz=10):
    return Font(bold=bold, color=color, size=sz)
def brd():
    s = Side(style="thin", color="FFD0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)
def aln(h="left"):
    return Alignment(horizontal=h, vertical="center", wrap_text=True)


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1: GUI — Three screens
# ─────────────────────────────────────────────────────────────────────────────
def get_all_config():
    """Single GUI with 3 tabs: Live Portal, Test Portal, Test Data"""
    config = {"live": {}, "test": {}, "data": {}}

    def submit():
        # Validate live portal
        live_url  = live_url_var.get().strip()
        live_user = live_user_var.get().strip()
        live_pwd  = live_pwd_var.get().strip()
        # Validate test portal
        test_url  = test_url_var.get().strip()
        test_user = test_user_var.get().strip()
        test_pwd  = test_pwd_var.get().strip()

        if not all([live_url, live_user, live_pwd, test_url, test_user, test_pwd]):
            messagebox.showerror("Error", "All URL, username and password fields are required")
            return

        ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
        folder = test_user + "_" + ts
        ss_dir = os.path.join(folder, "screenshots")
        os.makedirs(ss_dir, exist_ok=True)

        config["live"] = {
            "url":      live_url,
            "username": live_user,
            "password": live_pwd,
            "base":     urlparse(live_url).scheme + "://" + urlparse(live_url).netloc,
            "label":    "LIVE (Read-Only)",
        }
        config["test"] = {
            "url":      test_url,
            "username": test_user,
            "password": test_pwd,
            "base":     urlparse(test_url).scheme + "://" + urlparse(test_url).netloc,
            "folder":      folder,
            "screenshots": ss_dir,
            "html":        os.path.join(folder, test_user + "_portal.html"),
            "live_html":   os.path.join(folder, "live_portal.html"),
            "csv":         os.path.join(folder, test_user + "_menu.csv"),
            "excel":       os.path.join(folder, test_user + "_test_report.xlsx"),
            "label":    "TEST (Full Testing)",
        }
        config["data"] = {
            "diary_no":   diary_var.get().strip(),
            "case_no":    case_var.get().strip(),
            "case_year":  year_var.get().strip(),
            "from_date":  from_var.get().strip(),
            "to_date":    to_var.get().strip(),
            "party_name": party_var.get().strip(),
            "skip_deep":  skip_var.get(),
        }
        root.destroy()

    def skip_all():
        for k in ["diary_no","case_no","case_year","from_date","to_date","party_name"]:
            config["data"][k] = ""
        config["data"]["skip_deep"] = True
        # Still need portal credentials
        submit()

    root = tk.Tk()
    root.title("Portal Testing Tool v2.0")
    root.geometry("640x580")
    root.resizable(False, False)

    tk.Label(root, text="Universal Portal Testing Tool v2.0",
             font=("Arial", 14, "bold"), fg="#003366").pack(pady=8)
    tk.Label(root,
             text="LIVE portal = read-only reference only  |  TEST portal = full testing",
             font=("Arial", 9), fg="#cc0000").pack()

    nb = ttk.Notebook(root)
    nb.pack(fill="both", expand=True, padx=10, pady=8)

    # ── Tab 1: Live Portal (READ ONLY) ───────────────────────────────────────
    tab1 = ttk.Frame(nb, padding=15)
    nb.add(tab1, text="  🔴  LIVE Portal (Read-Only)  ")

    tk.Label(tab1, text="LIVE Portal — Menu structure will be extracted ONLY.",
             font=("Arial", 10, "bold"), fg="#cc0000").grid(
             row=0, column=0, columnspan=2, pady=(0,10), sticky="w")
    tk.Label(tab1, text="No forms will be filled. No buttons clicked. No data submitted.",
             font=("Arial", 9), fg="#888").grid(
             row=1, column=0, columnspan=2, pady=(0,15), sticky="w")

    live_url_var  = tk.StringVar(value="https://cis.drt.gov.in/drtlive/login.php")
    live_user_var = tk.StringVar(value="")
    live_pwd_var  = tk.StringVar(value="")

    for ri, (lbl, var, show) in enumerate([
        ("Live Portal URL:", live_url_var,  ""),
        ("Username:",        live_user_var, ""),
        ("Password:",        live_pwd_var,  "*"),
    ], 2):
        ttk.Label(tab1, text=lbl).grid(row=ri, column=0, sticky="w", pady=8)
        ttk.Entry(tab1, textvariable=var, show=show, width=52).grid(
            row=ri, column=1, padx=8)

    tk.Label(tab1,
             text="⚠  This portal is LIVE. The script will NEVER submit or modify anything.",
             font=("Arial", 9, "bold"), fg="#cc0000").grid(
             row=5, column=0, columnspan=2, pady=(20,0), sticky="w")

    # ── Tab 2: Test Portal (FULL TESTING) ───────────────────────────────────
    tab2 = ttk.Frame(nb, padding=15)
    nb.add(tab2, text="  🟢  TEST Portal (Full Testing)  ")

    tk.Label(tab2, text="TEST Portal — Full testing will be performed here.",
             font=("Arial", 10, "bold"), fg="#006600").grid(
             row=0, column=0, columnspan=2, pady=(0,10), sticky="w")
    tk.Label(tab2, text="Forms filled, buttons clicked, screenshots taken.",
             font=("Arial", 9), fg="#888").grid(
             row=1, column=0, columnspan=2, pady=(0,15), sticky="w")

    test_url_var  = tk.StringVar(
        value="https://drt.etribunals.gov.in/cis2.0/filing/login")
    test_user_var = tk.StringVar(value="")
    test_pwd_var  = tk.StringVar(value="")

    for ri, (lbl, var, show) in enumerate([
        ("Test Portal URL:", test_url_var,  ""),
        ("Username:",        test_user_var, ""),
        ("Password:",        test_pwd_var,  "*"),
    ], 2):
        ttk.Label(tab2, text=lbl).grid(row=ri, column=0, sticky="w", pady=8)
        ttk.Entry(tab2, textvariable=var, show=show, width=52).grid(
            row=ri, column=1, padx=8)

    tk.Label(tab2,
             text="✅  Safe to test — this is the testing/staging environment.",
             font=("Arial", 9, "bold"), fg="#006600").grid(
             row=5, column=0, columnspan=2, pady=(20,0), sticky="w")

    # ── Tab 3: Test Data ──────────────────────────────────────────────────────
    tab3 = ttk.Frame(nb, padding=15)
    nb.add(tab3, text="  📋  Test Data  ")

    tk.Label(tab3,
             text="Test data used to fill forms on TEST portal only.\n"
                  "Live portal is NEVER touched with this data.",
             font=("Arial", 9), fg="#555").grid(
             row=0, column=0, columnspan=3, pady=(0,12), sticky="w")

    diary_var = tk.StringVar()
    case_var  = tk.StringVar()
    year_var  = tk.StringVar(value=str(datetime.now().year))
    from_var  = tk.StringVar(value="01/01/2026")
    to_var    = tk.StringVar(value="31/12/2026")
    party_var = tk.StringVar()
    skip_var  = tk.BooleanVar(value=False)

    fields = [
        ("Diary Number:",  diary_var,  "e.g. 100/2026"),
        ("Case Number:",   case_var,   "e.g. OA/5/2026"),
        ("Case Year:",     year_var,   "e.g. 2026"),
        ("From Date:",     from_var,   "DD/MM/YYYY"),
        ("To Date:",       to_var,     "DD/MM/YYYY"),
        ("Party Name:",    party_var,  "e.g. BANK OF INDIA"),
    ]
    for ri, (lbl, var, hint) in enumerate(fields, 1):
        ttk.Label(tab3, text=lbl).grid(row=ri, column=0, sticky="w", pady=5)
        ttk.Entry(tab3, textvariable=var, width=28).grid(
            row=ri, column=1, padx=8, sticky="w")
        ttk.Label(tab3, text=hint, foreground="#999",
                  font=("Arial", 8)).grid(row=ri, column=2, sticky="w")

    ttk.Checkbutton(tab3,
                    text="Skip form submission (extract components only, no form filling)",
                    variable=skip_var).grid(
                    row=len(fields)+1, column=0, columnspan=3, pady=10, sticky="w")

    # ── Bottom buttons ────────────────────────────────────────────────────────
    btn_frame = ttk.Frame(root)
    btn_frame.pack(pady=10)
    ttk.Button(btn_frame, text="Skip Form Tests",
               command=skip_all).pack(side="left", padx=8)
    ttk.Button(btn_frame, text="▶  Start Testing",
               command=submit).pack(side="left", padx=8)

    root.mainloop()
    return config


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2: Browser
# ─────────────────────────────────────────────────────────────────────────────
def init_driver():
    opts = webdriver.ChromeOptions()
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=opts)


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3: Login (used for both portals)
# ─────────────────────────────────────────────────────────────────────────────
def login(driver, portal_cfg, screenshots_dir=None, label=""):
    print("\n[LOGIN] Opening:", portal_cfg["label"])
    print("        URL:", portal_cfg["url"])
    driver.get(portal_cfg["url"])
    time.sleep(3)

    def fill_field(fname, value, locators):
        for loc in locators:
            try:
                el = driver.find_element(*loc)
                el.clear()
                el.send_keys(value)
                print("    [OK]", fname, "filled")
                return True
            except NoSuchElementException:
                pass
        print("    [!]", fname, "not found — fill manually")
        return False

    fill_field("Username", portal_cfg["username"], [
        (By.ID,   "user_name"),   (By.ID,   "username"),
        (By.NAME, "user_name"),   (By.NAME, "username"),
        (By.NAME, "txtUsername"), (By.ID,   "txtUsername"),
        (By.XPATH, "//input[@type='text'][1]"),
        (By.XPATH, "//input[contains(@name,'user') or contains(@id,'user')]"),
    ])
    fill_field("Password", portal_cfg["password"], [
        (By.ID,   "user_pass"),   (By.ID,   "password"),
        (By.NAME, "user_pass"),   (By.NAME, "password"),
        (By.NAME, "txtPassword"), (By.ID,   "txtPassword"),
        (By.XPATH, "//input[@type='password']"),
    ])

    if screenshots_dir:
        driver.save_screenshot(
            os.path.join(screenshots_dir, label + "_00_login.png"))

    print("\n" + "="*54)
    print("  Portal:", portal_cfg["label"])
    print("  ACTION: Solve CAPTCHA then click LOGIN")
    print("  Waiting up to 120 seconds...")
    print("="*54 + "\n")

    start_url = driver.current_url
    try:
        WebDriverWait(driver, 120).until(lambda d: d.current_url != start_url)
    except TimeoutException:
        print("  [!] URL did not change — continuing anyway...")

    time.sleep(3)
    if screenshots_dir:
        driver.save_screenshot(
            os.path.join(screenshots_dir, label + "_01_dashboard.png"))
    print("[OK] Logged in:", driver.current_url)


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4: Extract menu — READ ONLY (safe for live portal)
# ─────────────────────────────────────────────────────────────────────────────
def get_menu_readonly(driver, portal_cfg, html_path=None):
    """
    SAFE for live portal:
    - Only reads the page HTML
    - Does NOT click, scroll, fill, or interact with anything
    - Just expands menus via CSS class injection (no form submission)
    """
    print("\n[MENU] Extracting menu from:", portal_cfg["label"])
    print("       READ-ONLY — no clicks, no form fills, no data submission")

    # Only JS that adds CSS classes — safe, non-destructive
    driver.execute_script("""
        document.querySelectorAll('.collapse').forEach(el =>
            el.classList.add('show'));
        document.querySelectorAll('[aria-expanded]').forEach(el =>
            el.setAttribute('aria-expanded', 'true'));
        document.querySelectorAll('.dropdown-menu').forEach(el => {
            el.classList.add('show');
            el.style.display = 'block';
        });
    """)
    time.sleep(2)

    html = driver.page_source
    if html_path:
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html)
        print("       HTML saved:", html_path)

    rows = parse_menu_from_html(html, portal_cfg["base"])
    print("       Menu items found:", len(rows))
    for r in rows:
        label = "  📁 " + r["main"] if not r["sub"] else "     └─ " + r["sub"]
        print(label)
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5: URL cleaner
# ─────────────────────────────────────────────────────────────────────────────
def clean_url(href, base):
    if not href or not href.strip():
        return "", "Empty href"
    href = href.strip()
    if href.lower().startswith("javascript") or href in ("#",""):
        return "", "JavaScript link"
    if re.search(r"['\",]", href):
        clean = re.split(r"['\",]", href)[0].strip()
        if clean and clean.startswith("/"):
            return base + clean, "PORTAL BUG: broken routerlink. Original: " + href[:60]
        return "", "PORTAL BUG: broken routerlink - " + href[:60]
    if href.startswith("http"):
        return href, ""
    if href.startswith("/"):
        return base + href, ""
    return urljoin(base + "/", href), "Relative URL"


# ─────────────────────────────────────────────────────────────────────────────
# STEP 6: Parse menu from HTML (works for both old and new portal)
# ─────────────────────────────────────────────────────────────────────────────
def parse_menu_from_html(html, base):
    soup = BeautifulSoup(html, "html.parser")
    rows = []

    # Strategy 1: New portal — accordionSidebar
    sidebar = soup.find("ul", id="accordionSidebar")
    if sidebar:
        print("       Nav type: NEW portal (accordionSidebar)")
        for li in sidebar.find_all("li", class_="nav-item"):
            main_a = li.find("a", class_="nav-link")
            if not main_a: continue
            main_name = main_a.get_text(strip=True)
            if not main_name: continue
            collapses = li.find_all("div", class_="collapse-inner")
            if collapses:
                for collapse in collapses:
                    sub_links = collapse.find_all("a", class_="dropdown-item")
                    if not sub_links:
                        sub_links = collapse.find_all("a", href=True)
                    for a in sub_links:
                        sub = a.get_text(strip=True)
                        if not sub: continue
                        href = a.get("href","")
                        url, issue = clean_url(href, base)
                        rows.append({"main": main_name, "sub": sub,
                                     "url": url, "issue": issue, "href": href})
            else:
                href = main_a.get("href","")
                url, issue = clean_url(href, base)
                if url:
                    rows.append({"main": main_name, "sub": "",
                                 "url": url, "issue": issue, "href": href})
        if rows:
            return rows

    # Strategy 2: Old portal — navbar-nav + dropdown-menu
    nav    = soup.find("nav", class_=re.compile(r"navbar", re.I))
    nav_ul = None
    if nav:
        nav_ul = nav.find("ul", class_=re.compile(r"navbar-nav", re.I))
    if not nav_ul:
        for ul in soup.find_all("ul"):
            if len(ul.find_all("li", recursive=False)) >= 3:
                nav_ul = ul; break

    if nav_ul:
        print("       Nav type: OLD portal (navbar-nav + dropdown-menu)")
        for li in nav_ul.find_all("li", recursive=False):
            main_a = li.find("a", class_=re.compile(r"nav-link", re.I))
            if not main_a:
                main_a = li.find("a")
            if not main_a: continue
            main_name = main_a.get_text(strip=True).lstrip("*").strip()
            if not main_name: continue
            main_href = main_a.get("href","")
            main_url, main_issue = clean_url(main_href, base)

            dropdown = li.find("div", class_=re.compile(r"dropdown-menu", re.I))
            if not dropdown:
                dropdown = li.find("ul", class_=re.compile(r"dropdown-menu|sub-menu", re.I))

            if dropdown:
                sub_links = dropdown.find_all("a", class_=re.compile(r"dropdown-item", re.I))
                if not sub_links:
                    sub_links = dropdown.find_all("a", href=True)
                added = 0
                for a in sub_links:
                    sub = a.get_text(strip=True)
                    if not sub: continue
                    href = a.get("href","")
                    url, issue = clean_url(href, base)
                    rows.append({"main": main_name, "sub": sub,
                                 "url": url, "issue": issue, "href": href})
                    added += 1
                if not added:
                    rows.append({"main": main_name, "sub": "",
                                 "url": main_url, "issue": main_issue,
                                 "href": main_href})
            else:
                rows.append({"main": main_name, "sub": "",
                             "url": main_url, "issue": main_issue,
                             "href": main_href})
        if rows:
            return rows

    # Fallback
    print("       Nav type: FALLBACK (all anchors)")
    for a in soup.find_all("a", href=True):
        txt = a.get_text(strip=True)
        if txt and len(txt) < 80:
            href = a.get("href","")
            url, issue = clean_url(href, base)
            if url:
                rows.append({"main": txt, "sub": "", "url": url,
                             "issue": issue, "href": href})
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# STEP 7: Compare menus
# ─────────────────────────────────────────────────────────────────────────────
def compare_menus(live_rows, test_rows):
    """Compare live vs test portal menus and return differences"""
    def key(r):
        return (r["main"].strip().lower(), r["sub"].strip().lower())

    live_keys = {key(r) for r in live_rows}
    test_keys = {key(r) for r in test_rows}

    missing_in_test  = live_keys - test_keys  # in live but not in test
    extra_in_test    = test_keys - live_keys  # in test but not in live
    common           = live_keys & test_keys

    comparison = []
    for r in live_rows:
        k = key(r)
        comparison.append({
            "main":   r["main"],
            "sub":    r["sub"],
            "status": "MATCH" if k in test_keys else "MISSING IN TEST",
            "live_url": r["url"],
            "test_url": next((t["url"] for t in test_rows if key(t)==k), ""),
        })
    for r in test_rows:
        k = key(r)
        if k not in live_keys:
            comparison.append({
                "main":   r["main"],
                "sub":    r["sub"],
                "status": "EXTRA IN TEST",
                "live_url": "",
                "test_url": r["url"],
            })

    print("\n[COMPARE] Menu comparison:")
    print("          Live items    :", len(live_rows))
    print("          Test items    :", len(test_rows))
    print("          Matching      :", len(common))
    print("          Missing in test:", len(missing_in_test))
    print("          Extra in test  :", len(extra_in_test))
    return comparison


# ─────────────────────────────────────────────────────────────────────────────
# STEP 8: Save CSV
# ─────────────────────────────────────────────────────────────────────────────
def save_csv(rows, cfg):
    with open(cfg["test"]["csv"], "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Main Menu","Sub Menu","URL","Original Href","Issue"])
        for r in rows:
            w.writerow([r["main"],r["sub"],r["url"],r["href"],r["issue"]])
    print("\n[CSV] Saved:", cfg["test"]["csv"])


# ─────────────────────────────────────────────────────────────────────────────
# STEP 9: Expand test portal sidebar
# ─────────────────────────────────────────────────────────────────────────────
def expand_test_sidebar(driver, cfg):
    print("\n[EXPAND] Expanding test portal sidebar menus...")

    driver.execute_script("""
        document.querySelectorAll('.collapse').forEach(el => el.classList.add('show'));
        document.querySelectorAll('[aria-expanded]').forEach(el =>
            el.setAttribute('aria-expanded', 'true'));
    """)
    time.sleep(1)

    toggles = driver.find_elements(By.CSS_SELECTOR, "[data-bs-toggle='collapse']")
    for toggle in toggles:
        try:
            driver.execute_script("arguments[0].scrollIntoView(true);", toggle)
            time.sleep(0.1)
            target = (toggle.get_attribute("data-bs-target") or "").replace("#","")
            if target:
                try:
                    el = driver.find_element(By.ID, target)
                    if "show" not in (el.get_attribute("class") or ""):
                        driver.execute_script("arguments[0].click();", toggle)
                        time.sleep(0.3)
                except:
                    driver.execute_script("arguments[0].click();", toggle)
                    time.sleep(0.3)
            print("    [+]", toggle.text.strip()[:40] or target)
        except: pass

    driver.execute_script("""
        document.querySelectorAll('.collapse').forEach(el => el.classList.add('show'));
    """)
    time.sleep(2)

    driver.save_screenshot(
        os.path.join(cfg["test"]["screenshots"], "TEST_02_sidebar.png"))

    html = driver.page_source
    with open(cfg["test"]["html"], "w", encoding="utf-8") as f:
        f.write(html)

    soup  = BeautifulSoup(html, "html.parser")
    items = soup.find_all("a", class_="dropdown-item")
    print("    Captured", len(items), "dropdown items in test portal")
    return html


# ─────────────────────────────────────────────────────────────────────────────
# STEP 10: Extract components
# ─────────────────────────────────────────────────────────────────────────────
def extract_components(driver, soup):
    comps = {
        "forms": [], "buttons": [], "inputs": [],
        "tables": [], "dropdowns": [], "labels": [],
        "headings": [], "page_sections": [],
        "fields_detail": [],
    }

    for tag in ["h1","h2","h3","h4","h5"]:
        for h in soup.find_all(tag):
            txt = h.get_text(strip=True)
            if txt and len(txt) > 2:
                comps["headings"].append(tag.upper() + ": " + txt[:50])

    for card in soup.find_all(class_=re.compile(
            r"card-header|panel-heading|section-title", re.I)):
        txt = card.get_text(strip=True)
        if txt:
            comps["page_sections"].append(txt[:50])

    for form_idx, form in enumerate(soup.find_all("form"), 1):
        method = form.get("method","GET").upper()
        action = form.get("action","")[:40]
        fields_in_form = []

        for inp in form.find_all("input"):
            itype = inp.get("type","text").lower()
            if itype in ["hidden","submit","button","reset"]: continue
            name  = inp.get("name","")
            id_   = inp.get("id","")
            ph    = inp.get("placeholder","").replace("*","").strip()
            req   = (inp.has_attr("required") or "*" in inp.get("placeholder","") or
                     inp.get("ng-required","") == "true")
            lbl_txt = ""
            if id_:
                lbl = soup.find("label", {"for": id_})
                if lbl: lbl_txt = lbl.get_text(strip=True).replace("*","").strip()
            if not lbl_txt:
                parent = inp.parent
                if parent:
                    lbl = parent.find("label")
                    if lbl: lbl_txt = lbl.get_text(strip=True).replace("*","").strip()
            display = lbl_txt or ph or name or id_ or itype
            if not display or display.lower() in ["search for...","search:","search"]:
                continue
            mandatory = "MANDATORY" if req else "OPTIONAL"
            fs = itype.upper() + " | " + display[:40] + " | " + mandatory
            fields_in_form.append(fs)
            comps["inputs"].append(fs)

        for ta in form.find_all("textarea"):
            name = ta.get("name","") or ta.get("id","") or ta.get("placeholder","")
            req  = ta.has_attr("required") or "*" in (ta.get("placeholder",""))
            if name:
                fs = "TEXTAREA | " + name[:40] + " | " + ("MANDATORY" if req else "OPTIONAL")
                fields_in_form.append(fs); comps["inputs"].append(fs)

        for sel in form.find_all("select"):
            name = sel.get("name","") or sel.get("id","")
            req  = sel.has_attr("required")
            lbl_txt = ""
            if sel.get("id"):
                lbl = soup.find("label", {"for": sel.get("id")})
                if lbl: lbl_txt = lbl.get_text(strip=True).replace("*","").strip()
            display = lbl_txt or name
            opts = [o.get_text(strip=True) for o in sel.find_all("option")
                    if o.get_text(strip=True)][:6]
            mandatory = "MANDATORY" if req else "OPTIONAL"
            fs = ("SELECT | " + display[:30] + " | " + mandatory +
                  " | Options: " + " / ".join(opts))
            fields_in_form.append(fs); comps["dropdowns"].append(fs)

        radio_groups = {}
        for radio in form.find_all("input", {"type":"radio"}):
            name = radio.get("name","radio_group")
            val  = radio.get("value","")
            lbl  = ""
            rid  = radio.get("id","")
            if rid:
                lbl_el = soup.find("label", {"for": rid})
                if lbl_el: lbl = lbl_el.get_text(strip=True)
            if name not in radio_groups: radio_groups[name] = []
            radio_groups[name].append(lbl or val)
        for grp_name, options in radio_groups.items():
            fs = "RADIO | " + grp_name[:30] + " | OPTIONS: " + " / ".join(options[:5])
            fields_in_form.append(fs)

        for chk in form.find_all("input", {"type":"checkbox"}):
            name = chk.get("name","") or chk.get("id","")
            lbl  = ""
            cid  = chk.get("id","")
            if cid:
                lbl_el = soup.find("label", {"for": cid})
                if lbl_el: lbl = lbl_el.get_text(strip=True).replace("*","").strip()
            display = lbl or name
            if display:
                fs = "CHECKBOX | " + display[:40] + " | OPTIONAL"
                fields_in_form.append(fs)

        for f_inp in form.find_all("input", {"type":"file"}):
            name = f_inp.get("name","") or f_inp.get("id","") or "file"
            req  = f_inp.has_attr("required")
            fs   = "FILE | " + name[:40] + " | " + ("MANDATORY" if req else "OPTIONAL")
            fields_in_form.append(fs)

        comps["forms"].append(
            "Form" + str(form_idx) + " [" + method + "]" +
            (" action:" + action if action else "") +
            " — " + str(len(fields_in_form)) + " field(s)")
        for f in fields_in_form:
            comps["fields_detail"].append("Form" + str(form_idx) + " › " + f)

    seen_btn = set()
    for btn in soup.find_all(["button","input"]):
        btype = btn.get("type","").lower()
        if btype in ["submit","button","reset"] or btn.name == "button":
            label = (btn.get_text(strip=True) or btn.get("value","") or
                     btn.get("name","") or btype)
            if label and label not in seen_btn and len(label) > 1:
                comps["buttons"].append("[" + btype.upper() + "] " + label[:35])
                seen_btn.add(label)

    for i, tbl in enumerate(soup.find_all("table"), 1):
        hdrs = [th.get_text(strip=True) for th in tbl.find_all("th")][:8]
        rc   = len(tbl.find_all("tr")) - 1
        comps["tables"].append(
            "Table" + str(i) + " [" + str(rc) + " rows] " +
            ("cols: " + " | ".join(hdrs) if hdrs else "no headers"))

    seen_lbl = set()
    for lbl in soup.find_all("label"):
        txt = lbl.get_text(strip=True)
        if txt and txt not in seen_lbl and len(txt) > 1:
            comps["labels"].append(txt[:40]); seen_lbl.add(txt)

    return comps


# ─────────────────────────────────────────────────────────────────────────────
# STEP 11: Deep page testing (TEST PORTAL ONLY)
# ─────────────────────────────────────────────────────────────────────────────
def deep_test_page(driver, url, main, sub, cfg, index, test_data):
    results = []
    label   = (main + " > " + sub) if sub else main
    diary_no = test_data.get("diary_no","")
    case_no  = test_data.get("case_no","")
    from_dt  = test_data.get("from_date","01/01/2026")
    to_dt    = test_data.get("to_date","31/12/2026")

    def screenshot(suffix):
        safe    = re.sub(r"[^\w\-]", "_", label)[:30]
        ss_name = str(index).zfill(3) + "_" + safe + "_" + suffix + ".png"
        driver.save_screenshot(
            os.path.join(cfg["test"]["screenshots"], ss_name))
        return ss_name

    def add(component, action, result, notes, ss=""):
        results.append({"component": component, "action": action,
                        "result": result, "notes": notes, "screenshot": ss})
        print("         [" + result + "] " + component + " — " + notes[:60])

    try:
        # 1. Radio buttons
        radios = driver.find_elements(By.CSS_SELECTOR, "input[type='radio']")
        if radios:
            try:
                for r in radios[:3]:
                    driver.execute_script("arguments[0].click();", r)
                    time.sleep(0.3)
                ss = screenshot("radio")
                add("Radio Buttons",
                    "Clicked " + str(min(3,len(radios))) + " radio(s)",
                    "PASS",
                    str(len(radios)) + " radio(s). Values: " +
                    ", ".join([r.get_attribute("value") or "?" for r in radios[:3]]),
                    ss)
            except Exception as e:
                add("Radio Buttons","Click","FAIL",str(e)[:50])

        # 2. Search with diary/case no
        search_val = diary_no or case_no
        if search_val:
            try:
                for radio in driver.find_elements(By.CSS_SELECTOR,"input[type='radio']"):
                    val = (radio.get_attribute("value") or "").lower()
                    lbl = ""
                    try:
                        pid = radio.get_attribute("id")
                        if pid:
                            lbl_el = driver.find_element(
                                By.CSS_SELECTOR, "label[for='" + pid + "']")
                            lbl = lbl_el.text.lower()
                    except: pass
                    if diary_no and ("diary" in val or "diary" in lbl):
                        driver.execute_script("arguments[0].click();", radio)
                        time.sleep(0.3); break
                    if case_no and ("case" in val or "case" in lbl):
                        driver.execute_script("arguments[0].click();", radio)
                        time.sleep(0.3); break

                search_filled = False
                for selector in [
                    "input[type='text']:not([placeholder*='Search']):not([placeholder*='search'])",
                    "input[name*='diary']","input[name*='case']",
                    "input[name*='no']","input[placeholder*='No']","input[type='text']",
                ]:
                    for el in driver.find_elements(By.CSS_SELECTOR, selector):
                        try:
                            if el.is_displayed() and el.is_enabled():
                                el.clear(); el.send_keys(search_val)
                                search_filled = True; break
                        except: pass
                    if search_filled: break

                if search_filled:
                    ss = screenshot("search_filled")
                    add("Search Form","Filled: " + search_val,
                        "PASS","Search field filled with test data",ss)
                    for btn_text in ["Search","Get","Find","Submit","Go","OK"]:
                        try:
                            btn = driver.find_element(By.XPATH,
                                "//button[contains(translate(text(),"
                                "'abcdefghijklmnopqrstuvwxyz',"
                                "'ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'" + btn_text.upper() + "')] | "
                                "//input[@type='submit'][contains(translate(@value,"
                                "'abcdefghijklmnopqrstuvwxyz',"
                                "'ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'" + btn_text.upper() + "')]")
                            if btn.is_displayed():
                                btn.click(); time.sleep(2.5)
                                ss = screenshot("search_result")
                                new_soup = BeautifulSoup(driver.page_source,"html.parser")
                                body     = new_soup.get_text(strip=True).lower()
                                tbls     = new_soup.find_all("table")
                                errs     = [k for k in ["invalid","error","not found",
                                            "required","no record"] if k in body]
                                add("Search Submit","Clicked: " + btn_text,
                                    "FAIL" if errs else "PASS",
                                    ("Errors: " + ", ".join(errs)) if errs
                                    else str(len(tbls)) + " table(s) returned",ss)
                                driver.back(); time.sleep(1.5); break
                        except: pass
                else:
                    add("Search Form","Fill","SKIP","No suitable input found")
            except Exception as e:
                add("Search Form","Fill+Submit","ERROR",str(e)[:60])

        # 3. Date range forms
        if from_dt and to_dt:
            date_inputs = driver.find_elements(By.CSS_SELECTOR,
                "input[placeholder*='Date'], input[name*='date'], "
                "input[name*='Date'], input[name*='from'], input[name*='to']")
            if len(date_inputs) >= 2:
                try:
                    driver.execute_script(
                        "arguments[0].value=arguments[1];", date_inputs[0], from_dt)
                    driver.execute_script(
                        "arguments[0].value=arguments[1];", date_inputs[1], to_dt)
                    for di in [date_inputs[0], date_inputs[1]]:
                        driver.execute_script(
                            "arguments[0].dispatchEvent(new Event('input'));"
                            "arguments[0].dispatchEvent(new Event('change'));", di)
                    time.sleep(0.5)
                    ss = screenshot("dates_filled")
                    add("Date Range","Filled " + from_dt + " to " + to_dt,
                        "PASS","Date fields filled",ss)
                    for btn_text in ["Generate Report","Generate","Submit","View","Search"]:
                        try:
                            btn = driver.find_element(By.XPATH,
                                "//button[contains(text(),'" + btn_text + "')] | "
                                "//input[@type='submit'][@value='" + btn_text + "']")
                            if btn.is_displayed():
                                btn.click(); time.sleep(3)
                                ss = screenshot("report_result")
                                new_soup = BeautifulSoup(driver.page_source,"html.parser")
                                body     = new_soup.get_text(strip=True).lower()
                                tbls     = new_soup.find_all("table")
                                errs     = [k for k in ["invalid","error","required",
                                            "no record","please"] if k in body]
                                add("Report Generate","Clicked: " + btn_text,
                                    "FAIL" if errs else "PASS",
                                    ("Errors: " + ", ".join(errs)) if errs
                                    else str(len(tbls)) + " result table(s) returned",ss)
                                driver.back(); time.sleep(1.5); break
                        except: pass
                except Exception as e:
                    add("Date Range","Fill+Submit","ERROR",str(e)[:60])

        # 4. Dropdowns
        for sel in driver.find_elements(By.CSS_SELECTOR,"select")[:4]:
            try:
                s    = SeleniumSelect(sel)
                name = sel.get_attribute("name") or sel.get_attribute("id") or "dropdown"
                opts = [o.text.strip() for o in s.options if o.text.strip()]
                if len(opts) > 1:
                    s.select_by_index(1); time.sleep(0.3)
                    add("Dropdown [" + name[:20] + "]",
                        "Selected: " + opts[1][:30],
                        "PASS",str(len(opts)) + " options available")
            except Exception as e:
                add("Dropdown","Select","FAIL",str(e)[:50])

        # 5. Tables
        for idx, tbl in enumerate(
                driver.find_elements(By.CSS_SELECTOR,"table")[:3],1):
            try:
                data_rows = max(0, len(tbl.find_elements(By.TAG_NAME,"tr")) - 1)
                add("Table " + str(idx),"Row count",
                    "PASS" if data_rows > 0 else "WARN",
                    str(data_rows) + " data row(s) visible")
            except: pass

    except Exception as e:
        add("Page","Deep test","ERROR",str(e)[:80])

    return results


# ─────────────────────────────────────────────────────────────────────────────
# STEP 12: Test every URL (TEST PORTAL ONLY)
# ─────────────────────────────────────────────────────────────────────────────
def test_urls(driver, rows, cfg, test_data):
    print("\n[TEST] Testing", len(rows), "URLs on TEST portal...")
    print("       Live portal is NOT touched from this point.")
    results = []

    for i, row in enumerate(rows):
        main  = row["main"]
        sub   = row["sub"]
        url   = row["url"]
        issue = row["issue"]
        href  = row["href"]
        label = (main + " > " + sub) if sub else main

        print("\n  [" + str(i+1).rjust(3) + "/" + str(len(rows)) + "] " + label)
        print("       URL:", (url or href)[:70])

        if not url:
            print("       STATUS: PORTAL BUG")
            results.append({
                "main": main, "sub": sub, "url": href,
                "status": "PORTAL BUG",
                "issue": issue if issue else "Broken URL: " + href,
                "load": "", "title": "", "has_form": "",
                "form_count": "", "has_table": "", "file_upload": "",
                "headings": "", "sections": "", "forms": "",
                "labels": "", "inputs": "", "buttons": "",
                "tables": "", "dropdowns": "",
                "fields_detail": "",
                "deep_results": "", "deep_notes": "", "screenshot": ""
            })
            continue

        try:
            t0 = time.time()
            driver.get(url)
            try:
                WebDriverWait(driver, 8).until(
                    lambda d: d.execute_script(
                        "return document.readyState") == "complete")
            except: pass
            time.sleep(2)

            load  = round(time.time() - t0, 2)
            title = driver.title[:55]

            safe    = re.sub(r"[^\w\-]", "_", label)[:45]
            ss_name = str(i+1).zfill(3) + "_" + safe + ".png"
            driver.save_screenshot(
                os.path.join(cfg["test"]["screenshots"], ss_name))

            soup_p = BeautifulSoup(driver.page_source, "html.parser")
            body   = soup_p.get_text(separator=" ", strip=True).lower()
            comps  = extract_components(driver, soup_p)

            forms_all   = soup_p.find_all("form")
            has_form    = "YES" if forms_all else "NO"
            form_count  = len(forms_all)
            has_table   = "YES" if soup_p.find("table") else "NO"
            file_upload = "YES" if soup_p.find("input",{"type":"file"}) else "NO"

            err_kw = ["access denied","not found","unauthorized","404","403","500"]
            errors = [k for k in err_kw if k in body]
            status = ("FAIL" if errors else
                      "PASS*" if "PORTAL BUG" in issue else "PASS")

            deep         = []
            deep_summary = ""
            deep_notes   = ""
            if not test_data.get("skip_deep", False):
                print("       Running deep component tests...")
                deep = deep_test_page(driver, url, main, sub, cfg, i+1, test_data)
                deep_summary = " | ".join(
                    r["component"] + ":" + r["result"] for r in deep)
                deep_notes = " | ".join(
                    r["component"] + "→" + r["notes"] for r in deep)
                try:
                    driver.get(url); time.sleep(1.5)
                except: pass

            if comps["headings"]:
                print("       Headings:", " | ".join(comps["headings"][:2]))
            if comps["fields_detail"]:
                print("       Fields  :", len(comps["fields_detail"]), "fields found")
            if comps["buttons"]:
                print("       Buttons :", " | ".join(comps["buttons"][:3]))
            if deep_summary:
                print("       DeepTest:", deep_summary[:80])
            print("       STATUS  :", status, "(" + str(load) + "s)")

            results.append({
                "main":          main,
                "sub":           sub,
                "url":           url,
                "status":        status,
                "issue":         issue,
                "load":          load,
                "title":         title,
                "has_form":      has_form,
                "form_count":    form_count,
                "has_table":     has_table,
                "file_upload":   file_upload,
                "headings":      " | ".join(comps["headings"])[:100],
                "sections":      " | ".join(comps["page_sections"])[:80],
                "forms":         " | ".join(comps["forms"])[:120],
                "labels":        " | ".join(comps["labels"])[:100],
                "inputs":        " | ".join(comps["inputs"])[:200],
                "buttons":       " | ".join(comps["buttons"])[:100],
                "tables":        " | ".join(comps["tables"])[:100],
                "dropdowns":     " | ".join(comps["dropdowns"])[:150],
                "fields_detail": " || ".join(comps["fields_detail"]),
                "deep_results":  deep_summary[:150],
                "deep_notes":    deep_notes[:250],
                "screenshot":    ss_name,
            })

        except (TimeoutException, WebDriverException) as e:
            print("       STATUS: ERROR -", str(e)[:60])
            results.append({
                "main": main, "sub": sub, "url": url,
                "status": "ERROR", "issue": str(e)[:80],
                "load": "", "title": "", "has_form": "",
                "form_count": "", "has_table": "", "file_upload": "",
                "headings": "", "sections": "", "forms": "",
                "labels": "", "inputs": "", "buttons": "",
                "tables": "", "dropdowns": "",
                "fields_detail": "",
                "deep_results": "", "deep_notes": "", "screenshot": ""
            })

    return results


# ─────────────────────────────────────────────────────────────────────────────
# STEP 13: Excel report — 5 sheets
# ─────────────────────────────────────────────────────────────────────────────
def write_excel(live_rows, test_rows, comparison, results, cfg, test_data):
    print("\n[EXCEL] Writing report...")
    wb = Workbook()

    # ── Sheet 1: Live Menu Structure (READ ONLY reference) ───────────────────
    ws1 = wb.active
    ws1.title = "Live Menu (Reference)"
    tk_label = "LIVE PORTAL — READ ONLY REFERENCE"
    ws1.cell(1, 1, tk_label).font = fnt(True, "FFCC0000", 12)
    ws1.merge_cells("A1:E1")
    ws1.cell(1,1).alignment = aln("center")
    ws1.cell(1,1).fill = fill("FFFFF3CD")

    for ci, (h, w) in enumerate(zip(
        ["#","Level 1 (Main Menu)","Level 2 (Sub Menu)","URL","Issue"],
        [5,28,35,65,45]
    ), 1):
        c = ws1.cell(2, ci, h)
        c.fill = fill("FF8B0000"); c.font = fnt(True,"FFFFFFFF",11)
        c.alignment = aln("center"); c.border = brd()
        ws1.column_dimensions[c.column_letter].width = w

    for ri, row in enumerate(live_rows, 3):
        is_sub = bool(row["sub"])
        bg = ("FFFFF3CD" if row["issue"] else
              "FFFFE8E8" if not is_sub else
              "FFFAFAFA" if ri%2==0 else "FFF5F5F5")
        for ci, val in enumerate([ri-2, row["main"], row["sub"],
                                   row["url"] or row["href"], row["issue"]], 1):
            c = ws1.cell(ri, ci, val)
            c.fill = fill(bg); c.border = brd(); c.alignment = aln()
            c.font = (Font(color="FF8B0000", size=10, underline="single") if ci==4
                      else fnt(color="FFB71C1C", bold=True) if ci==5 and val
                      else fnt(bold=not is_sub and ci==2))
    ws1.freeze_panes = "A3"

    # ── Sheet 2: Menu Comparison ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Menu Comparison")
    for ci, (h, w) in enumerate(zip(
        ["#","Main Menu","Sub Menu","Status","Live URL","Test URL"],
        [4,25,32,18,55,55]
    ), 1):
        c = ws2.cell(1, ci, h)
        c.fill = fill("FF4A148C"); c.font = fnt(True,"FFFFFFFF",11)
        c.alignment = aln("center"); c.border = brd()
        ws2.column_dimensions[c.column_letter].width = w

    for ri, row in enumerate(comparison, 2):
        st = row["status"]
        if   st == "MATCH":           bg = "FFE8F5E9"; fc = "FF1B5E20"
        elif st == "MISSING IN TEST":  bg = "FFFFEBEE"; fc = "FFB71C1C"
        else:                          bg = "FFFFF8E1"; fc = "FFF57F17"
        row_bg = bg
        vals = [ri-1, row["main"], row["sub"], st,
                row["live_url"], row["test_url"]]
        for ci, val in enumerate(vals, 1):
            c = ws2.cell(ri, ci, val)
            c.fill = fill(row_bg) if ci != 4 else fill(bg)
            c.font = fnt(color=fc, bold=True) if ci==4 else fnt()
            c.alignment = aln(); c.border = brd()
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = "A1:F" + str(len(comparison)+1)

    # ── Sheet 3: Test Results ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Test Results")
    s3_cols = ["#","Main Menu","Sub Menu","URL","Result","Issue",
               "Load(s)","Page Title","Form?","Forms","Table?","File Upload?",
               "Page Headings","Forms Detail","Input Fields","Buttons",
               "Tables","Dropdowns","Component Tests","Component Notes","Screenshot"]
    s3_wids = [4,18,22,45,10,35,7,30,7,7,7,10,
               40,35,60,35,40,40,50,70,30]
    for ci, (h, w) in enumerate(zip(s3_cols, s3_wids), 1):
        c = ws3.cell(1, ci, h)
        c.fill = fill("FF005B96"); c.font = fnt(True,"FFFFFFFF",11)
        c.alignment = aln("center"); c.border = brd()
        ws3.column_dimensions[c.column_letter].width = w

    pass_c = fail_c = err_c = bug_c = 0
    for ri, r in enumerate(results, 2):
        st = r["status"]
        if   st=="PASS":       sf=fill("FFE8F5E9"); ff=fnt(True,"FF1B5E20")
        elif st=="PASS*":      sf=fill("FFF1F8E9"); ff=fnt(True,"FF33691E")
        elif st=="FAIL":       sf=fill("FFFFF3E0"); ff=fnt(True,"FFE65100")
        elif st=="PORTAL BUG": sf=fill("FFFCE4EC"); ff=fnt(True,"FFC62828")
        else:                  sf=fill("FFFFEBEE"); ff=fnt(True,"FFB71C1C")
        if   st in ["PASS","PASS*"]: pass_c+=1
        elif st=="FAIL":             fail_c+=1
        elif st=="PORTAL BUG":       bug_c+=1
        else:                        err_c+=1
        bg = "FFFAFAFA" if ri%2==0 else "FFFFFFFF"
        vals = [ri-1, r["main"], r["sub"],
                r["url"] or r.get("href",""),
                st, r["issue"], r["load"], r["title"],
                r.get("has_form",""), r.get("form_count",""),
                r.get("has_table",""), r.get("file_upload",""),
                r.get("headings",""), r.get("forms",""), r.get("inputs",""),
                r.get("buttons",""), r.get("tables",""), r.get("dropdowns",""),
                r.get("deep_results",""), r.get("deep_notes",""),
                r.get("screenshot","")]
        for ci, val in enumerate(vals, 1):
            c = ws3.cell(ri, ci, val)
            c.fill = sf if ci==5 else fill(bg)
            c.font = (ff if ci==5 else
                      fnt(color="FFB71C1C",bold=True) if ci==6 and val else fnt())
            c.alignment = aln(); c.border = brd()
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = "A1:U" + str(len(results)+1)

    # ── Sheet 4: Field Details ───────────────────────────────────────────────
    ws4 = wb.create_sheet("Field Details")
    for ci, (h, w) in enumerate(zip(
        ["#","Main Menu","Sub Menu","Page URL","Form",
         "Field Type","Field Name / Label","Mandatory?"],
        [4,20,25,52,8,12,50,12]
    ), 1):
        c = ws4.cell(1, ci, h)
        c.fill = fill("FF1A237E"); c.font = fnt(True,"FFFFFFFF",11)
        c.alignment = aln("center"); c.border = brd()
        ws4.column_dimensions[c.column_letter].width = w

    row_num = 2
    for r in results:
        fields_raw = r.get("fields_detail","")
        if not fields_raw: continue
        for field_line in fields_raw.split(" || "):
            field_line = field_line.strip()
            if not field_line: continue
            parts     = field_line.split(" › ", 1)
            form_part = parts[0].strip() if len(parts) > 1 else "Form?"
            rest      = parts[1].strip() if len(parts) > 1 else field_line
            fparts    = [p.strip() for p in rest.split(" | ")]
            ftype     = fparts[0] if len(fparts) > 0 else ""
            fname     = fparts[1] if len(fparts) > 1 else ""
            fmand     = fparts[2] if len(fparts) > 2 else ""
            if "Options:" in fmand:
                fmand = fmand.split("|")[0].strip()
            is_mand   = "MANDATORY" in fmand.upper()
            mand_bg   = "FFFCE4EC" if is_mand else "FFE8F5E9"
            mand_font = (fnt(color="FFB71C1C",bold=True) if is_mand
                         else fnt(color="FF1B5E20",bold=True))
            row_bg    = "FFFAFAFA" if row_num%2==0 else "FFFFFFFF"
            vals = [row_num-1, r["main"], r["sub"],
                    r["url"], form_part, ftype, fname, fmand]
            for ci, val in enumerate(vals, 1):
                c = ws4.cell(row_num, ci, val)
                c.fill = fill(mand_bg) if ci==8 else fill(row_bg)
                c.font = mand_font if ci==8 else fnt()
                c.alignment = aln(); c.border = brd()
            row_num += 1
    ws4.freeze_panes = "A2"
    ws4.auto_filter.ref = "A1:H" + str(row_num)

    # ── Sheet 5: Summary ─────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Summary")
    ws5.column_dimensions["A"].width = 32
    ws5.column_dimensions["B"].width = 38

    live_mains  = len({r["main"] for r in live_rows})
    test_mains  = len({r["main"] for r in test_rows})
    match_count = sum(1 for c in comparison if c["status"]=="MATCH")
    missing     = sum(1 for c in comparison if c["status"]=="MISSING IN TEST")
    extra       = sum(1 for c in comparison if c["status"]=="EXTRA IN TEST")
    total_fields = sum(
        len(r.get("fields_detail","").split(" || "))
        for r in results if r.get("fields_detail",""))
    mand_fields = sum(
        1 for r in results
        for f in r.get("fields_detail","").split(" || ")
        if "MANDATORY" in f.upper())

    summary = [
        ("Report Generated",    datetime.now().strftime("%Y-%m-%d %H:%M:%S"), None),
        ("", "", None),
        ("── PORTALS ──",       "", None),
        ("Live Portal URL",     cfg["live"]["url"],       None),
        ("Live Username",       cfg["live"]["username"],  None),
        ("Test Portal URL",     cfg["test"]["url"],       None),
        ("Test Username",       cfg["test"]["username"],  None),
        ("Output Folder",       cfg["test"]["folder"],    None),
        ("", "", None),
        ("── TEST DATA USED ──","", None),
        ("Diary Number",        test_data.get("diary_no","(not provided)"),   None),
        ("Case Number",         test_data.get("case_no","(not provided)"),    None),
        ("Date Range",          test_data.get("from_date","") + " to " +
                                test_data.get("to_date",""),                  None),
        ("Party Name",          test_data.get("party_name","(not provided)"), None),
        ("Deep Testing",        "SKIPPED" if test_data.get("skip_deep")
                                else "ENABLED",                               None),
        ("", "", None),
        ("── MENU COMPARISON ──","", None),
        ("Live Menu Items",     len(live_rows),   None),
        ("Test Menu Items",     len(test_rows),   None),
        ("Matching Items",      match_count,      "FF1B5E20"),
        ("Missing in Test",     missing,          "FFB71C1C"),
        ("Extra in Test",       extra,            "FFF57F17"),
        ("", "", None),
        ("── TEST RESULTS ──",  "", None),
        ("URLs Tested",         len(results),     None),
        ("Total Fields Found",  total_fields,     None),
        ("Mandatory Fields",    mand_fields,      None),
        ("Optional Fields",     total_fields - mand_fields, None),
        ("", "", None),
        ("PASS",                pass_c,  "FF1B5E20"),
        ("FAIL",                fail_c,  "FFE65100"),
        ("PORTAL BUG",          bug_c,   "FFC62828"),
        ("ERROR",               err_c,   "FFB71C1C"),
    ]
    for ri, (lbl, val, color) in enumerate(summary, 1):
        l = ws5.cell(ri, 1, lbl)
        v = ws5.cell(ri, 2, str(val))
        l.font = fnt(bold=True, sz=11)
        v.font = fnt(color=color, sz=12, bold=True) if color else fnt(sz=11)

    wb.save(cfg["test"]["excel"])
    print("    Saved:", os.path.abspath(cfg["test"]["excel"]))
    print("\n    PASS:", pass_c, "| FAIL:", fail_c,
          "| PORTAL BUG:", bug_c, "| ERROR:", err_c)
    print("    Menu match:", match_count, "| Missing in test:", missing,
          "| Extra:", extra)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    config = get_all_config()
    if not config["live"] or not config["test"]:
        print("Cancelled."); return

    live_cfg  = config["live"]
    test_cfg  = config["test"]
    test_data = config["data"]

    print("\n" + "="*56)
    print("  LIVE portal :", live_cfg["url"])
    print("  LIVE user   :", live_cfg["username"], "  [READ ONLY]")
    print("  TEST portal :", test_cfg["url"])
    print("  TEST user   :", test_cfg["username"], "  [FULL TESTING]")
    print("  Folder      :", test_cfg["folder"])
    print("  Deep Test   :", "NO" if test_data.get("skip_deep") else "YES")
    print("="*56)

    # We use ONE browser but visit two portals sequentially
    driver = init_driver()
    live_rows = []
    test_rows = []

    try:
        # ── Phase 1: LIVE portal — READ ONLY ────────────────────────────────
        print("\n" + "="*56)
        print("  PHASE 1: LIVE PORTAL — READ ONLY")
        print("  No data will be submitted or modified")
        print("="*56)

        login(driver, live_cfg, test_cfg["screenshots"], "LIVE")
        live_rows = get_menu_readonly(
            driver, live_cfg, test_cfg["live_html"])

        print("\n[OK] Live portal menu extracted. Closing live session.")
        driver.delete_all_cookies()
        time.sleep(2)

        # ── Phase 2: TEST portal — FULL TESTING ─────────────────────────────
        print("\n" + "="*56)
        print("  PHASE 2: TEST PORTAL — FULL TESTING")
        print("  Forms will be filled, buttons clicked, screenshots taken")
        print("="*56)

        login(driver, test_cfg, test_cfg["screenshots"], "TEST")
        html      = expand_test_sidebar(driver, config)
        test_rows = parse_menu_from_html(html, test_cfg["base"])
        save_csv(test_rows, config)

        # Compare menus
        comparison = compare_menus(live_rows, test_rows)

        # Test all URLs
        results = test_urls(driver, test_rows, config, test_data)

        # Write Excel
        write_excel(live_rows, test_rows, comparison, results, config, test_data)

        print("\n" + "="*56)
        print("  DONE!")
        print("  Folder :", test_cfg["folder"])
        print("  CSV    :", test_cfg["csv"])
        print("  Excel  :", test_cfg["excel"])
        print("="*56)

    except Exception as e:
        print("\n[ERROR]", e)
        import traceback; traceback.print_exc()
    finally:
        input("\nPress ENTER to close browser...")
        driver.quit()


if __name__ == "__main__":
    main()
