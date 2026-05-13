"""
Universal Portal Testing Tool v1.1
- GUI window for URL, username, password input
- Login with CAPTCHA support
- Extract sidebar menu (Level 1 + Level 2)
- Save to CSV
- Hit every URL, take screenshot, check status
- Create Excel test report with screenshot paths
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import tkinter as tk
from tkinter import ttk, messagebox
import time, os, csv
from datetime import datetime

# ── STEP 1: GUI — Get credentials from user ───────────────────────────────────
def get_config():
    config = {}

    def submit():
        url  = url_var.get().strip()
        user = user_var.get().strip()
        pwd  = pass_var.get().strip()
        if not url or not user or not pwd:
            messagebox.showerror("Error", "All fields are required")
            return
        config["url"]      = url
        config["username"] = user
        config["password"] = pwd

        # Create output folder named after username + timestamp
        ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
        folder = user + "_" + ts
        os.makedirs(folder, exist_ok=True)
        os.makedirs(os.path.join(folder, "screenshots"), exist_ok=True)

        config["folder"]      = folder
        config["screenshots"] = os.path.join(folder, "screenshots")
        config["html"]        = os.path.join(folder, "portal.html")
        config["csv"]         = os.path.join(folder, "menu.csv")
        config["excel"]       = os.path.join(folder, "test_report.xlsx")

        root.destroy()

    root = tk.Tk()
    root.title("Portal Testing Tool v1.1")
    root.geometry("520x320")
    root.resizable(False, False)

    # Header
    tk.Label(root, text="Universal Portal Testing Tool",
             font=("Arial", 15, "bold"), fg="#003366").pack(pady=14)
    tk.Label(root, text="Enter credentials to begin automated testing",
             font=("Arial", 10), fg="#555555").pack()

    frame = ttk.Frame(root, padding=20)
    frame.pack(fill="both", expand=True)

    url_var  = tk.StringVar(value="https://drt.etribunals.gov.in/cis2.0/filing/login")
    user_var = tk.StringVar(value="")
    pass_var = tk.StringVar(value="")

    ttk.Label(frame, text="Portal URL:").grid(row=0, column=0, sticky="w", pady=8)
    ttk.Entry(frame, textvariable=url_var, width=45).grid(row=0, column=1, padx=8)

    ttk.Label(frame, text="Username:").grid(row=1, column=0, sticky="w", pady=8)
    ttk.Entry(frame, textvariable=user_var, width=45).grid(row=1, column=1, padx=8)

    ttk.Label(frame, text="Password:").grid(row=2, column=0, sticky="w", pady=8)
    ttk.Entry(frame, textvariable=pass_var, show="*", width=45).grid(row=2, column=1, padx=8)

    ttk.Button(frame, text="▶  Start Testing", command=submit).grid(
        row=3, column=1, pady=20, sticky="e")

    root.mainloop()
    return config


# ── STEP 2: Launch Chrome ─────────────────────────────────────────────────────
def init_driver():
    opts = webdriver.ChromeOptions()
    opts.add_argument("--start-maximized")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=opts)


# ── STEP 3: Login ─────────────────────────────────────────────────────────────
def login(driver, cfg):
    print("\n[1] Opening login page...")
    driver.get(cfg["url"])
    time.sleep(3)

    try:
        driver.find_element(By.ID, "user_name").send_keys(cfg["username"])
        print("    Username filled")
    except:
        try:
            driver.find_element(By.NAME, "username").send_keys(cfg["username"])
            print("    Username filled")
        except:
            print("    [!] Fill username manually in browser")

    try:
        driver.find_element(By.ID, "user_pass").send_keys(cfg["password"])
        print("    Password filled")
    except:
        try:
            driver.find_element(By.NAME, "password").send_keys(cfg["password"])
            print("    Password filled")
        except:
            print("    [!] Fill password manually in browser")

    # Take screenshot of login page
    driver.save_screenshot(os.path.join(cfg["screenshots"], "01_login_page.png"))

    print("\n" + "="*52)
    print("  ACTION: Solve CAPTCHA then click LOGIN")
    print("  Waiting up to 120 seconds...")
    print("="*52 + "\n")

    WebDriverWait(driver, 120).until(
        lambda d: "login" not in d.current_url
    )
    time.sleep(3)

    # Screenshot after login
    driver.save_screenshot(os.path.join(cfg["screenshots"], "02_dashboard.png"))
    print("[OK] Logged in:", driver.current_url)


# ── STEP 4: Expand sidebar + save HTML ───────────────────────────────────────
def get_html(driver, cfg):
    print("\n[2] Expanding sidebar menus...")
    driver.execute_script("""
        document.querySelectorAll('.collapse').forEach(el => {
            el.classList.add('show');
        });
        document.querySelectorAll('[aria-expanded]').forEach(el => {
            el.setAttribute('aria-expanded', 'true');
        });
    """)
    time.sleep(2)

    html = driver.page_source
    with open(cfg["html"], "w", encoding="utf-8") as f:
        f.write(html)
    print("    HTML saved:", cfg["html"])
    return html


# ── STEP 5: Parse menu ────────────────────────────────────────────────────────
def parse_menu(html, cfg):
    print("\n[3] Parsing menu items...")
    from urllib.parse import urlparse
    parsed = urlparse(cfg["url"])
    base   = parsed.scheme + "://" + parsed.netloc

    soup = BeautifulSoup(html, "html.parser")
    rows = []

    sidebar = soup.find("ul", id="accordionSidebar")
    if not sidebar:
        sidebar = soup.find("ul", class_=lambda c: c and "navbar-nav" in c)

    if sidebar:
        for li in sidebar.find_all("li", class_="nav-item"):
            main_a = li.find("a", class_="nav-link")
            if not main_a:
                continue
            main_name = main_a.get_text(strip=True)
            collapse  = li.find("div", class_="collapse-inner")
            if collapse:
                for a in collapse.find_all("a", class_="dropdown-item"):
                    sub  = a.get_text(strip=True)
                    href = a.get("href", "")
                    url  = (base + href) if href.startswith("/") else href
                    rows.append([main_name, sub, url])
                    print("    +", main_name, ">", sub)
            else:
                href = main_a.get("href", "")
                if href and "javascript" not in href:
                    url = (base + href) if href.startswith("/") else href
                    rows.append([main_name, "", url])
                    print("    +", main_name, "(direct)")

    print("    Total:", len(rows), "items")
    return rows


# ── STEP 6: Save CSV ──────────────────────────────────────────────────────────
def save_csv(rows, cfg):
    with open(cfg["csv"], "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Main Menu", "Sub Menu", "URL"])
        w.writerows(rows)
    print("\n[4] CSV saved:", cfg["csv"])


# ── STEP 7: Test URLs + take screenshots ─────────────────────────────────────
def test_urls(driver, rows, cfg):
    print("\n[5] Testing", len(rows), "URLs + taking screenshots...")
    results = []

    for i, (main, sub, url) in enumerate(rows):
        label = (main + " > " + sub) if sub else main
        print("  [" + str(i+1).rjust(3) + "/" + str(len(rows)) + "]",
              label[:42].ljust(42), end=" ... ", flush=True)

        if not url:
            print("NO URL")
            results.append([main, sub, url, "NO URL", "", "", "", "", ""])
            continue

        try:
            t0 = time.time()
            driver.get(url)
            time.sleep(1.5)
            load  = round(time.time() - t0, 2)
            title = driver.title[:50]

            # Take screenshot — name based on index + menu label
            safe_label = "".join(c if c.isalnum() or c in "-_" else "_"
                                 for c in label)[:50]
            ss_name = str(i+1).zfill(3) + "_" + safe_label + ".png"
            ss_path = os.path.join(cfg["screenshots"], ss_name)
            driver.save_screenshot(ss_path)

            soup   = BeautifulSoup(driver.page_source, "html.parser")
            body   = soup.get_text(separator=" ", strip=True).lower()
            forms  = len(soup.find_all("form"))
            tables = len(soup.find_all("table"))

            err_kw = ["access denied", "not found", "unauthorized", "404", "403", "500"]
            errors = [k for k in err_kw if k in body]
            status = "FAIL" if errors else "PASS"

            print(status + " (" + str(load) + "s)")
            results.append([main, sub, url, status, load, title,
                            str(forms) + " form(s)", str(tables) + " table(s)",
                            ss_name])
        except Exception as e:
            print("ERROR")
            results.append([main, sub, url, "ERROR", "", "", str(e)[:60], "", ""])

    return results


# ── STEP 8: Write Excel report ────────────────────────────────────────────────
def write_excel(rows, results, cfg):
    print("\n[6] Writing Excel report...")

    def fill(c):  return PatternFill("solid", fgColor=c)
    def fnt(bold=False, color="FF000000", sz=10):
        return Font(bold=bold, color=color, size=sz)
    def brd():
        s = Side(style="thin", color="FFD0D0D0")
        return Border(left=s, right=s, top=s, bottom=s)
    def aln(h="left"):
        return Alignment(horizontal=h, vertical="center", wrap_text=True)

    wb = Workbook()

    # ── Sheet 1: Menu Structure ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Menu Structure"
    for ci, (h, w) in enumerate(zip(
        ["#", "Level 1 (Main Menu)", "Level 2 (Sub Menu)", "URL"],
        [5, 28, 35, 70]
    ), 1):
        c = ws1.cell(1, ci, h)
        c.fill = fill("FF003366"); c.font = fnt(True, "FFFFFFFF", 11)
        c.alignment = aln("center"); c.border = brd()
        ws1.column_dimensions[c.column_letter].width = w

    for ri, (main, sub, url) in enumerate(rows, 2):
        is_sub = bool(sub)
        bg     = "FFE8F4FD" if not is_sub else ("FFFAFAFA" if ri%2==0 else "FFF5F5F5")
        for ci, val in enumerate([ri-1, main, sub, url], 1):
            c = ws1.cell(ri, ci, val)
            c.fill = fill(bg); c.border = brd(); c.alignment = aln()
            c.font = Font(color="FF0055AA", size=10, underline="single") \
                     if ci == 4 else fnt(bold=not is_sub and ci == 2)
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Test Results ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Test Results")
    for ci, (h, w) in enumerate(zip(
        ["#", "Main Menu", "Sub Menu", "URL", "Result",
         "Load(s)", "Page Title", "Forms", "Tables", "Screenshot"],
        [4, 22, 30, 52, 9, 8, 35, 10, 10, 35]
    ), 1):
        c = ws2.cell(1, ci, h)
        c.fill = fill("FF005B96"); c.font = fnt(True, "FFFFFFFF", 11)
        c.alignment = aln("center"); c.border = brd()
        ws2.column_dimensions[c.column_letter].width = w

    pass_count = fail_count = error_count = 0
    for ri, res in enumerate(results, 2):
        st = res[3] if len(res) > 3 else ""
        if   st == "PASS":  bg_st = fill("FFE8F5E9"); ft_st = fnt(True, "FF1B5E20")
        elif st == "FAIL":  bg_st = fill("FFFFF3E0"); ft_st = fnt(True, "FFE65100")
        else:               bg_st = fill("FFFFEBEE"); ft_st = fnt(True, "FFB71C1C")
        if   st == "PASS":  pass_count  += 1
        elif st == "FAIL":  fail_count  += 1
        else:               error_count += 1

        row_bg = "FFFAFAFA" if ri%2==0 else "FFFFFFFF"
        while len(res) < 10: res.append("")
        for ci, val in enumerate(res[:10], 1):
            c = ws2.cell(ri, ci, val if ci != 1 else ri-1)
            c.fill = bg_st if ci == 5 else fill(row_bg)
            c.font = ft_st if ci == 5 else fnt()
            c.alignment = aln(); c.border = brd()
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Summary ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 25
    for ri, (lbl, val) in enumerate([
        ("Report Date",       datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Portal URL",        cfg["url"]),
        ("Username",          cfg["username"]),
        ("Output Folder",     cfg["folder"]),
        ("", ""),
        ("Total Menu Items",  len(rows)),
        ("Total URLs Tested", len(results)),
        ("", ""),
        ("✅  PASS",          pass_count),
        ("⚠️   FAIL",         fail_count),
        ("❌  ERROR",         error_count),
    ], 1):
        ws3.cell(ri, 1, lbl).font = fnt(True, sz=11)
        ws3.cell(ri, 2, str(val)).font = fnt(sz=11)

    wb.save(cfg["excel"])
    print("    Excel saved:", os.path.abspath(cfg["excel"]))
    print("\n    PASS:", pass_count, "| FAIL:", fail_count, "| ERROR:", error_count)


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    cfg = get_config()
    if not cfg:
        print("Cancelled.")
        return

    print("\n" + "="*52)
    print("  Output folder:", cfg["folder"])
    print("  Screenshots  :", cfg["screenshots"])
    print("="*52)

    driver = init_driver()
    try:
        login(driver, cfg)
        html    = get_html(driver, cfg)
        rows    = parse_menu(html, cfg)
        save_csv(rows, cfg)
        results = test_urls(driver, rows, cfg)
        write_excel(rows, results, cfg)

        print("\n" + "="*52)
        print("  DONE!")
        print("  Folder :", cfg["folder"])
        print("  CSV    :", cfg["csv"])
        print("  Excel  :", cfg["excel"])
        print("  Screenshots:", cfg["screenshots"])
        print("="*52)

    except Exception as e:
        print("\n[ERROR]", e)
    finally:
        input("\nPress ENTER to close browser...")
        driver.quit()

if __name__ == "__main__":
    main()