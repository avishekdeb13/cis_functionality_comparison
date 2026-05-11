"""
DRT Portal - Menu Structure Extractor + Automated Testing Tool
==============================================================
PHASE 1 : Login → Save full HTML → Parse ribbon + dropdowns → Export CSV/Excel
PHASE 2 : Visit every sub-menu URL → Check if page loads → Detect forms/tables/
           buttons → Record functionality → Append results to Excel

Run:
    pip install selenium webdriver-manager openpyxl beautifulsoup4
    python drt_scraper_v2.py
"""

import os, csv, json, time, re
from datetime import datetime
from urllib.parse import urljoin, urlparse

from bs4 import BeautifulSoup

# ── selenium ──────────────────────────────────────────────────────────────────
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, StaleElementReferenceException,
    WebDriverException,
)

# ── openpyxl ──────────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)

# ─────────────────────────────────────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────────────────────────────────────
USERNAME       = "filingdrt1"
PASSWORD       = "Sodrt1@6060"
BASE_URL       = "https://cis.drt.gov.in/drtlive/index.php"
CAPTCHA_WAIT   = 90          # seconds for manual CAPTCHA entry
HTML_DUMP      = "drt_page.html"
MENU_CSV       = "drt_menu_structure.csv"
FULL_REPORT    = "drt_full_report.xlsx"
PAGE_TIMEOUT   = 15
HOVER_PAUSE    = 1.2

# ─────────────────────────────────────────────────────────────────────────────
#  DRIVER
# ─────────────────────────────────────────────────────────────────────────────
def init_driver():
    opts = webdriver.ChromeOptions()
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument("--disable-popup-blocking")
    try:
        from webdriver_manager.chrome import ChromeDriverManager
        from selenium.webdriver.chrome.service import Service
        driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()), options=opts)
    except Exception:
        driver = webdriver.Chrome(options=opts)
    driver.set_page_load_timeout(PAGE_TIMEOUT + 10)
    return driver


# ─────────────────────────────────────────────────────────────────────────────
#  LOGIN
# ─────────────────────────────────────────────────────────────────────────────
def login(driver):
    print("\n[STEP 1] Opening login page …")
    driver.get(BASE_URL)
    time.sleep(3)

    # Try multiple possible field selectors
    def fill(field_type, value, locators):
        for loc in locators:
            try:
                el = driver.find_element(*loc)
                el.clear(); el.send_keys(value)
                print(f"  ✔ {field_type} filled ({loc})")
                return True
            except NoSuchElementException:
                pass
        print(f"  ✘ Could not find {field_type} field — fill manually")
        return False

    fill("Username", USERNAME, [
        (By.NAME, "username"), (By.NAME, "user_name"), (By.NAME, "txtUsername"),
        (By.ID,   "username"), (By.ID,   "txtUsername"),
        (By.XPATH, "//input[@type='text'][1]"),
        (By.XPATH, "//input[contains(@name,'user') or contains(@id,'user')]"),
    ])
    fill("Password", PASSWORD, [
        (By.NAME, "password"), (By.NAME, "txtPassword"),
        (By.ID,   "password"), (By.ID,   "txtPassword"),
        (By.XPATH, "//input[@type='password']"),
    ])

    print(f"""
{'='*62}
  ACTION REQUIRED — solve CAPTCHA in the browser window
  then click LOGIN.  You have {CAPTCHA_WAIT} seconds.
{'='*62}""")

    start_url = driver.current_url
    try:
        WebDriverWait(driver, CAPTCHA_WAIT).until(
            lambda d: d.current_url != start_url
        )
        print(f"\n  ✔ Logged in  →  {driver.current_url}")
    except TimeoutException:
        print("\n  [!] URL did not change — continuing anyway …")

    time.sleep(4)


# ─────────────────────────────────────────────────────────────────────────────
#  STEP 2 — SAVE HTML
# ─────────────────────────────────────────────────────────────────────────────
def save_html(driver, path=HTML_DUMP):
    html = driver.page_source
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    size = os.path.getsize(path)
    print(f"\n[STEP 2] HTML saved → {path}  ({size:,} bytes)")
    return html


# ─────────────────────────────────────────────────────────────────────────────
#  STEP 3 — PARSE MENU FROM HTML  (BeautifulSoup)
# ─────────────────────────────────────────────────────────────────────────────
def parse_menu_from_html(html, base_url):
    """
    Tries several common nav-bar patterns.
    Returns list of:
        { 'main': str, 'main_url': str,
          'sub':  str, 'sub_url':  str }   ← sub is '' when no dropdown
    """
    soup = BeautifulSoup(html, "html.parser")
    rows = []

    def abs_url(href):
        if not href or href.strip() in ("#", "", "javascript:void(0)", "javascript:;"):
            return ""
        return urljoin(base_url, href.strip())

    # ── Strategy A: look for <ul> that has nested <ul> (standard dropdowns) ──
    # Find the nav container that holds the most top-level <li> items
    best_nav = None
    best_count = 0
    for tag in soup.find_all(["nav", "ul", "div"]):
        # Only direct children li
        direct_li = tag.find_all("li", recursive=False)
        if len(direct_li) > best_count:
            best_count = len(direct_li)
            best_nav = tag

    if best_nav and best_count >= 3:
        print(f"  ✔ Strategy A: found nav with {best_count} top-level <li>")
        for li in best_nav.find_all("li", recursive=False):
            main_a   = li.find("a", recursive=False) or li.find("a")
            main_txt = (main_a.get_text(strip=True) if main_a else
                        li.get_text(separator=" ", strip=True)[:40])
            main_url = abs_url(main_a["href"]) if main_a and main_a.get("href") else ""

            # Sub-menu: nested <ul> or sibling <ul>
            sub_ul = li.find("ul")
            if sub_ul:
                for sub_li in sub_ul.find_all("li"):
                    sub_a   = sub_li.find("a")
                    sub_txt = sub_a.get_text(strip=True) if sub_a else sub_li.get_text(strip=True)
                    sub_url = abs_url(sub_a["href"]) if sub_a and sub_a.get("href") else ""
                    if sub_txt.strip("|").strip():
                        rows.append({
                            "main": main_txt, "main_url": main_url,
                            "sub":  sub_txt,  "sub_url":  sub_url,
                        })
            else:
                if main_txt.strip("|").strip():
                    rows.append({
                        "main": main_txt, "main_url": main_url,
                        "sub":  "",       "sub_url":  main_url,
                    })
        if rows:
            return rows

    # ── Strategy B: table-based nav (older gov portals) ──────────────────────
    print("  [!] Strategy A yielded nothing, trying Strategy B (table nav) …")
    for table in soup.find_all("table"):
        links = table.find_all("a")
        if len(links) >= 4:
            for a in links:
                txt = a.get_text(strip=True)
                url = abs_url(a.get("href", ""))
                if txt.strip("|").strip():
                    rows.append({"main": txt, "main_url": url,
                                 "sub":  "",  "sub_url":  url})
            if rows:
                print(f"  ✔ Strategy B: {len(rows)} links from table nav")
                return rows

    # ── Strategy C: collect ALL anchors and let caller filter ────────────────
    print("  [!] Strategy B also empty, collecting all visible anchors …")
    for a in soup.find_all("a", href=True):
        txt = a.get_text(strip=True)
        url = abs_url(a["href"])
        if txt and url and len(txt) < 60:
            rows.append({"main": txt, "main_url": url,
                         "sub":  "",  "sub_url":  url})
    return rows


# ─────────────────────────────────────────────────────────────────────────────
#  STEP 4 — LIVE HOVER to capture JS-generated dropdowns
# ─────────────────────────────────────────────────────────────────────────────
def scrape_live_dropdowns(driver, base_url):
    """
    Hovers over every ribbon item, grabs dropdown links that appear in DOM.
    Returns same structure as parse_menu_from_html.
    """
    rows = []
    actions = ActionChains(driver)

    # Find top-level nav items — broad selector then filter
    NAV_SELECTORS = [
        "//nav//li[not(ancestor::li)]/a",
        "//ul[contains(@class,'nav') or contains(@class,'menu') or contains(@class,'ribbon')]//li[not(ancestor::ul//li)]/a",
        "//div[contains(@class,'nav') or contains(@class,'menu')]//a[not(ancestor::ul//ul)]",
        "//table[1]//td/a",
        "//ul[1]//li/a",
    ]

    top_items = []
    for sel in NAV_SELECTORS:
        items = driver.find_elements(By.XPATH, sel)
        items = [i for i in items if i.is_displayed() and i.text.strip()]
        if len(items) >= 3:
            top_items = items
            print(f"  ✔ Found {len(items)} ribbon items with: {sel}")
            break

    if not top_items:
        print("  [!] No ribbon items found via live hover.")
        return rows

    # Snapshot text+href so we can re-find after hover
    snapshot = []
    for el in top_items:
        try:
            snapshot.append({
                "text": el.text.strip(),
                "href": el.get_attribute("href") or "",
            })
        except StaleElementReferenceException:
            pass

    def abs_url(href):
        if not href or href.strip() in ("#", "", "javascript:void(0)", "javascript:;"):
            return ""
        return urljoin(base_url, href.strip())

    for idx, snap in enumerate(snapshot):
        main_txt = snap["text"]
        main_url = abs_url(snap["href"])
        if not main_txt.strip("|").strip():
            continue

        print(f"\n  Ribbon [{idx+1}/{len(snapshot)}] : {main_txt}")

        # Re-locate element
        try:
            candidates = driver.find_elements(
                By.XPATH, f"//a[normalize-space(text())='{main_txt}']"
            )
            el = next((c for c in candidates if c.is_displayed()), None)
            if not el:
                rows.append({"main": main_txt, "main_url": main_url,
                             "sub": "", "sub_url": main_url})
                continue

            # Move to element
            driver.execute_script("arguments[0].scrollIntoView(true);", el)
            actions.move_to_element(el).perform()
            time.sleep(HOVER_PAUSE)

            # Capture any newly-visible dropdown links
            dropdown_links = driver.find_elements(
                By.XPATH,
                "//ul[contains(@class,'dropdown') or contains(@style,'block') "
                "or contains(@style,'visible') or contains(@class,'open')]//a"
                " | //li[contains(@class,'open') or contains(@class,'active')]//ul//a"
                " | //div[contains(@class,'dropdown-menu') and not(contains(@style,'none'))]//a"
            )
            dropdown_links = [l for l in dropdown_links
                              if l.is_displayed() and l.text.strip()
                              and l.text.strip() != main_txt]

            if dropdown_links:
                for sub in dropdown_links:
                    sub_txt = sub.text.strip()
                    sub_url = abs_url(sub.get_attribute("href") or "")
                    rows.append({"main": main_txt, "main_url": main_url,
                                 "sub":  sub_txt,  "sub_url":  sub_url})
                    print(f"    └─ {sub_txt}  →  {sub_url}")
            else:
                rows.append({"main": main_txt, "main_url": main_url,
                             "sub": "", "sub_url": main_url})
                print(f"    └─ (no dropdown)")

            # Move away to close dropdown
            actions.move_by_offset(0, 300).perform()
            time.sleep(0.4)

        except Exception as e:
            print(f"    [!] Error: {e}")
            rows.append({"main": main_txt, "main_url": main_url,
                         "sub": "", "sub_url": main_url})

    return rows


# ─────────────────────────────────────────────────────────────────────────────
#  STEP 5 — FUNCTIONAL TEST each URL
# ─────────────────────────────────────────────────────────────────────────────
def test_url(driver, url, label):
    """
    Visit a URL and return a dict with:
        status, load_time, title, has_form, has_table, has_buttons,
        input_count, button_labels, error_text, notes
    """
    result = {
        "label":         label,
        "url":           url,
        "status":        "NOT_TESTED",
        "http_code":     "",
        "load_time_sec": "",
        "page_title":    "",
        "has_form":      "",
        "has_table":     "",
        "has_buttons":   "",
        "input_count":   "",
        "button_labels": "",
        "error_text":    "",
        "notes":         "",
    }
    if not url:
        result["status"] = "NO_URL"
        return result

    try:
        t0 = time.time()
        driver.get(url)
        time.sleep(2)
        elapsed = round(time.time() - t0, 2)

        soup = BeautifulSoup(driver.page_source, "html.parser")

        result["status"]        = "OK"
        result["load_time_sec"] = elapsed
        result["page_title"]    = driver.title.strip()[:80]

        # Forms
        forms = soup.find_all("form")
        result["has_form"]    = "YES" if forms else "NO"
        result["input_count"] = sum(len(f.find_all(["input","select","textarea"]))
                                    for f in forms)

        # Tables
        result["has_table"] = "YES" if soup.find("table") else "NO"

        # Buttons
        buttons = soup.find_all(["button", "input"],
                                 attrs={"type": re.compile(r"submit|button", re.I)})
        result["has_buttons"]   = "YES" if buttons else "NO"
        result["button_labels"] = " | ".join(
            (b.get_text(strip=True) or b.get("value","") or b.get("name",""))[:30]
            for b in buttons[:8]
        )

        # Detect error messages on page
        error_keywords = ["error", "invalid", "not found", "access denied",
                          "unauthorized", "exception", "500", "404"]
        body_text = soup.get_text(separator=" ", strip=True).lower()[:1000]
        found_errors = [kw for kw in error_keywords if kw in body_text]
        result["error_text"] = ", ".join(found_errors) if found_errors else ""

        # Functional notes
        notes = []
        if soup.find("table"):
            notes.append("Data table present")
        if forms:
            notes.append(f"{len(forms)} form(s)")
        if soup.find(["select"]):
            notes.append("Dropdown selects")
        if soup.find("input", {"type": "file"}):
            notes.append("File upload")
        if soup.find(attrs={"class": re.compile(r"alert|message|success", re.I)}):
            notes.append("Alert/message box")
        result["notes"] = " | ".join(notes)

    except TimeoutException:
        result["status"] = "TIMEOUT"
    except WebDriverException as e:
        result["status"] = "ERROR"
        result["error_text"] = str(e)[:100]

    return result


# ─────────────────────────────────────────────────────────────────────────────
#  STEP 6 — WRITE EXCEL REPORT
# ─────────────────────────────────────────────────────────────────────────────
CYAN   = "FF00BCD4"
WHITE  = "FFFFFFFF"
DKBLUE = "FF0277BD"
LTGRAY = "FFF5F5F5"
GREEN  = "FF4CAF50"
RED    = "FFF44336"
ORANGE = "FFFF9800"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="FF000000", size=10):
    return Font(bold=bold, color=color, size=size)

def _border():
    s = Side(style="thin", color="FFD0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def write_excel(menu_rows, test_results, path=FULL_REPORT):
    wb = Workbook()

    # ── Sheet 1: Menu Hierarchy ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Menu Hierarchy"

    headers1 = ["#", "Main Menu", "Sub Menu", "URL", "Type"]
    col_widths1 = [5, 28, 32, 60, 12]

    for ci, (h, w) in enumerate(zip(headers1, col_widths1), 1):
        cell = ws1.cell(1, ci, h)
        cell.fill      = _fill(DKBLUE)
        cell.font      = _font(bold=True, color=WHITE, size=11)
        cell.alignment = _center()
        cell.border    = _border()
        ws1.column_dimensions[cell.column_letter].width = w

    ws1.row_dimensions[1].height = 22

    for ri, row in enumerate(menu_rows, 2):
        menu_type = "Sub-item" if row["sub"] else "Top-level"
        values = [ri - 1, row["main"], row["sub"], row.get("sub_url") or row.get("main_url",""), menu_type]
        bg = LTGRAY if ri % 2 == 0 else WHITE
        for ci, val in enumerate(values, 1):
            cell = ws1.cell(ri, ci, val)
            cell.fill      = _fill(bg) if menu_type == "Sub-item" else _fill("FFE1F5FE")
            cell.font      = _font(size=10)
            cell.alignment = _left()
            cell.border    = _border()
            if ci == 4 and val:  # URL column — blue hyperlink look
                cell.font = Font(color="FF0277BD", size=10, underline="single")

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:E{len(menu_rows)+1}"

    # ── Sheet 2: Functional Test Results ─────────────────────────────────────
    if test_results:
        ws2 = wb.create_sheet("Functional Tests")
        headers2 = [
            "#", "Main Menu", "Sub Menu", "URL", "Status",
            "Load(s)", "Page Title", "Form?", "Table?", "Buttons?",
            "Input Count", "Button Labels", "Error Text", "Notes",
        ]
        col_widths2 = [4, 22, 24, 50, 10, 8, 30, 7, 7, 8, 10, 35, 25, 35]

        for ci, (h, w) in enumerate(zip(headers2, col_widths2), 1):
            cell = ws2.cell(1, ci, h)
            cell.fill      = _fill(CYAN)
            cell.font      = _font(bold=True, color=WHITE, size=11)
            cell.alignment = _center()
            cell.border    = _border()
            ws2.column_dimensions[cell.column_letter].width = w

        ws2.row_dimensions[1].height = 22

        for ri, (mr, tr) in enumerate(zip(menu_rows, test_results), 2):
            status = tr["status"]
            if status == "OK":
                status_fill = _fill("FFE8F5E9")
                status_font = _font(color="FF2E7D32", bold=True)
            elif status == "TIMEOUT":
                status_fill = _fill("FFFFF8E1")
                status_font = _font(color="FFF57F17", bold=True)
            else:
                status_fill = _fill("FFFFEBEE")
                status_font = _font(color="FFC62828", bold=True)

            row_bg = LTGRAY if ri % 2 == 0 else WHITE
            values = [
                ri-1, mr["main"], mr["sub"],
                tr["url"], tr["status"], tr["load_time_sec"], tr["page_title"],
                tr["has_form"], tr["has_table"], tr["has_buttons"],
                tr["input_count"], tr["button_labels"], tr["error_text"], tr["notes"],
            ]
            for ci, val in enumerate(values, 1):
                cell = ws2.cell(ri, ci, val)
                if ci == 5:
                    cell.fill  = status_fill
                    cell.font  = status_font
                else:
                    cell.fill = _fill(row_bg)
                    cell.font = _font()
                cell.alignment = _left()
                cell.border    = _border()

        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = f"A1:N{len(test_results)+1}"

    # ── Sheet 3: Summary ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 20

    summary_data = [
        ("Report Generated",  datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Site",              BASE_URL),
        ("Total Menu Items",  len(menu_rows)),
        ("Top-Level Menus",   len({r["main"] for r in menu_rows})),
        ("Sub-Menu Items",    sum(1 for r in menu_rows if r["sub"])),
    ]
    if test_results:
        ok      = sum(1 for t in test_results if t["status"] == "OK")
        timeout = sum(1 for t in test_results if t["status"] == "TIMEOUT")
        err     = sum(1 for t in test_results if t["status"] == "ERROR")
        summary_data += [
            ("URLs Tested",       len(test_results)),
            ("✅ OK",             ok),
            ("⚠️ Timeout",        timeout),
            ("❌ Error / Other",  err),
        ]

    for ri, (label, value) in enumerate(summary_data, 1):
        ws3.cell(ri, 1, label).font  = _font(bold=True)
        ws3.cell(ri, 2, str(value))
        ws3.cell(ri, 1).alignment = ws3.cell(ri, 2).alignment = _left()

    wb.save(path)
    print(f"\n[✔] Excel report saved → {path}")


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    driver = init_driver()
    menu_rows    = []
    test_results = []

    try:
        # ── Phase 1: Login ────────────────────────────────────────────────────
        login(driver)
        current_url = driver.current_url

        # ── Phase 2: Save HTML ────────────────────────────────────────────────
        html = save_html(driver)

        # ── Phase 3: Parse from HTML (static) ────────────────────────────────
        print("\n[STEP 3] Parsing menu from HTML (static) …")
        static_rows = parse_menu_from_html(html, current_url)
        print(f"  → {len(static_rows)} items found via HTML parsing")

        # ── Phase 4: Live hover to catch JS dropdowns ─────────────────────────
        print("\n[STEP 4] Live-hover ribbon items to capture JS dropdowns …")
        live_rows = scrape_live_dropdowns(driver, current_url)
        print(f"  → {len(live_rows)} items found via live hover")

        # Merge: prefer live results if richer
        if len(live_rows) >= len(static_rows):
            menu_rows = live_rows
            print("  → Using LIVE results (richer)")
        else:
            menu_rows = static_rows
            print("  → Using STATIC parse results (richer)")

        # Deduplicate
        seen = set()
        unique = []
        for r in menu_rows:
            key = (r["main"], r["sub"])
            if key not in seen:
                seen.add(key); unique.append(r)
        menu_rows = unique
        print(f"  → {len(menu_rows)} unique menu rows after dedup")

        # ── Write CSV ─────────────────────────────────────────────────────────
        with open(MENU_CSV, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Main Menu", "Sub Menu", "URL"])
            for r in menu_rows:
                writer.writerow([r["main"], r["sub"],
                                 r.get("sub_url") or r.get("main_url","")])
        print(f"\n[✔] Menu CSV saved → {MENU_CSV}")

        # ── Phase 5: Functional Testing ───────────────────────────────────────
        do_test = input("\n[STEP 5] Run functional test on all URLs now? (y/n): ").strip().lower()
        if do_test == "y":
            print(f"\n  Testing {len(menu_rows)} URLs …")
            for idx, row in enumerate(menu_rows):
                url   = row.get("sub_url") or row.get("main_url","")
                label = f"{row['main']} > {row['sub']}" if row["sub"] else row["main"]
                print(f"  [{idx+1}/{len(menu_rows)}] {label[:60]} …", end=" ", flush=True)
                result = test_url(driver, url, label)
                test_results.append(result)
                print(result["status"])
        else:
            print("  Skipping functional tests — only menu CSV will be written.")

        # ── Phase 6: Excel report ─────────────────────────────────────────────
        write_excel(menu_rows, test_results if do_test == "y" else [], FULL_REPORT)

        print(f"""
{'='*62}
  DONE!
  • Menu CSV   : {MENU_CSV}
  • Excel      : {FULL_REPORT}
  • HTML dump  : {HTML_DUMP}
{'='*62}""")

    finally:
        input("\nPress ENTER to close browser …")
        driver.quit()


if __name__ == "__main__":
    main()