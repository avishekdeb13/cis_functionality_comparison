"""
Universal Portal Testing Tool v3.0
====================================
NEW IN v3:
  - Screen comparison: screenshots from both portals → pixel diff → highlighted diff image
  - Field extraction on LIVE portal too (read-only, no interaction)
  - Field-by-field comparison: LIVE vs TEST (missing fields, extra fields, type changes)
  - Excel report: 7 sheets including Screen Diff and Field Comparison
  - All comparison images embedded in Excel sheet

Install:
  pip install selenium webdriver-manager openpyxl beautifulsoup4 Pillow numpy scikit-image

Run:
  python portal_compare_v3.py
"""

import os, re, csv, time, sys, json
from datetime import datetime
from urllib.parse import urlparse, urljoin
from io import BytesIO

# ── Image ─────────────────────────────────────────────────────────────────────
from PIL import Image, ImageChops, ImageDraw, ImageEnhance
import numpy as np

# ── Selenium ──────────────────────────────────────────────────────────────────
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select as SeleniumSelect
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, WebDriverException)
from webdriver_manager.chrome import ChromeDriverManager

# ── BS4 ───────────────────────────────────────────────────────────────────────
from bs4 import BeautifulSoup

# ── Excel ─────────────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
#  STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def xfill(c):   return PatternFill("solid", fgColor=c)
def xfnt(bold=False, color="FF000000", sz=10):
    return Font(bold=bold, color=color, size=sz)
def xbrd():
    s = Side(style="thin", color="FFD0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)
def xaln(h="left"):
    return Alignment(horizontal=h, vertical="center", wrap_text=True)

def header_cell(ws, row, col, text, bg="FF003366", fg="FFFFFFFF", sz=11, w=None):
    c = ws.cell(row, col, text)
    c.fill = xfill(bg); c.font = xfnt(True, fg, sz)
    c.alignment = xaln("center"); c.border = xbrd()
    if w:
        ws.column_dimensions[get_column_letter(col)].width = w
    return c

def data_cell(ws, row, col, value, bg="FFFFFFFF", bold=False,
              color="FF000000", h="left"):
    c = ws.cell(row, col, value)
    c.fill = xfill(bg); c.font = xfnt(bold, color)
    c.alignment = xaln(h); c.border = xbrd()
    return c


# ─────────────────────────────────────────────────────────────────────────────
#  GUI  (tkinter — 3 tabs)
# ─────────────────────────────────────────────────────────────────────────────
def get_all_config():
    import tkinter as tk
    from tkinter import ttk, messagebox

    config = {"live": {}, "test": {}, "data": {}}

    def submit():
        for url_var, label in [(live_url_var, "Live URL"), (test_url_var, "Test URL")]:
            if not url_var.get().strip():
                messagebox.showerror("Error", f"{label} is required"); return
        for u_var, label in [(live_user_var, "Live username"), (test_user_var, "Test username")]:
            if not u_var.get().strip():
                messagebox.showerror("Error", f"{label} is required"); return
        for p_var, label in [(live_pwd_var, "Live password"), (test_pwd_var, "Test password")]:
            if not p_var.get().strip():
                messagebox.showerror("Error", f"{label} is required"); return

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        folder = (test_user_var.get().strip() or "test") + "_" + ts
        ss_dir = os.path.join(folder, "screenshots")
        diff_dir = os.path.join(folder, "diffs")
        os.makedirs(ss_dir,   exist_ok=True)
        os.makedirs(diff_dir, exist_ok=True)

        def base(url):
            p = urlparse(url.strip())
            return p.scheme + "://" + p.netloc

        config["live"] = {
            "url":      live_url_var.get().strip(),
            "username": live_user_var.get().strip(),
            "password": live_pwd_var.get().strip(),
            "base":     base(live_url_var.get()),
            "label":    "LIVE",
        }
        config["test"] = {
            "url":      test_url_var.get().strip(),
            "username": test_user_var.get().strip(),
            "password": test_pwd_var.get().strip(),
            "base":     base(test_url_var.get()),
            "folder":      folder,
            "screenshots": ss_dir,
            "diffs":       diff_dir,
            "live_html":   os.path.join(folder, "live_portal.html"),
            "test_html":   os.path.join(folder, "test_portal.html"),
            "csv":         os.path.join(folder, "menu.csv"),
            "excel":       os.path.join(folder, "comparison_report.xlsx"),
            "label":    "TEST",
        }
        config["data"] = {
            "diary_no":   diary_var.get().strip(),
            "case_no":    case_var.get().strip(),
            "from_date":  from_var.get().strip(),
            "to_date":    to_var.get().strip(),
            "party_name": party_var.get().strip(),
            "skip_deep":  skip_var.get(),
        }
        root.destroy()

    root = tk.Tk()
    root.title("Portal Comparison Tool v3.0")
    root.geometry("660x600")
    root.resizable(False, False)

    tk.Label(root, text="Portal Comparison Tool v3.0",
             font=("Arial", 14, "bold"), fg="#003366").pack(pady=8)
    tk.Label(root,
             text="Compares menus · fields · screenshots between LIVE and TEST portals",
             font=("Arial", 9), fg="#555").pack()

    nb = ttk.Notebook(root)
    nb.pack(fill="both", expand=True, padx=10, pady=8)

    def make_portal_tab(nb, title, color, note, url_default, user_default=""):
        tab = ttk.Frame(nb, padding=15)
        nb.add(tab, text=title)
        tk.Label(tab, text=note, font=("Arial", 10, "bold"),
                 fg=color).grid(row=0, column=0, columnspan=2, pady=(0,12), sticky="w")
        url_v  = tk.StringVar(value=url_default)
        user_v = tk.StringVar(value=user_default)
        pwd_v  = tk.StringVar()
        for ri, (lbl, var, show) in enumerate([
            ("Portal URL:", url_v,  ""),
            ("Username:",   user_v, ""),
            ("Password:",   pwd_v,  "*"),
        ], 1):
            ttk.Label(tab, text=lbl).grid(row=ri, column=0, sticky="w", pady=8)
            ttk.Entry(tab, textvariable=var, show=show, width=54).grid(
                row=ri, column=1, padx=8)
        return url_v, user_v, pwd_v

    live_url_var, live_user_var, live_pwd_var = make_portal_tab(
        nb, "  🔴  LIVE Portal (Read-Only)  ", "#cc0000",
        "READ-ONLY — fields extracted but nothing submitted",
        "https://cis.drt.gov.in/drtlive/index.php", "filingdrt1")

    test_url_var, test_user_var, test_pwd_var = make_portal_tab(
        nb, "  🟢  TEST Portal (Full Testing)  ", "#006600",
        "FULL TESTING — forms filled, screenshots, screen diff",
        "https://drt.etribunals.gov.in/cis2.0/filing/login")

    # Test Data tab
    tab3 = ttk.Frame(nb, padding=15)
    nb.add(tab3, text="  📋  Test Data  ")
    tk.Label(tab3, text="Used to fill forms on TEST portal only.",
             font=("Arial", 9), fg="#555").grid(
             row=0, column=0, columnspan=3, pady=(0,12), sticky="w")

    diary_var = tk.StringVar()
    case_var  = tk.StringVar()
    from_var  = tk.StringVar(value="01/01/2026")
    to_var    = tk.StringVar(value="31/12/2026")
    party_var = tk.StringVar()
    skip_var  = tk.BooleanVar(value=False)

    for ri, (lbl, var, hint) in enumerate([
        ("Diary Number:", diary_var, "e.g. 100/2026"),
        ("Case Number:",  case_var,  "e.g. OA/5/2026"),
        ("From Date:",    from_var,  "DD/MM/YYYY"),
        ("To Date:",      to_var,    "DD/MM/YYYY"),
        ("Party Name:",   party_var, "e.g. BANK OF INDIA"),
    ], 1):
        ttk.Label(tab3, text=lbl).grid(row=ri, column=0, sticky="w", pady=6)
        ttk.Entry(tab3, textvariable=var, width=30).grid(row=ri, column=1, padx=8, sticky="w")
        ttk.Label(tab3, text=hint, foreground="#999",
                  font=("Arial", 8)).grid(row=ri, column=2, sticky="w")

    ttk.Checkbutton(tab3,
                    text="Skip form filling (extract + screenshot compare only)",
                    variable=skip_var).grid(row=7, column=0, columnspan=3,
                                            pady=10, sticky="w")

    btn_frame = ttk.Frame(root)
    btn_frame.pack(pady=10)
    ttk.Button(btn_frame, text="▶  Start Comparison",
               command=submit).pack(side="left", padx=12)

    root.mainloop()
    return config


# ─────────────────────────────────────────────────────────────────────────────
#  DRIVER
# ─────────────────────────────────────────────────────────────────────────────
def init_driver():
    opts = webdriver.ChromeOptions()
    opts.add_argument("--start-maximized")
    opts.add_argument("--force-device-scale-factor=1")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=opts)


# ─────────────────────────────────────────────────────────────────────────────
#  LOGIN
# ─────────────────────────────────────────────────────────────────────────────
USER_LOCS = [
    (By.ID, "user_name"), (By.ID, "username"), (By.NAME, "user_name"),
    (By.NAME, "username"), (By.NAME, "txtUsername"), (By.ID, "txtUsername"),
    (By.XPATH, "//input[@type='text'][1]"),
    (By.XPATH, "//input[contains(@name,'user') or contains(@id,'user')]"),
]
PASS_LOCS = [
    (By.ID, "user_pass"), (By.ID, "password"), (By.NAME, "user_pass"),
    (By.NAME, "password"), (By.NAME, "txtPassword"), (By.ID, "txtPassword"),
    (By.XPATH, "//input[@type='password']"),
]

def login(driver, pcfg, ss_dir, tag):
    print(f"\n{'='*58}")
    print(f"  LOGIN → {pcfg['label']}  |  {pcfg['url']}")
    print(f"{'='*58}")
    driver.get(pcfg["url"])
    time.sleep(3)

    def fill(label, val, locs):
        for loc in locs:
            try:
                el = driver.find_element(*loc)
                if el.is_displayed():
                    el.clear(); el.send_keys(val)
                    print(f"  ✔ {label} filled")
                    return
            except NoSuchElementException:
                pass
        print(f"  ✘ {label} — fill manually in browser")

    fill("Username", pcfg["username"], USER_LOCS)
    fill("Password", pcfg["password"], PASS_LOCS)

    print(f"""
  ┌──────────────────────────────────────────────────┐
  │  ACTION: Solve CAPTCHA → click Login             │
  │  Waiting up to 120 seconds...                    │
  └──────────────────────────────────────────────────┘""")

    start = driver.current_url
    try:
        WebDriverWait(driver, 120).until(lambda d: d.current_url != start)
    except TimeoutException:
        print("  [!] URL did not change — continuing...")

    time.sleep(3)
    driver.save_screenshot(os.path.join(ss_dir, f"{tag}_00_dashboard.png"))
    print(f"  ✔ Logged in → {driver.current_url}")
    return driver.current_url


# ─────────────────────────────────────────────────────────────────────────────
#  URL CLEANER
# ─────────────────────────────────────────────────────────────────────────────
def clean_url(href, base):
    if not href or not href.strip(): return "", "Empty href"
    href = href.strip()
    if href.lower().startswith("javascript") or href in ("#", ""):
        return "", "JS link"
    if re.search(r"['\",]", href):
        clean = re.split(r"['\",]", href)[0].strip()
        if clean.startswith("/"):
            return base + clean, "BUG:broken-routerlink"
        return "", "BUG:broken-href"
    if href.startswith("http"):  return href, ""
    if href.startswith("/"):     return base + href, ""
    return urljoin(base + "/", href), "relative"


# ─────────────────────────────────────────────────────────────────────────────
#  MENU PARSER  (Bootstrap + sidebar + fallback)
# ─────────────────────────────────────────────────────────────────────────────
def parse_menu(html, base, label=""):
    soup = BeautifulSoup(html, "html.parser")
    rows = []

    def row(main, sub, href):
        url, issue = clean_url(href, base)
        return {"main": main, "sub": sub, "url": url,
                "issue": issue, "href": href}

    # Strategy 1: accordionSidebar (new portal)
    sidebar = soup.find("ul", id="accordionSidebar")
    if sidebar:
        print(f"  [{label}] Nav: accordionSidebar")
        for li in sidebar.find_all("li", class_="nav-item"):
            main_a = li.find("a", class_="nav-link")
            if not main_a: continue
            main_name = main_a.get_text(strip=True)
            if not main_name: continue
            for ci in li.find_all("div", class_="collapse-inner"):
                for a in (ci.find_all("a", class_="dropdown-item") or
                          ci.find_all("a", href=True)):
                    sub = a.get_text(strip=True)
                    if sub: rows.append(row(main_name, sub, a.get("href","")))
            if not any(r["main"] == main_name for r in rows):
                rows.append(row(main_name, "", main_a.get("href","")))
        if rows: return rows

    # Strategy 2: Bootstrap navbar-nav + dropdown-menu div (old portal / DRT live)
    nav = soup.find("nav", class_=re.compile(r"navbar", re.I))
    nav_ul = (nav.find("ul", class_=re.compile(r"navbar-nav", re.I)) if nav
              else None)
    if not nav_ul:
        for ul in soup.find_all("ul"):
            if len(ul.find_all("li", recursive=False)) >= 3:
                nav_ul = ul; break

    if nav_ul:
        print(f"  [{label}] Nav: navbar-nav+dropdown-menu")
        for li in nav_ul.find_all("li", recursive=False):
            main_a = (li.find("a", class_=re.compile(r"nav-link", re.I))
                      or li.find("a"))
            if not main_a: continue
            main_name = main_a.get_text(strip=True).lstrip("*").strip()
            if not main_name: continue
            main_href = main_a.get("href","")
            dd = (li.find("div", class_=re.compile(r"dropdown-menu", re.I)) or
                  li.find("ul",  class_=re.compile(r"dropdown-menu|sub-menu", re.I)))
            if dd:
                subs = (dd.find_all("a", class_=re.compile(r"dropdown-item", re.I))
                        or dd.find_all("a", href=True))
                added = 0
                for a in subs:
                    sub = a.get_text(strip=True)
                    if sub:
                        rows.append(row(main_name, sub, a.get("href","")))
                        added += 1
                if not added:
                    rows.append(row(main_name, "", main_href))
            else:
                rows.append(row(main_name, "", main_href))
        if rows: return rows

    # Fallback
    print(f"  [{label}] Nav: all-anchors fallback")
    for a in soup.find_all("a", href=True):
        txt = a.get_text(strip=True)
        if txt and len(txt) < 80:
            rows.append(row(txt, "", a["href"]))
    return rows


# ─────────────────────────────────────────────────────────────────────────────
#  FIELD EXTRACTOR  (works on any soup — used for LIVE and TEST)
# ─────────────────────────────────────────────────────────────────────────────
def extract_fields(soup):
    """
    Returns list of field dicts:
      { form_idx, type, name, label, mandatory, options, placeholder }
    Works read-only from HTML — never touches the browser.
    """
    fields = []

    def find_label(el, soup):
        eid = el.get("id","")
        if eid:
            lbl = soup.find("label", {"for": eid})
            if lbl:
                return lbl.get_text(strip=True).replace("*","").strip()
        parent = el.parent
        if parent:
            lbl = parent.find("label")
            if lbl:
                return lbl.get_text(strip=True).replace("*","").strip()
        return ""

    for fi, form in enumerate(soup.find_all("form"), 1):
        for inp in form.find_all("input"):
            itype = inp.get("type","text").lower()
            if itype in ("hidden",): continue
            name  = inp.get("name","") or inp.get("id","")
            lbl   = find_label(inp, soup)
            ph    = inp.get("placeholder","").replace("*","").strip()
            req   = inp.has_attr("required") or "*" in inp.get("placeholder","")
            display = lbl or ph or name or itype
            if not display or display.lower() in ("search","search for..."):
                continue
            fields.append({
                "form": fi, "type": itype.upper(),
                "name": name, "label": display,
                "mandatory": "MANDATORY" if req else "OPTIONAL",
                "options": "", "placeholder": ph,
            })

        for ta in form.find_all("textarea"):
            name = ta.get("name","") or ta.get("id","")
            lbl  = find_label(ta, soup)
            req  = ta.has_attr("required")
            display = lbl or name or "textarea"
            fields.append({
                "form": fi, "type": "TEXTAREA",
                "name": name, "label": display,
                "mandatory": "MANDATORY" if req else "OPTIONAL",
                "options": "", "placeholder": "",
            })

        for sel in form.find_all("select"):
            name = sel.get("name","") or sel.get("id","")
            lbl  = find_label(sel, soup)
            req  = sel.has_attr("required")
            opts = [o.get_text(strip=True)
                    for o in sel.find_all("option")
                    if o.get_text(strip=True)][:8]
            display = lbl or name or "select"
            fields.append({
                "form": fi, "type": "SELECT",
                "name": name, "label": display,
                "mandatory": "MANDATORY" if req else "OPTIONAL",
                "options": " / ".join(opts), "placeholder": "",
            })

        radio_groups = {}
        for r in form.find_all("input", {"type":"radio"}):
            grp = r.get("name","grp")
            lbl = find_label(r, soup) or r.get("value","")
            radio_groups.setdefault(grp, []).append(lbl)
        for grp, opts in radio_groups.items():
            fields.append({
                "form": fi, "type": "RADIO",
                "name": grp, "label": grp,
                "mandatory": "OPTIONAL",
                "options": " / ".join(opts[:6]), "placeholder": "",
            })

        for chk in form.find_all("input", {"type":"checkbox"}):
            name = chk.get("name","") or chk.get("id","")
            lbl  = find_label(chk, soup) or name
            if lbl:
                fields.append({
                    "form": fi, "type": "CHECKBOX",
                    "name": name, "label": lbl,
                    "mandatory": "OPTIONAL",
                    "options": "", "placeholder": "",
                })

        for fi_up in form.find_all("input", {"type":"file"}):
            name = fi_up.get("name","") or fi_up.get("id","") or "file"
            lbl  = find_label(fi_up, soup) or name
            req  = fi_up.has_attr("required")
            fields.append({
                "form": fi, "type": "FILE",
                "name": name, "label": lbl,
                "mandatory": "MANDATORY" if req else "OPTIONAL",
                "options": "", "placeholder": "",
            })

        for btn in form.find_all(
                lambda t: t.name == "button" or
                (t.name == "input" and
                 t.get("type","").lower() in ("submit","button","reset"))):
            btype = (btn.get("type","button") or "button").upper()
            lbl   = (btn.get_text(strip=True) or
                     btn.get("value","") or
                     btn.get("name","") or btype)
            if lbl and len(lbl) > 1:
                fields.append({
                    "form": fi, "type": f"BUTTON[{btype}]",
                    "name": lbl, "label": lbl,
                    "mandatory": "", "options": "", "placeholder": "",
                })

    return fields


# ─────────────────────────────────────────────────────────────────────────────
#  FIELD COMPARATOR
# ─────────────────────────────────────────────────────────────────────────────
def compare_fields(live_fields, test_fields, page_key):
    """
    Compare live vs test fields for one page.
    Returns list of comparison dicts.
    """
    def field_key(f):
        # Normalize label for matching
        return re.sub(r"[^a-z0-9]", "", f["label"].lower())

    live_map = {field_key(f): f for f in live_fields}
    test_map = {field_key(f): f for f in test_fields}

    live_keys = set(live_map.keys())
    test_keys = set(test_map.keys())

    results = []

    for k in sorted(live_keys | test_keys):
        lf = live_map.get(k)
        tf = test_map.get(k)

        if lf and tf:
            # Both have it — check for differences
            diffs = []
            if lf["type"] != tf["type"]:
                diffs.append(f"Type: LIVE={lf['type']} TEST={tf['type']}")
            if lf["mandatory"] != tf["mandatory"]:
                diffs.append(f"Mandatory: LIVE={lf['mandatory']} TEST={tf['mandatory']}")
            status = "MISMATCH" if diffs else "MATCH"
            results.append({
                "page":      page_key,
                "field":     lf["label"],
                "status":    status,
                "live_type": lf["type"],
                "test_type": tf["type"],
                "live_mand": lf["mandatory"],
                "test_mand": tf["mandatory"],
                "live_opts": lf["options"],
                "test_opts": tf["options"],
                "diff_notes": " | ".join(diffs),
            })
        elif lf and not tf:
            results.append({
                "page":      page_key,
                "field":     lf["label"],
                "status":    "MISSING IN TEST",
                "live_type": lf["type"],
                "test_type": "",
                "live_mand": lf["mandatory"],
                "test_mand": "",
                "live_opts": lf["options"],
                "test_opts": "",
                "diff_notes": "Field present in LIVE but absent in TEST",
            })
        else:
            results.append({
                "page":      page_key,
                "field":     tf["label"],
                "status":    "EXTRA IN TEST",
                "live_type": "",
                "test_type": tf["type"],
                "live_mand": "",
                "test_mand": tf["mandatory"],
                "live_opts": "",
                "test_opts": tf["options"],
                "diff_notes": "Field present in TEST but absent in LIVE",
            })

    return results


# ─────────────────────────────────────────────────────────────────────────────
#  SCREEN COMPARATOR
# ─────────────────────────────────────────────────────────────────────────────
def compare_screenshots(live_ss_path, test_ss_path, diff_out_path, label=""):
    """
    Compare two screenshots pixel-by-pixel.
    Returns dict: { diff_pct, diff_path, summary }
    Saves a side-by-side diff image with differences highlighted in red.
    """
    result = {
        "diff_pct":  0.0,
        "diff_path": "",
        "summary":   "No comparison",
        "status":    "NOT_COMPARED",
    }

    if not (os.path.exists(live_ss_path) and os.path.exists(test_ss_path)):
        result["summary"] = "Screenshot missing"
        return result

    try:
        live_img = Image.open(live_ss_path).convert("RGB")
        test_img = Image.open(test_ss_path).convert("RGB")

        # Resize to same dimensions (use larger)
        W = max(live_img.width,  test_img.width)
        H = max(live_img.height, test_img.height)
        live_img = live_img.resize((W, H), Image.LANCZOS)
        test_img = test_img.resize((W, H), Image.LANCZOS)

        live_arr = np.array(live_img, dtype=np.int32)
        test_arr = np.array(test_img, dtype=np.int32)

        # Pixel difference
        diff_arr   = np.abs(live_arr - test_arr)
        diff_gray  = diff_arr.max(axis=2)            # max channel diff per pixel
        threshold  = 25
        diff_mask  = diff_gray > threshold

        diff_pct   = float(diff_mask.sum()) / (W * H) * 100
        result["diff_pct"] = round(diff_pct, 2)

        # Highlight diff on test image in red
        highlighted = test_img.copy()
        hi_arr      = np.array(highlighted)
        hi_arr[diff_mask] = [220, 30, 30]           # red highlight
        highlighted = Image.fromarray(hi_arr.astype(np.uint8))

        # Side-by-side: LIVE | DIFF-HIGHLIGHTED TEST | DIFF MASK
        thumb_w = 640
        thumb_h = int(H * thumb_w / W)
        live_thumb = live_img.resize((thumb_w, thumb_h), Image.LANCZOS)
        test_thumb = highlighted.resize((thumb_w, thumb_h), Image.LANCZOS)

        # Diff heatmap
        diff_norm  = (diff_gray / diff_gray.max() * 255
                      if diff_gray.max() > 0 else diff_gray).astype(np.uint8)
        diff_hm    = Image.fromarray(diff_norm, mode="L").convert("RGB")
        hm_thumb   = diff_hm.resize((thumb_w, thumb_h), Image.LANCZOS)

        # Header bar
        bar_h = 36
        canvas = Image.new("RGB", (thumb_w * 3, thumb_h + bar_h), (30, 30, 30))
        draw   = ImageDraw.Draw(canvas)

        for i, title in enumerate(["LIVE", f"TEST  (red=diff {diff_pct:.1f}%)", "DIFF HEATMAP"]):
            x = i * thumb_w + thumb_w // 2
            draw.text((x, 10), title, fill=(255, 255, 255))

        canvas.paste(live_thumb, (0,         bar_h))
        canvas.paste(test_thumb, (thumb_w,   bar_h))
        canvas.paste(hm_thumb,   (thumb_w*2, bar_h))

        # Red border if large diff
        if diff_pct > 5:
            brd = ImageDraw.Draw(canvas)
            brd.rectangle([0,0, canvas.width-1, canvas.height-1],
                          outline=(220,30,30), width=4)

        canvas.save(diff_out_path, "PNG")
        result["diff_path"] = diff_out_path

        if   diff_pct < 1:  status = "IDENTICAL"
        elif diff_pct < 5:  status = "MINOR_DIFF"
        elif diff_pct < 20: status = "MODERATE_DIFF"
        else:               status = "MAJOR_DIFF"

        result["status"]  = status
        result["summary"] = (f"{diff_pct:.1f}% pixels differ  |  "
                             f"{diff_mask.sum():,} px changed  |  {status}")
        print(f"  Screen diff [{label}]: {result['summary']}")

    except Exception as e:
        result["summary"] = f"Error: {e}"
        result["status"]  = "ERROR"

    return result


# ─────────────────────────────────────────────────────────────────────────────
#  PAGE VISIT  (for LIVE — read-only, extract fields + screenshot only)
# ─────────────────────────────────────────────────────────────────────────────
def visit_live_page(driver, url, label, ss_path, timeout=12):
    """Visit a page on LIVE portal. Take screenshot. Extract fields. NOTHING ELSE."""
    result = {"fields": [], "title": "", "has_form": False,
              "screenshot": ss_path, "status": "OK"}
    if not url:
        result["status"] = "NO_URL"; return result
    try:
        driver.get(url)
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script("return document.readyState") == "complete")
        time.sleep(1.5)
        result["title"] = driver.title[:60]
        driver.save_screenshot(ss_path)
        soup = BeautifulSoup(driver.page_source, "html.parser")
        result["fields"]   = extract_fields(soup)
        result["has_form"] = bool(soup.find("form"))
    except TimeoutException:
        result["status"] = "TIMEOUT"
        try: driver.save_screenshot(ss_path)
        except: pass
    except WebDriverException as e:
        result["status"] = f"ERROR:{str(e)[:60]}"
    return result


# ─────────────────────────────────────────────────────────────────────────────
#  PAGE VISIT  (for TEST — full interaction + screenshot)
# ─────────────────────────────────────────────────────────────────────────────
def visit_test_page(driver, url, label, ss_path, test_data, timeout=15):
    """Visit TEST portal page. Fill forms. Screenshot. Extract fields."""
    result = {"fields": [], "title": "", "has_form": False,
              "screenshot": ss_path, "status": "OK",
              "buttons": [], "tables": 0, "deep": []}
    if not url:
        result["status"] = "NO_URL"; return result
    try:
        driver.get(url)
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script("return document.readyState") == "complete")
        time.sleep(2)
        result["title"] = driver.title[:60]
        driver.save_screenshot(ss_path)

        soup = BeautifulSoup(driver.page_source, "html.parser")
        result["fields"]   = extract_fields(soup)
        result["has_form"] = bool(soup.find("form"))
        result["tables"]   = len(soup.find_all("table"))

        # Collect button labels
        result["buttons"] = [
            (b.get_text(strip=True) or b.get("value",""))[:30]
            for b in soup.find_all(
                lambda t: t.name=="button" or
                (t.name=="input" and
                 t.get("type","").lower() in ("submit","button","reset")))
            if (b.get_text(strip=True) or b.get("value",""))
        ][:8]

        # Light interaction if not skipped
        if not test_data.get("skip_deep"):
            _light_interact(driver, test_data, result)

    except TimeoutException:
        result["status"] = "TIMEOUT"
        try: driver.save_screenshot(ss_path)
        except: pass
    except WebDriverException as e:
        result["status"] = f"ERROR:{str(e)[:60]}"
    return result


def _light_interact(driver, test_data, result):
    """Non-destructive interactions: fill search, select dropdowns."""
    try:
        # Fill search field
        search_val = test_data.get("diary_no") or test_data.get("case_no","")
        if search_val:
            for loc in [(By.CSS_SELECTOR, "input[type='text']"),
                        (By.CSS_SELECTOR, "input[name*='diary']"),
                        (By.CSS_SELECTOR, "input[name*='case']")]:
                els = driver.find_elements(*loc)
                for el in els:
                    try:
                        if el.is_displayed() and el.is_enabled():
                            el.clear(); el.send_keys(search_val)
                            result["deep"].append(f"Filled search: {search_val}")
                            break
                    except: pass
                else: continue
                break

        # Select dropdowns (first option only)
        for sel_el in driver.find_elements(By.CSS_SELECTOR, "select")[:3]:
            try:
                s    = SeleniumSelect(sel_el)
                opts = [o.text for o in s.options if o.text.strip()]
                if len(opts) > 1:
                    s.select_by_index(1)
                    result["deep"].append(
                        f"Dropdown '{sel_el.get_attribute('name') or '?'}'"
                        f" → '{opts[1][:20]}'")
                    time.sleep(0.2)
            except: pass
    except Exception as e:
        result["deep"].append(f"Interact error: {e!s:.50}")


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN COMPARISON LOOP
# ─────────────────────────────────────────────────────────────────────────────
def run_comparison(driver, live_rows, test_rows, cfg, test_data):
    """
    For each URL that appears in BOTH portals:
      1. Visit LIVE  → screenshot, extract fields
      2. Visit TEST  → screenshot, extract fields, light interact
      3. Compare screenshots pixel-by-pixel
      4. Compare fields
    Returns list of page_result dicts.
    """
    ss_dir   = cfg["test"]["screenshots"]
    diff_dir = cfg["test"]["diffs"]

    # Build URL map: key = (main.lower, sub.lower)
    def row_key(r):
        return (r["main"].strip().lower(), r["sub"].strip().lower())

    live_map = {row_key(r): r for r in live_rows}
    test_map = {row_key(r): r for r in test_rows}
    all_keys = sorted(set(list(live_map.keys()) + list(test_map.keys())))

    page_results = []
    total = len(all_keys)

    for idx, key in enumerate(all_keys):
        lr = live_map.get(key)
        tr = test_map.get(key)
        label = f"{key[0]} > {key[1]}" if key[1] else key[0]
        print(f"\n  [{idx+1:>3}/{total}] {label[:58]}")

        live_url = lr["url"] if lr else ""
        test_url = tr["url"] if tr else ""

        ss_live  = os.path.join(ss_dir,  f"{idx+1:03d}_LIVE_{re.sub(r'[^\\w]','_',label)[:35]}.png")
        ss_test  = os.path.join(ss_dir,  f"{idx+1:03d}_TEST_{re.sub(r'[^\\w]','_',label)[:35]}.png")
        ss_diff  = os.path.join(diff_dir,f"{idx+1:03d}_DIFF_{re.sub(r'[^\\w]','_',label)[:35]}.png")

        # ── Visit LIVE ────────────────────────────────────────────────────────
        live_result = {"fields":[], "title":"", "status":"NOT_IN_LIVE",
                       "screenshot": ss_live, "has_form": False}
        if live_url:
            print(f"    LIVE  → {live_url[:65]}")
            live_result = visit_live_page(driver, live_url, label, ss_live)
            print(f"    LIVE  status={live_result['status']}"
                  f"  fields={len(live_result['fields'])}"
                  f"  form={live_result['has_form']}")

        # ── Visit TEST ────────────────────────────────────────────────────────
        test_result = {"fields":[], "title":"", "status":"NOT_IN_TEST",
                       "screenshot": ss_test, "has_form": False,
                       "buttons":[], "tables":0, "deep":[]}
        if test_url:
            print(f"    TEST  → {test_url[:65]}")
            test_result = visit_test_page(driver, test_url, label,
                                          ss_test, test_data)
            print(f"    TEST  status={test_result['status']}"
                  f"  fields={len(test_result['fields'])}"
                  f"  tables={test_result['tables']}")

        # ── Screen compare ────────────────────────────────────────────────────
        diff_result = {"diff_pct": 0, "diff_path": "", "summary": "N/A",
                       "status": "NO_SCREENSHOT"}
        if live_url and test_url:
            diff_result = compare_screenshots(ss_live, ss_test, ss_diff, label)

        # ── Field compare ─────────────────────────────────────────────────────
        field_cmp = []
        if live_result["fields"] or test_result["fields"]:
            field_cmp = compare_fields(
                live_result["fields"], test_result["fields"], label)
            match_f   = sum(1 for f in field_cmp if f["status"]=="MATCH")
            miss_f    = sum(1 for f in field_cmp if f["status"]=="MISSING IN TEST")
            extra_f   = sum(1 for f in field_cmp if f["status"]=="EXTRA IN TEST")
            mis_f     = sum(1 for f in field_cmp if f["status"]=="MISMATCH")
            print(f"    Fields match={match_f} missing={miss_f}"
                  f" extra={extra_f} mismatch={mis_f}")

        page_results.append({
            "key":          key,
            "label":        label,
            "main":         key[0],
            "sub":          key[1],
            "live_url":     live_url,
            "test_url":     test_url,
            "live_status":  live_result["status"],
            "test_status":  test_result["status"],
            "live_title":   live_result["title"],
            "test_title":   test_result["title"],
            "live_fields":  live_result["fields"],
            "test_fields":  test_result["fields"],
            "field_cmp":    field_cmp,
            "diff":         diff_result,
            "ss_live":      ss_live,
            "ss_test":      ss_test,
            "ss_diff":      ss_diff if os.path.exists(ss_diff) else "",
            "test_buttons": test_result.get("buttons", []),
            "test_tables":  test_result.get("tables", 0),
            "test_deep":    test_result.get("deep", []),
        })

    return page_results


# ─────────────────────────────────────────────────────────────────────────────
#  MENU COMPARE
# ─────────────────────────────────────────────────────────────────────────────
def compare_menus(live_rows, test_rows):
    def key(r): return (r["main"].strip().lower(), r["sub"].strip().lower())
    live_keys = {key(r): r for r in live_rows}
    test_keys = {key(r): r for r in test_rows}
    all_keys  = sorted(set(list(live_keys.keys()) + list(test_keys.keys())))
    out = []
    for k in all_keys:
        lr = live_keys.get(k)
        tr = test_keys.get(k)
        if lr and tr:
            st = "MATCH"
        elif lr:
            st = "MISSING IN TEST"
        else:
            st = "EXTRA IN TEST"
        out.append({
            "main": (lr or tr)["main"], "sub": (lr or tr)["sub"],
            "status": st,
            "live_url": lr["url"] if lr else "",
            "test_url": tr["url"] if tr else "",
            "live_issue": lr.get("issue","") if lr else "",
            "test_issue": tr.get("issue","") if tr else "",
        })
    return out


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL WRITER  — 7 sheets
# ─────────────────────────────────────────────────────────────────────────────
STATUS_COLORS = {
    "MATCH":           ("FFE8F5E9", "FF1B5E20"),
    "MISSING IN TEST": ("FFFFEBEE", "FFB71C1C"),
    "EXTRA IN TEST":   ("FFFFF8E1", "FFF57F17"),
    "MISMATCH":        ("FFFFF3E0", "FFE65100"),
    "IDENTICAL":       ("FFE8F5E9", "FF1B5E20"),
    "MINOR_DIFF":      ("FFF1F8E9", "FF33691E"),
    "MODERATE_DIFF":   ("FFFFF3E0", "FFE65100"),
    "MAJOR_DIFF":      ("FFFFEBEE", "FFB71C1C"),
    "NOT_COMPARED":    ("FFFFFDE7", "FFF9A825"),
}

def _st_cells(ws, ri, ci_status, ci_end, status, vals, widths=None):
    bg, fc = STATUS_COLORS.get(status, ("FFFFFFFF","FF000000"))
    for ci, val in enumerate(vals, 1):
        c = ws.cell(ri, ci, val)
        if ci == ci_status:
            c.fill = xfill(bg); c.font = xfnt(True, fc)
        else:
            row_bg = "FFFAFAFA" if ri%2==0 else "FFFFFFFF"
            c.fill = xfill(row_bg); c.font = xfnt()
        c.alignment = xaln(); c.border = xbrd()


def write_excel(live_rows, test_rows, menu_cmp, page_results, cfg, test_data):
    print("\n[EXCEL] Writing 7-sheet report …")
    wb = Workbook()

    # ── Sheet 1: Live Menu ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Live Menu"
    cols1 = ["#","Main Menu","Sub Menu","URL","Issue"]
    wids1 = [5,  28,          32,        65,   40]
    ws1.cell(1,1,"LIVE PORTAL — READ-ONLY REFERENCE").font = xfnt(True,"FF8B0000",12)
    ws1.merge_cells("A1:E1")
    ws1.cell(1,1).alignment = xaln("center")
    ws1.cell(1,1).fill = xfill("FFFFF3CD")
    for ci,(h,w) in enumerate(zip(cols1,wids1),1):
        header_cell(ws1, 2, ci, h, "FF8B0000", w=w)
    for ri,r in enumerate(live_rows,3):
        is_sub = bool(r["sub"])
        bg = "FFFFEBEE" if r.get("issue") else ("FFFFE8E8" if not is_sub else
              "FFFAFAFA" if ri%2==0 else "FFFFFFFF")
        vals = [ri-2, r["main"], r["sub"], r["url"] or r["href"], r.get("issue","")]
        for ci,val in enumerate(vals,1):
            c = ws1.cell(ri,ci,val)
            c.fill=xfill(bg); c.border=xbrd(); c.alignment=xaln()
            c.font = (Font(color="FF8B0000",size=10,underline="single") if ci==4
                      else xfnt(bold=not is_sub and ci==2))
    ws1.freeze_panes = "A3"

    # ── Sheet 2: Test Menu ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Test Menu")
    ws2.cell(1,1,"TEST PORTAL — FULL TESTING").font = xfnt(True,"FF006600",12)
    ws2.merge_cells("A1:E1")
    ws2.cell(1,1).alignment = xaln("center"); ws2.cell(1,1).fill = xfill("FFE8F5E9")
    for ci,(h,w) in enumerate(zip(cols1,wids1),1):
        header_cell(ws2, 2, ci, h, "FF006600", w=w)
    for ri,r in enumerate(test_rows,3):
        is_sub = bool(r["sub"])
        bg = "FFFFEBEE" if r.get("issue") else ("FFE8F5E9" if not is_sub else
              "FFFAFAFA" if ri%2==0 else "FFFFFFFF")
        vals = [ri-2, r["main"], r["sub"], r["url"] or r["href"], r.get("issue","")]
        for ci,val in enumerate(vals,1):
            c = ws2.cell(ri,ci,val)
            c.fill=xfill(bg); c.border=xbrd(); c.alignment=xaln(); c.font=xfnt()
    ws2.freeze_panes = "A3"

    # ── Sheet 3: Menu Comparison ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Menu Comparison")
    cols3 = ["#","Main Menu","Sub Menu","Status","Live URL","Test URL","Live Issue","Test Issue"]
    wids3 = [4,  22,          28,        16,       52,        52,         30,           30]
    for ci,(h,w) in enumerate(zip(cols3,wids3),1):
        header_cell(ws3, 1, ci, h, "FF4A148C", w=w)
    for ri,r in enumerate(menu_cmp,2):
        _st_cells(ws3, ri, 4, 8, r["status"],
                  [ri-1,r["main"],r["sub"],r["status"],
                   r["live_url"],r["test_url"],
                   r.get("live_issue",""),r.get("test_issue","")])
    ws3.freeze_panes="A2"
    ws3.auto_filter.ref=f"A1:H{len(menu_cmp)+1}"

    # ── Sheet 4: Screen Comparison ────────────────────────────────────────────
    ws4 = wb.create_sheet("Screen Comparison")
    cols4 = ["#","Page","Live Status","Test Status","Screen Diff %","Diff Status",
             "Diff Summary","Live Title","Test Title","Diff Image Path"]
    wids4 = [4,  38,    14,           14,            14,             16,
             55,          30,          30,             45]
    for ci,(h,w) in enumerate(zip(cols4,wids4),1):
        header_cell(ws4, 1, ci, h, "FF1A237E", w=w)

    # Row height for diff images
    ws4.sheet_properties.outlinePr.summaryBelow = False
    img_row_map = {}  # ri → diff_path

    for ri,pr in enumerate(page_results,2):
        diff = pr["diff"]
        st   = diff.get("status","NOT_COMPARED")
        vals = [ri-1, pr["label"],
                pr["live_status"], pr["test_status"],
                diff.get("diff_pct",""), st,
                diff.get("summary",""), pr["live_title"],
                pr["test_title"], diff.get("diff_path","")]
        _st_cells(ws4, ri, 6, 10, st, vals)
        if diff.get("diff_path") and os.path.exists(diff["diff_path"]):
            img_row_map[ri] = diff["diff_path"]

    ws4.freeze_panes="A2"

    # ── Sheet 5: Screen Diff Images ───────────────────────────────────────────
    ws5 = wb.create_sheet("Screen Diff Images")
    ws5.column_dimensions["A"].width = 20
    ws5.column_dimensions["B"].width = 130

    ws5.cell(1,1,"Page").font  = xfnt(True,"FFFFFFFF",11)
    ws5.cell(1,1).fill         = xfill("FF1A237E")
    ws5.cell(1,2,"Side-by-Side Diff  (LIVE | TEST with red diff | HEATMAP)").font = xfnt(True,"FFFFFFFF",11)
    ws5.cell(1,2).fill         = xfill("FF1A237E")
    ws5.row_dimensions[1].height = 20

    img_row = 2
    inserted = 0
    for pr in page_results:
        diff_path = pr.get("ss_diff","")
        if not diff_path or not os.path.exists(diff_path):
            continue
        try:
            ws5.cell(img_row, 1, pr["label"]).font = xfnt(bold=True)
            ws5.cell(img_row, 1).alignment = xaln("left")

            img_pil = Image.open(diff_path)
            # Resize to ~1400px wide for readability in Excel
            max_w   = 1400
            ratio   = min(1.0, max_w / img_pil.width)
            new_w   = int(img_pil.width  * ratio)
            new_h   = int(img_pil.height * ratio)
            img_pil = img_pil.resize((new_w, new_h), Image.LANCZOS)

            buf = BytesIO()
            img_pil.save(buf, format="PNG")
            buf.seek(0)

            xl_img = XLImage(buf)
            xl_img.anchor = f"B{img_row}"
            ws5.add_image(xl_img)

            # Adjust row height (~0.75 pt per pixel)
            row_h = max(60, int(new_h * 0.75))
            ws5.row_dimensions[img_row].height = row_h

            img_row += 1
            inserted += 1
        except Exception as e:
            ws5.cell(img_row, 1, pr["label"])
            ws5.cell(img_row, 2, f"Image error: {e}")
            img_row += 1

    print(f"  ✔ {inserted} diff images embedded in 'Screen Diff Images' sheet")

    # ── Sheet 6: Field Comparison ─────────────────────────────────────────────
    ws6 = wb.create_sheet("Field Comparison")
    cols6 = ["#","Page","Field Name/Label","Status",
             "Live Type","Test Type","Live Mandatory","Test Mandatory",
             "Live Options","Test Options","Notes"]
    wids6 = [4,  30,    38,               16,
             14,         14,          16,              16,
             35,           35,           45]
    for ci,(h,w) in enumerate(zip(cols6,wids6),1):
        header_cell(ws6, 1, ci, h, "FF006064", w=w)

    field_ri = 2
    for pr in page_results:
        for fc in pr["field_cmp"]:
            st   = fc["status"]
            bg,col = STATUS_COLORS.get(st, ("FFFFFFFF","FF000000"))
            row_bg = "FFFAFAFA" if field_ri%2==0 else "FFFFFFFF"
            vals = [field_ri-1, pr["label"], fc["field"], st,
                    fc["live_type"], fc["test_type"],
                    fc["live_mand"], fc["test_mand"],
                    fc["live_opts"], fc["test_opts"],
                    fc["diff_notes"]]
            for ci,val in enumerate(vals,1):
                c = ws6.cell(field_ri, ci, val)
                if ci==4:
                    c.fill=xfill(bg); c.font=xfnt(True,col)
                else:
                    c.fill=xfill(row_bg); c.font=xfnt()
                c.alignment=xaln(); c.border=xbrd()
            field_ri += 1

    ws6.freeze_panes="A2"
    ws6.auto_filter.ref=f"A1:K{field_ri}"

    # ── Sheet 7: Summary ──────────────────────────────────────────────────────
    ws7 = wb.create_sheet("Summary")
    ws7.column_dimensions["A"].width=34
    ws7.column_dimensions["B"].width=40

    # Menu stats
    menu_match   = sum(1 for r in menu_cmp if r["status"]=="MATCH")
    menu_miss    = sum(1 for r in menu_cmp if r["status"]=="MISSING IN TEST")
    menu_extra   = sum(1 for r in menu_cmp if r["status"]=="EXTRA IN TEST")

    # Screen stats
    screen_ident = sum(1 for pr in page_results if pr["diff"]["status"]=="IDENTICAL")
    screen_minor = sum(1 for pr in page_results if pr["diff"]["status"]=="MINOR_DIFF")
    screen_mod   = sum(1 for pr in page_results if pr["diff"]["status"]=="MODERATE_DIFF")
    screen_major = sum(1 for pr in page_results if pr["diff"]["status"]=="MAJOR_DIFF")

    # Field stats
    all_fc    = [fc for pr in page_results for fc in pr["field_cmp"]]
    f_match   = sum(1 for f in all_fc if f["status"]=="MATCH")
    f_miss    = sum(1 for f in all_fc if f["status"]=="MISSING IN TEST")
    f_extra   = sum(1 for f in all_fc if f["status"]=="EXTRA IN TEST")
    f_mis     = sum(1 for f in all_fc if f["status"]=="MISMATCH")

    summary = [
        ("Report Generated",    datetime.now().strftime("%Y-%m-%d %H:%M:%S"), None),
        ("","",""),
        ("── PORTALS ──","",""),
        ("Live Portal",         cfg["live"]["url"],      None),
        ("Live Username",       cfg["live"]["username"], None),
        ("Test Portal",         cfg["test"]["url"],      None),
        ("Test Username",       cfg["test"]["username"], None),
        ("Output Folder",       cfg["test"]["folder"],   None),
        ("","",""),
        ("── MENU COMPARISON ──","",""),
        ("Live Menu Items",     len(live_rows),   None),
        ("Test Menu Items",     len(test_rows),   None),
        ("Matching Menus",      menu_match,       "FF1B5E20"),
        ("Missing in Test",     menu_miss,        "FFB71C1C"),
        ("Extra in Test",       menu_extra,       "FFF57F17"),
        ("","",""),
        ("── SCREEN COMPARISON ──","",""),
        ("Pages Compared",      len(page_results),None),
        ("Identical",           screen_ident,     "FF1B5E20"),
        ("Minor Diff (<5%)",    screen_minor,     "FF33691E"),
        ("Moderate (5-20%)",    screen_mod,       "FFE65100"),
        ("Major Diff (>20%)",   screen_major,     "FFB71C1C"),
        ("","",""),
        ("── FIELD COMPARISON ──","",""),
        ("Total Fields Compared",len(all_fc),     None),
        ("Matching Fields",     f_match,          "FF1B5E20"),
        ("Missing in Test",     f_miss,           "FFB71C1C"),
        ("Extra in Test",       f_extra,          "FFF57F17"),
        ("Mismatched Fields",   f_mis,            "FFE65100"),
    ]

    for ri,(lbl,val,color) in enumerate(summary,1):
        l = ws7.cell(ri,1,lbl)
        v = ws7.cell(ri,2,str(val))
        l.font = xfnt(bold="──" in lbl,sz=11)
        v.font = xfnt(color=color,sz=12,bold=bool(color)) if color else xfnt(sz=11)
        l.alignment=v.alignment=xaln()

    wb.save(cfg["test"]["excel"])
    print(f"\n  ✔ Excel saved → {os.path.abspath(cfg['test']['excel'])}")
    print(f"  Sheets: Live Menu | Test Menu | Menu Comparison | "
          f"Screen Comparison | Screen Diff Images | Field Comparison | Summary")


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    config    = get_all_config()
    if not config["live"] or not config["test"]:
        print("Cancelled."); return

    live_cfg  = config["live"]
    test_cfg  = config["test"]
    test_data = config["data"]

    print(f"""
{'='*60}
  LIVE  : {live_cfg['url']}  [{live_cfg['username']}]  READ ONLY
  TEST  : {test_cfg['url']}  [{test_cfg['username']}]  FULL TEST
  Folder: {test_cfg['folder']}
  Deep  : {'NO' if test_data.get('skip_deep') else 'YES'}
{'='*60}""")

    driver     = init_driver()
    live_rows  = []
    test_rows  = []

    try:
        # ── PHASE 1: LIVE — extract menus (read-only) ─────────────────────────
        print("\n" + "="*60)
        print("  PHASE 1: LIVE PORTAL — READ ONLY")
        print("="*60)
        login(driver, live_cfg, test_cfg["screenshots"], "LIVE")

        # Open all dropdowns via CSS injection (non-destructive)
        driver.execute_script("""
            document.querySelectorAll('.dropdown-menu,.collapse').forEach(
                el=>{el.classList.add('show');el.style.display='block'});
            document.querySelectorAll('[aria-expanded]').forEach(
                el=>el.setAttribute('aria-expanded','true'));
        """)
        time.sleep(2)
        live_html = driver.page_source
        with open(test_cfg["live_html"], "w", encoding="utf-8") as f:
            f.write(live_html)
        live_rows = parse_menu(live_html, live_cfg["base"], "LIVE")
        print(f"  ✔ Live menu: {len(live_rows)} items")

        driver.delete_all_cookies()
        time.sleep(2)

        # ── PHASE 2: TEST — extract menus ────────────────────────────────────
        print("\n" + "="*60)
        print("  PHASE 2: TEST PORTAL — FULL TESTING")
        print("="*60)
        login(driver, test_cfg, test_cfg["screenshots"], "TEST")

        driver.execute_script("""
            document.querySelectorAll('.collapse,[data-bs-toggle],.dropdown-menu')
                .forEach(el=>{el.classList.add('show');el.style.display='block'});
        """)
        time.sleep(2)
        test_html = driver.page_source
        with open(test_cfg["test_html"], "w", encoding="utf-8") as f:
            f.write(test_html)
        test_rows = parse_menu(test_html, test_cfg["base"], "TEST")
        print(f"  ✔ Test menu: {len(test_rows)} items")

        # ── PHASE 3: Menu comparison ──────────────────────────────────────────
        menu_cmp = compare_menus(live_rows, test_rows)
        match_m  = sum(1 for c in menu_cmp if c["status"]=="MATCH")
        miss_m   = sum(1 for c in menu_cmp if c["status"]=="MISSING IN TEST")
        print(f"\n  Menu: {match_m} match | {miss_m} missing in test"
              f" | {len(menu_cmp)-match_m-miss_m} extra in test")

        # ── PHASE 4: Page-by-page comparison ─────────────────────────────────
        print(f"\n{'='*60}")
        print("  PHASE 3: PAGE COMPARISON (screens + fields)")
        print(f"{'='*60}")
        page_results = run_comparison(
            driver, live_rows, test_rows, config, test_data)

        # ── PHASE 5: Write Excel ──────────────────────────────────────────────
        write_excel(live_rows, test_rows, menu_cmp,
                    page_results, config, test_data)

        print(f"""
{'='*60}
  DONE!
  Folder : {test_cfg['folder']}
  Excel  : {test_cfg['excel']}
  Diffs  : {test_cfg['diffs']}
{'='*60}""")

    except KeyboardInterrupt:
        print("\n[!] Interrupted.")
    except Exception as e:
        print(f"\n[ERROR] {e}")
        import traceback; traceback.print_exc()
    finally:
        input("\nPress ENTER to close browser …")
        driver.quit()


if __name__ == "__main__":
    main()