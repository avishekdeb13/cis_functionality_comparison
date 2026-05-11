"""
Universal Web Portal Menu Scraper & Tester
==========================================
A fully interactive tool that works on ANY website.

Features:
  - Interactive CLI for all inputs (URL, credentials, options)
  - Auto-detects nav structure (Bootstrap, custom ul/li, table-based)
  - Saves HTML dump, CSV, and color-coded Excel report
  - Functional URL testing (forms, tables, buttons, errors)
  - Supports multiple sites in one session for comparison
  - Config save/load so you don't re-type credentials

Install:
  pip install selenium webdriver-manager openpyxl beautifulsoup4

Run:
  python universal_scraper.py
"""

import os, csv, re, sys, json, time, getpass
from datetime import datetime
from urllib.parse import urljoin, urlparse

from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, WebDriverException,
)
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

CONFIG_FILE = "scraper_configs.json"

# ─────────────────────────────────────────────────────────────────────────────
#  TERMINAL HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def clr(text, code): return f"\033[{code}m{text}\033[0m"
def bold(t):   return clr(t, "1")
def green(t):  return clr(t, "32")
def yellow(t): return clr(t, "33")
def red(t):    return clr(t, "31")
def cyan(t):   return clr(t, "36")
def dim(t):    return clr(t, "2")

def hr(char="─", width=64): print(dim(char * width))

def header(title):
    hr("═")
    print(bold(f"  {title}"))
    hr("═")

def section(title):
    print()
    hr()
    print(cyan(f"  {title}"))
    hr()

def ask(prompt, default=None, secret=False):
    """Prompt user, show default, return value."""
    hint = f" {dim('[' + str(default) + ']')}" if default is not None else ""
    full_prompt = f"  {prompt}{hint}: "
    if secret:
        val = getpass.getpass(full_prompt)
    else:
        val = input(full_prompt).strip()
    if val == "" and default is not None:
        return str(default)
    return val

def ask_int(prompt, default=None, min_val=None, max_val=None):
    while True:
        raw = ask(prompt, default)
        try:
            n = int(raw)
            if min_val is not None and n < min_val:
                print(red(f"  ✘ Must be ≥ {min_val}"))
                continue
            if max_val is not None and n > max_val:
                print(red(f"  ✘ Must be ≤ {max_val}"))
                continue
            return n
        except ValueError:
            print(red("  ✘ Please enter a number"))

def ask_yes(prompt, default="y"):
    val = ask(f"{prompt} (y/n)", default).lower()
    return val.startswith("y")

def pick(prompt, options, default=1):
    """Numbered menu picker. Returns chosen item."""
    print(f"\n  {prompt}")
    for i, opt in enumerate(options, 1):
        marker = green("▶") if i == default else " "
        print(f"  {marker} {i}. {opt}")
    idx = ask_int("  Enter number", default, 1, len(options))
    return options[idx - 1]


# ─────────────────────────────────────────────────────────────────────────────
#  CONFIG SAVE / LOAD
# ─────────────────────────────────────────────────────────────────────────────
def load_saved_configs():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r") as f:
                return json.load(f)
        except Exception:
            pass
    return []

def save_config(cfg):
    configs = load_saved_configs()
    # Update existing or append
    for i, c in enumerate(configs):
        if c.get("url") == cfg["url"] and c.get("username") == cfg["username"]:
            configs[i] = cfg
            break
    else:
        configs.append(cfg)
    try:
        with open(CONFIG_FILE, "w") as f:
            json.dump(configs[-10:], f, indent=2)  # keep last 10
    except Exception:
        pass

def domain_prefix(url):
    try:
        parts = urlparse(url)
        return re.sub(r"[^a-zA-Z0-9]", "_", parts.netloc)[:30]
    except Exception:
        return "site"


# ─────────────────────────────────────────────────────────────────────────────
#  USER INPUT WIZARD
# ─────────────────────────────────────────────────────────────────────────────
def get_config():
    header("Universal Web Portal Scraper & Tester")
    print(dim("  Scrapes navigation menus, tests URLs, exports Excel/CSV"))
    print()

    saved = load_saved_configs()

    # ── Choose saved or new ───────────────────────────────────────────────────
    cfg = {}
    if saved:
        choices = [f"{c.get('url','?')}  ({c.get('username','?')})" for c in saved]
        choices = ["Enter new site details"] + choices
        choice = pick("Load a saved site or enter new?", choices, default=1)
        if choice != "Enter new site details":
            idx = choices.index(choice) - 1
            cfg = dict(saved[idx])
            print(green(f"\n  ✔ Loaded config for: {cfg['url']}"))
            if not ask_yes("  Use these saved credentials?", "y"):
                cfg = {}

    # ── URL ───────────────────────────────────────────────────────────────────
    section("SITE DETAILS")
    if not cfg.get("url"):
        raw_url = ask("Site URL (include https://)",
                      "https://cis.drt.gov.in/drtlive/index.php")
        # ensure scheme
        if not raw_url.startswith("http"):
            raw_url = "https://" + raw_url
        cfg["url"] = raw_url.rstrip("/")

    # ── Credentials ───────────────────────────────────────────────────────────
    if not cfg.get("username"):
        cfg["username"] = ask("Username", "filingdrt1")

    change_pass = not cfg.get("password") or ask_yes("  Change password?", "n")
    if change_pass:
        cfg["password"] = ask("Password (hidden)", secret=True) or cfg.get("password","")

    # ── CAPTCHA wait ──────────────────────────────────────────────────────────
    section("LOGIN OPTIONS")
    cfg["captcha_wait"] = ask_int(
        "Seconds to wait for manual CAPTCHA + login",
        default=cfg.get("captcha_wait", 90), min_val=10, max_val=300
    )

    login_indicator = ask(
        "Text/URL fragment that confirms successful login\n"
        "  (leave blank to rely on URL change only)",
        default=cfg.get("login_indicator", "")
    )
    cfg["login_indicator"] = login_indicator

    # ── Nav selector hint ─────────────────────────────────────────────────────
    section("NAV DETECTION")
    nav_modes = [
        "Auto-detect (recommended)",
        "Bootstrap navbar (nav.navbar > ul.navbar-nav)",
        "Standard UL/LI dropdown",
        "Table-based nav",
        "All anchor tags (fallback)",
    ]
    cfg["nav_mode"] = pick("Navigation structure", nav_modes,
                           default=nav_modes.index(cfg.get("nav_mode", nav_modes[0])) + 1)

    # ── Output options ────────────────────────────────────────────────────────
    section("OUTPUT OPTIONS")
    prefix = domain_prefix(cfg["url"])
    cfg["output_dir"] = ask("Output folder", default=cfg.get("output_dir", "."))
    cfg["prefix"]     = ask("File prefix", default=cfg.get("prefix", prefix))

    # Derived filenames
    base = os.path.join(cfg["output_dir"], cfg["prefix"])
    cfg["html_file"]  = base + "_page.html"
    cfg["csv_file"]   = base + "_menu.csv"
    cfg["excel_file"] = base + "_report.xlsx"

    os.makedirs(cfg["output_dir"], exist_ok=True)

    # ── Testing options ───────────────────────────────────────────────────────
    section("URL TESTING")
    cfg["run_tests"] = ask_yes("Run functional URL tests after scraping?",
                               "y" if cfg.get("run_tests", True) else "n")
    if cfg["run_tests"]:
        cfg["test_pause"] = ask_int(
            "Pause between URL tests (seconds, 0-5)",
            default=cfg.get("test_pause", 1), min_val=0, max_val=10
        )
        cfg["skip_no_url"] = ask_yes("Skip rows with no URL?",
                                     "y" if cfg.get("skip_no_url", True) else "n")

    # ── Headless browser ─────────────────────────────────────────────────────
    section("BROWSER OPTIONS")
    cfg["headless"] = ask_yes(
        "Run browser in headless mode?\n"
        "  (Say NO — you need to see the CAPTCHA)",
        "n"
    )
    cfg["page_timeout"] = ask_int(
        "Page load timeout (seconds)",
        default=cfg.get("page_timeout", 20), min_val=5, max_val=120
    )

    # ── Save config ───────────────────────────────────────────────────────────
    section("SAVE CONFIG")
    if ask_yes("Save this config for next time?", "y"):
        save_config({k: v for k, v in cfg.items()
                     if k not in ("html_file","csv_file","excel_file")})
        print(green(f"  ✔ Config saved to {CONFIG_FILE}"))

    # ── Confirm ───────────────────────────────────────────────────────────────
    section("READY TO RUN")
    print(f"  {'URL':<20} {cfg['url']}")
    print(f"  {'Username':<20} {cfg['username']}")
    print(f"  {'Password':<20} {'*' * len(cfg['password'])}")
    print(f"  {'CAPTCHA wait':<20} {cfg['captcha_wait']}s")
    print(f"  {'Nav mode':<20} {cfg['nav_mode']}")
    print(f"  {'Run URL tests':<20} {cfg['run_tests']}")
    print(f"  {'Output dir':<20} {cfg['output_dir']}")
    print(f"  {'Excel file':<20} {cfg['excel_file']}")
    print()
    if not ask_yes("Start scraping now?", "y"):
        print(yellow("  Aborted."))
        sys.exit(0)

    return cfg


# ─────────────────────────────────────────────────────────────────────────────
#  DRIVER
# ─────────────────────────────────────────────────────────────────────────────
def init_driver(cfg):
    opts = webdriver.ChromeOptions()
    if cfg.get("headless"):
        opts.add_argument("--headless=new")
        opts.add_argument("--window-size=1920,1080")
    else:
        opts.add_argument("--start-maximized")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument("--disable-popup-blocking")
    opts.add_argument("--disable-notifications")

    try:
        from webdriver_manager.chrome import ChromeDriverManager
        from selenium.webdriver.chrome.service import Service
        driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()), options=opts)
    except Exception:
        driver = webdriver.Chrome(options=opts)

    driver.set_page_load_timeout(cfg.get("page_timeout", 20) + 10)
    return driver


# ─────────────────────────────────────────────────────────────────────────────
#  LOGIN
# ─────────────────────────────────────────────────────────────────────────────
USERNAME_LOCATORS = [
    (By.NAME, "username"), (By.NAME, "user_name"), (By.NAME, "txtUsername"),
    (By.NAME, "loginid"),  (By.NAME, "login"),     (By.NAME, "email"),
    (By.ID,   "username"), (By.ID,   "txtUsername"),(By.ID,   "loginid"),
    (By.ID,   "email"),    (By.ID,   "user"),
    (By.XPATH, "//input[@type='text'][1]"),
    (By.XPATH, "//input[contains(@name,'user') or contains(@placeholder,'user') or contains(@id,'user')]"),
    (By.XPATH, "//input[contains(@name,'login') or contains(@placeholder,'login')]"),
]
PASSWORD_LOCATORS = [
    (By.NAME, "password"), (By.NAME, "txtPassword"), (By.NAME, "pwd"),
    (By.NAME, "pass"),     (By.NAME, "passwd"),
    (By.ID,   "password"), (By.ID,   "txtPassword"), (By.ID,   "pwd"),
    (By.XPATH, "//input[@type='password']"),
]

def login(driver, cfg):
    section("LOGIN")
    print(f"  Opening: {cfg['url']}")
    driver.get(cfg["url"])
    time.sleep(3)

    def fill(label, value, locators):
        for loc in locators:
            try:
                el = driver.find_element(*loc)
                if el.is_displayed():
                    el.clear()
                    el.send_keys(value)
                    print(green(f"  ✔ {label} filled  ({loc[1]})"))
                    return True
            except NoSuchElementException:
                pass
        print(yellow(f"  ✘ {label} field not found — fill manually in browser"))
        return False

    fill("Username", cfg["username"], USERNAME_LOCATORS)
    fill("Password", cfg["password"], PASSWORD_LOCATORS)

    print(f"""
  {yellow('ACTION REQUIRED')}
  ┌─────────────────────────────────────────────┐
  │  1. Look at the browser window              │
  │  2. Solve the CAPTCHA                       │
  │  3. Click Login / Submit                    │
  │  You have {cfg['captcha_wait']} seconds              │
  └─────────────────────────────────────────────┘""")

    start_url = driver.current_url
    indicator = cfg.get("login_indicator", "")

    def logged_in(d):
        if indicator and indicator in d.current_url:
            return True
        if d.current_url != start_url:
            return True
        return False

    try:
        WebDriverWait(driver, cfg["captcha_wait"]).until(logged_in)
        print(green(f"\n  ✔ Login detected  →  {driver.current_url}"))
    except TimeoutException:
        print(yellow("\n  [!] Timeout waiting for login — continuing anyway …"))

    time.sleep(4)
    return driver.current_url


# ─────────────────────────────────────────────────────────────────────────────
#  HTML DUMP
# ─────────────────────────────────────────────────────────────────────────────
def save_html(driver, path):
    html = driver.page_source
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    size = os.path.getsize(path)
    print(green(f"  ✔ HTML saved → {path}  ({size:,} bytes)"))
    return html


# ─────────────────────────────────────────────────────────────────────────────
#  MENU PARSER  (multi-strategy)
# ─────────────────────────────────────────────────────────────────────────────
def abs_url(href, base):
    if not href: return ""
    href = href.strip()
    if href in ("#", "javascript:void(0)", "javascript:;", ""): return ""
    return urljoin(base, href)

def _bootstrap_nav(soup, base):
    """Bootstrap 4/5: nav.navbar > ul.navbar-nav > li.nav-item"""
    rows = []
    nav = soup.find("nav", class_=re.compile(r"navbar", re.I)) or soup
    ul  = nav.find("ul",  class_=re.compile(r"navbar-nav", re.I))
    if not ul:
        return rows
    for li in ul.find_all("li", recursive=False):
        main_a   = li.find("a", class_=re.compile(r"nav-link", re.I)) or li.find("a")
        if not main_a: continue
        main_txt = main_a.get_text(strip=True).lstrip("*").strip()
        main_url = abs_url(main_a.get("href",""), base)
        dropdown = li.find("div", class_=re.compile(r"dropdown-menu", re.I))
        if dropdown:
            subs = dropdown.find_all("a", class_=re.compile(r"dropdown-item", re.I))
            if not subs: subs = dropdown.find_all("a", href=True)
            for a in subs:
                sub_txt = a.get_text(strip=True)
                sub_url = abs_url(a.get("href",""), base)
                if sub_txt:
                    rows.append({"main": main_txt, "main_url": main_url,
                                 "sub": sub_txt, "sub_url": sub_url})
            if not subs:
                rows.append({"main": main_txt, "main_url": main_url,
                             "sub": "", "sub_url": main_url})
        else:
            if main_txt:
                rows.append({"main": main_txt, "main_url": main_url,
                             "sub": "", "sub_url": main_url})
    return rows

def _ul_li_nav(soup, base):
    """Generic UL/LI with nested UL dropdowns."""
    rows = []
    best_ul, best_n = None, 0
    for ul in soup.find_all("ul"):
        n = len(ul.find_all("li", recursive=False))
        if n > best_n:
            best_n = n; best_ul = ul
    if not best_ul or best_n < 3:
        return rows
    for li in best_ul.find_all("li", recursive=False):
        main_a   = li.find("a", recursive=False) or li.find("a")
        if not main_a: continue
        main_txt = main_a.get_text(strip=True)
        main_url = abs_url(main_a.get("href",""), base)
        sub_ul   = li.find("ul")
        if sub_ul:
            for sub_li in sub_ul.find_all("li"):
                sub_a   = sub_li.find("a")
                sub_txt = sub_a.get_text(strip=True) if sub_a else ""
                sub_url = abs_url(sub_a.get("href","") if sub_a else "", base)
                if sub_txt:
                    rows.append({"main": main_txt, "main_url": main_url,
                                 "sub": sub_txt, "sub_url": sub_url})
        else:
            if main_txt:
                rows.append({"main": main_txt, "main_url": main_url,
                             "sub": "", "sub_url": main_url})
    return rows

def _table_nav(soup, base):
    """Table-based nav (older gov portals)."""
    rows = []
    for table in soup.find_all("table"):
        links = table.find_all("a", href=True)
        if len(links) >= 4:
            for a in links:
                txt = a.get_text(strip=True)
                url = abs_url(a["href"], base)
                if txt:
                    rows.append({"main": txt, "main_url": url,
                                 "sub": "", "sub_url": url})
            if rows:
                return rows
    return rows

def _all_anchors(soup, base):
    """Fallback: every anchor on the page."""
    rows = []
    for a in soup.find_all("a", href=True):
        txt = a.get_text(strip=True)
        url = abs_url(a["href"], base)
        if txt and url and len(txt) < 80:
            rows.append({"main": txt, "main_url": url,
                         "sub": "", "sub_url": url})
    return rows

def parse_menu(html, base_url, nav_mode):
    soup = BeautifulSoup(html, "html.parser")
    mode = nav_mode

    strategy_map = {
        "Bootstrap navbar (nav.navbar > ul.navbar-nav)": [_bootstrap_nav],
        "Standard UL/LI dropdown":                      [_ul_li_nav],
        "Table-based nav":                               [_table_nav],
        "All anchor tags (fallback)":                   [_all_anchors],
    }

    if mode == "Auto-detect (recommended)":
        strategies = [_bootstrap_nav, _ul_li_nav, _table_nav, _all_anchors]
    else:
        strategies = strategy_map.get(mode, [_bootstrap_nav, _ul_li_nav,
                                              _table_nav, _all_anchors])

    for fn in strategies:
        rows = fn(soup, base_url)
        if rows:
            print(green(f"  ✔ [{fn.__name__}] found {len(rows)} menu rows"))
            return rows
        else:
            print(dim(f"  – [{fn.__name__}] found nothing"))

    return []


# ─────────────────────────────────────────────────────────────────────────────
#  CSV
# ─────────────────────────────────────────────────────────────────────────────
def save_csv(rows, path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Main Menu", "Sub Menu", "URL"])
        for r in rows:
            w.writerow([r["main"], r["sub"], r["sub_url"] or r["main_url"]])
    print(green(f"  ✔ CSV saved → {path}  ({len(rows)} rows)"))


# ─────────────────────────────────────────────────────────────────────────────
#  URL TESTER
# ─────────────────────────────────────────────────────────────────────────────
def test_url(driver, url, pause=1):
    r = dict(url=url, status="NOT_TESTED", load_sec="", page_title="",
             has_form="", form_count="", has_table="", has_buttons="",
             input_count="", button_labels="", select_count="",
             file_upload="", error_on_page="", page_notes="")
    if not url:
        r["status"] = "NO_URL"; return r
    try:
        t0 = time.time()
        driver.get(url)
        time.sleep(pause)
        r["load_sec"]   = round(time.time() - t0, 2)
        r["page_title"] = driver.title.strip()[:80]

        soup = BeautifulSoup(driver.page_source, "html.parser")
        body = soup.get_text(separator=" ", strip=True).lower()[:3000]

        forms = soup.find_all("form")
        r["has_form"]    = "YES" if forms else "NO"
        r["form_count"]  = len(forms)
        r["input_count"] = sum(len(f.find_all(["input","textarea"])) for f in forms)
        r["select_count"]= len(soup.find_all("select"))
        r["file_upload"] = "YES" if soup.find("input", {"type":"file"}) else "NO"
        r["has_table"]   = "YES" if soup.find("table") else "NO"

        btns = soup.find_all(
            lambda t: t.name == "button" or
                      (t.name == "input" and
                       t.get("type","").lower() in ("submit","button","reset"))
        )
        r["has_buttons"]   = "YES" if btns else "NO"
        r["button_labels"] = " | ".join(
            (b.get_text(strip=True) or b.get("value","") or b.get("name",""))[:25]
            for b in btns[:6])

        err_kw = ["error","invalid","access denied","not found",
                  "unauthorized","exception","500","403","404"]
        found  = [k for k in err_kw if k in body]
        r["error_on_page"] = ", ".join(found) if found else ""

        notes = []
        if forms:                   notes.append(f"{len(forms)} form(s)")
        if soup.find("table"):      notes.append("data table")
        if r["select_count"]:       notes.append(f"{r['select_count']} dropdown(s)")
        if r["file_upload"]=="YES": notes.append("file upload")
        if soup.find(attrs={"class": re.compile(r"alert|success|danger", re.I)}):
            notes.append("alert box")
        r["page_notes"] = " | ".join(notes)
        r["status"] = "OK"

    except TimeoutException:
        r["status"] = "TIMEOUT"
    except WebDriverException as e:
        r["status"] = "ERROR"; r["error_on_page"] = str(e)[:120]
    return r


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL WRITER
# ─────────────────────────────────────────────────────────────────────────────
def _fill_c(h): return PatternFill("solid", fgColor=h)
def _fnt(bold=False, color="FF000000", sz=10): return Font(bold=bold, color=color, size=sz)
def _brd():
    s = Side(style="thin", color="FFD0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)
def _aln(h="left"): return Alignment(horizontal=h, vertical="center", wrap_text=True)

def write_excel(menu_rows, test_results, cfg):
    wb = Workbook()

    # ── Sheet 1: Menu Hierarchy ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Menu Hierarchy"
    h1 = ["#","Main Menu","Sub Menu","URL","Type"]
    w1 = [5,  28,          32,        65,   12]
    for ci,(h,w) in enumerate(zip(h1,w1),1):
        c = ws1.cell(1,ci,h)
        c.fill=_fill_c("FF0277BD"); c.font=_fnt(True,"FFFFFFFF",11)
        c.alignment=_aln("center"); c.border=_brd()
        ws1.column_dimensions[c.column_letter].width=w
    ws1.row_dimensions[1].height=22

    for ri,row in enumerate(menu_rows,2):
        is_sub = bool(row["sub"])
        url    = row["sub_url"] or row["main_url"]
        vals   = [ri-1, row["main"], row["sub"], url,
                  "Sub-item" if is_sub else "Top-level"]
        bg = "FFF0F4F8" if ri%2==0 else "FFFAFAFA"
        row_bg = bg if is_sub else "FFE1F5FE"
        for ci,val in enumerate(vals,1):
            c = ws1.cell(ri,ci,val)
            c.fill=_fill_c(row_bg); c.border=_brd(); c.alignment=_aln()
            if ci==4 and val:
                c.font=Font(color="FF0277BD",size=10,underline="single")
            elif ci==2:
                c.font=_fnt(bold=not is_sub,size=10)
            else:
                c.font=_fnt(size=10)
    ws1.freeze_panes="A2"
    ws1.auto_filter.ref=f"A1:E{len(menu_rows)+1}"

    # ── Sheet 2: URL Test Results ─────────────────────────────────────────────
    if test_results:
        ws2 = wb.create_sheet("URL Test Results")
        h2 = ["#","Main Menu","Sub Menu","URL","Status","Load(s)","Page Title",
              "Form?","Forms","Table?","Buttons?","Inputs","Selects",
              "File Upload?","Error on Page","Notes"]
        w2 = [4, 22,24,52,9,7,30,7,6,7,8,7,7,10,25,32]
        for ci,(h,w) in enumerate(zip(h2,w2),1):
            c = ws2.cell(1,ci,h)
            c.fill=_fill_c("FF00838F"); c.font=_fnt(True,"FFFFFFFF",11)
            c.alignment=_aln("center"); c.border=_brd()
            ws2.column_dimensions[c.column_letter].width=w
        ws2.row_dimensions[1].height=22

        for ri,(mr,tr) in enumerate(zip(menu_rows,test_results),2):
            st = tr["status"]
            if   st=="OK":      sf,ff=_fill_c("FFE8F5E9"),_fnt(True,"FF1B5E20")
            elif st=="TIMEOUT": sf,ff=_fill_c("FFFFF8E1"),_fnt(True,"FFF57F17")
            else:               sf,ff=_fill_c("FFFFEBEE"),_fnt(True,"FFB71C1C")
            bg = "FFF0F4F8" if ri%2==0 else "FFFFFFFF"
            vals=[ri-1,mr["main"],mr["sub"],tr["url"],st,tr["load_sec"],
                  tr["page_title"],tr["has_form"],tr["form_count"],tr["has_table"],
                  tr["has_buttons"],tr["input_count"],tr["select_count"],
                  tr["file_upload"],tr["error_on_page"],tr["page_notes"]]
            for ci,val in enumerate(vals,1):
                c = ws2.cell(ri,ci,val)
                if ci==5: c.fill=sf; c.font=ff
                else:     c.fill=_fill_c(bg); c.font=_fnt()
                c.alignment=_aln(); c.border=_brd()
        ws2.freeze_panes="A2"
        ws2.auto_filter.ref=f"A1:P{len(test_results)+1}"

    # ── Sheet 3: Summary ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3.column_dimensions["A"].width=32
    ws3.column_dimensions["B"].width=24

    total_mains = len({r["main"] for r in menu_rows})
    total_subs  = sum(1 for r in menu_rows if r["sub"])

    rows_s = [
        ("Report Generated",    datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Site URL",            cfg["url"]),
        ("Username",            cfg["username"]),
        ("Nav Mode",            cfg["nav_mode"]),
        ("",""),
        ("── MENU STRUCTURE ──",""),
        ("Total Rows",          len(menu_rows)),
        ("Distinct Main Menus", total_mains),
        ("Sub-menu Items",      total_subs),
        ("Top-level Only",      len(menu_rows)-total_subs),
    ]
    if test_results:
        ok  = sum(1 for t in test_results if t["status"]=="OK")
        to  = sum(1 for t in test_results if t["status"]=="TIMEOUT")
        err = sum(1 for t in test_results
                  if t["status"] not in ("OK","TIMEOUT","NOT_TESTED","NO_URL"))
        rows_s+=[("",""),("── URL TESTS ──",""),
                 ("URLs Tested",len(test_results)),
                 ("✅  OK",ok),("⚠️  Timeout",to),("❌  Error",err)]

    for ri,(lbl,val) in enumerate(rows_s,1):
        ws3.cell(ri,1,lbl).font=_fnt(bold="──" in lbl,sz=10)
        ws3.cell(ri,2,str(val)).font=_fnt(sz=10)
        ws3.cell(ri,1).alignment=ws3.cell(ri,2).alignment=_aln()

    wb.save(cfg["excel_file"])
    print(green(f"  ✔ Excel saved → {cfg['excel_file']}"))


# ─────────────────────────────────────────────────────────────────────────────
#  PRINT MENU TREE
# ─────────────────────────────────────────────────────────────────────────────
def print_tree(rows):
    section("MENU STRUCTURE FOUND")
    current_main = None
    for r in rows:
        if r["main"] != current_main:
            print(f"  {cyan('📁')} {bold(r['main'])}")
            current_main = r["main"]
        if r["sub"]:
            print(f"      {dim('└─')} {r['sub']}")
    print()
    mains = len({r["main"] for r in rows})
    subs  = sum(1 for r in rows if r["sub"])
    print(f"  Main menus: {bold(str(mains))}   Sub-items: {bold(str(subs))}   Total rows: {bold(str(len(rows)))}")


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    cfg    = get_config()
    driver = init_driver(cfg)
    menu_rows    = []
    test_results = []

    try:
        # Login
        current_url = login(driver, cfg)

        # Save HTML
        section("SCRAPING HTML")
        html = save_html(driver, cfg["html_file"])

        # Parse menu
        section("PARSING NAVIGATION")
        menu_rows = parse_menu(html, current_url, cfg["nav_mode"])

        if not menu_rows:
            print(red("  ✘ No menu items found."))
            print(yellow("  Try a different nav mode or check the HTML dump."))
        else:
            print_tree(menu_rows)
            save_csv(menu_rows, cfg["csv_file"])

        # URL Testing
        if menu_rows and cfg.get("run_tests"):
            section("FUNCTIONAL URL TESTING")
            to_test = [(r, r["sub_url"] or r["main_url"]) for r in menu_rows]
            if cfg.get("skip_no_url"):
                to_test = [(r, u) for r, u in to_test if u]

            total = len(to_test)
            print(f"  Testing {total} URLs …\n")

            for idx, (row, url) in enumerate(to_test):
                label = f"{row['main']} › {row['sub']}" if row["sub"] else row["main"]
                bar   = f"[{idx+1:>3}/{total}]"
                print(f"  {dim(bar)} {label[:52]:<52}", end=" ", flush=True)
                result = test_url(driver, url, cfg.get("test_pause", 1))
                test_results.append(result)
                if   result["status"] == "OK":      print(green("OK"))
                elif result["status"] == "TIMEOUT": print(yellow("TIMEOUT"))
                elif result["status"] == "NO_URL":  print(dim("NO URL"))
                else:                               print(red(result["status"]))

        # Excel report
        section("WRITING REPORT")
        write_excel(menu_rows, test_results, cfg)

        # Done
        hr("═")
        print(bold("  DONE!"))
        hr("═")
        print(f"  {'HTML':15} {cfg['html_file']}")
        print(f"  {'CSV':15} {cfg['csv_file']}")
        print(f"  {'Excel':15} {cfg['excel_file']}")
        if cfg.get("output_dir") and cfg["output_dir"] != ".":
            print(f"  {'Output folder':15} {os.path.abspath(cfg['output_dir'])}")
        hr("═")

    except KeyboardInterrupt:
        print(yellow("\n\n  [!] Interrupted by user."))
        if menu_rows:
            print(yellow("  Saving partial results …"))
            save_csv(menu_rows, cfg["csv_file"])
            write_excel(menu_rows, test_results, cfg)
    finally:
        input("\n  Press ENTER to close the browser …")
        driver.quit()


if __name__ == "__main__":
    main()