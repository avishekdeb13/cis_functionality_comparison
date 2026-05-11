"""
DRT Portal Menu Scraper & URL Tester - v3
==========================================
FIXES:
  - Parses Bootstrap dropdown-menu correctly (div.dropdown-menu > a.dropdown-item)
  - Legacy Data menu now captured
  - User provides URL / username / password at runtime (universal)
  - Full URL functional test with status report
  - Excel report with color-coded results

STEPS THIS SCRIPT DOES:
  Step 1 → Ask user for site URL, username, password
  Step 2 → Open browser, auto-fill credentials, wait for CAPTCHA + login
  Step 3 → Dump page HTML to file
  Step 4 → Parse HTML with BeautifulSoup (Bootstrap nav structure)
  Step 5 → Save Menu CSV
  Step 6 → Ask user whether to run URL tests
  Step 7 → Hit every sub-menu URL, record status/features
  Step 8 → Write Excel report (Menu Hierarchy + Test Results + Summary)

Install:
  pip install selenium webdriver-manager openpyxl beautifulsoup4

Run:
  python drt_scraper_v3.py
"""

import os, csv, re, time, sys
from datetime import datetime
from urllib.parse import urljoin

from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException,
    StaleElementReferenceException, WebDriverException,
)
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ─────────────────────────────────────────────────────────────────────────────
#  STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="FF000000", size=10):
    return Font(bold=bold, color=color, size=size)

def _border():
    s = Side(style="thin", color="FFD0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left"):
    return Alignment(horizontal=h, vertical="center", wrap_text=True)

# Colour palette
C_HEADER_BLUE  = "FF0277BD"
C_HEADER_CYAN  = "FF00838F"
C_MAIN_ROW     = "FFE1F5FE"
C_SUB_ROW_ODD  = "FFFAFAFA"
C_SUB_ROW_EVEN = "FFF0F4F8"
C_WHITE        = "FFFFFFFF"
C_OK           = "FFE8F5E9"
C_OK_TEXT      = "FF1B5E20"
C_WARN         = "FFFFF8E1"
C_WARN_TEXT    = "FFF57F17"
C_ERR          = "FFFFEBEE"
C_ERR_TEXT     = "FFB71C1C"


# ─────────────────────────────────────────────────────────────────────────────
#  STEP 1 — USER INPUT UI
# ─────────────────────────────────────────────────────────────────────────────
def get_user_config():
    import tkinter as tk
    from tkinter import ttk, messagebox
    from urllib.parse import urlparse

    config = {}

    def submit():
        url = url_var.get().strip()
        user = user_var.get().strip()
        pwd = pass_var.get().strip()

        if not url or not user or not pwd:
            messagebox.showerror("Error", "All fields are required")
            return

        domain = urlparse(url).netloc.replace(".", "_").replace("/", "_")
        prefix = domain[:30] if domain else "site"

        config.update({
            "url": url,
            "username": user,
            "password": pwd,
            "prefix": prefix,
            "html_file": f"{prefix}_page.html",
            "csv_file": f"{prefix}_menu.csv",
            "excel_file": f"{prefix}_report.xlsx",
        })

        root.destroy()

    # ---------------- UI ----------------
    root = tk.Tk()
    root.title("DRT Portal Scraper Login")
    root.geometry("500x280")
    root.resizable(False, False)

    ttk.Label(root, text="DRT Portal Scraper & Tester",
              font=("Arial", 16, "bold")).pack(pady=15)

    frame = ttk.Frame(root, padding=20)
    frame.pack(fill="both", expand=True)

    # Default values
    default_url = "https://cis.drt.gov.in/drtlive/index.php"
    default_user = "filingdrt1"
    default_pass = "Sodrt1@6060"

    url_var = tk.StringVar(value=default_url)
    user_var = tk.StringVar(value=default_user)
    pass_var = tk.StringVar(value=default_pass)

    # URL
    ttk.Label(frame, text="Site URL:").grid(row=0, column=0, sticky="w", pady=10)
    ttk.Entry(frame, textvariable=url_var, width=50).grid(row=0, column=1)

    # Username
    ttk.Label(frame, text="Username:").grid(row=1, column=0, sticky="w", pady=10)
    ttk.Entry(frame, textvariable=user_var, width=50).grid(row=1, column=1)

    # Password
    ttk.Label(frame, text="Password:").grid(row=2, column=0, sticky="w", pady=10)
    ttk.Entry(frame, textvariable=pass_var, show="*", width=50).grid(row=2, column=1)

    # Button
    ttk.Button(frame, text="Start Scraper", command=submit)\
        .grid(row=3, column=1, pady=25)

    root.mainloop()

    return config

# ─────────────────────────────────────────────────────────────────────────────
#  STEP 2 — DRIVER
# ─────────────────────────────────────────────────────────────────────────────
def init_driver():
    opts = webdriver.ChromeOptions()
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    try:
        from webdriver_manager.chrome import ChromeDriverManager
        from selenium.webdriver.chrome.service import Service
        return webdriver.Chrome(
            service=Service(ChromeDriverManager().install()), options=opts)
    except Exception:
        return webdriver.Chrome(options=opts)


# ─────────────────────────────────────────────────────────────────────────────
#  STEP 3 — LOGIN
# ─────────────────────────────────────────────────────────────────────────────
CAPTCHA_WAIT = 90

def login(driver, cfg):
    print("[STEP 2] Opening login page …")
    driver.get(cfg["url"])
    time.sleep(3)

    def fill(label, value, locators):
        for loc in locators:
            try:
                el = driver.find_element(*loc)
                el.clear(); el.send_keys(value)
                print(f"  ✔ {label} filled")
                return
            except NoSuchElementException:
                pass
        print(f"  ✘ {label} field not found — fill manually in browser")

    fill("Username", cfg["username"], [
        (By.NAME, "username"), (By.NAME, "user_name"), (By.NAME, "txtUsername"),
        (By.ID,   "username"), (By.ID,   "txtUsername"),
        (By.XPATH, "//input[@type='text'][1]"),
        (By.XPATH, "//input[contains(@name,'user') or contains(@id,'user')]"),
    ])
    fill("Password", cfg["password"], [
        (By.NAME, "password"), (By.NAME, "txtPassword"),
        (By.ID,   "password"), (By.ID,   "txtPassword"),
        (By.XPATH, "//input[@type='password']"),
    ])

    print(f"""
{'='*62}
  ACTION REQUIRED:
    1. Solve the CAPTCHA in the browser window
    2. Click the LOGIN / SUBMIT button
  You have {CAPTCHA_WAIT} seconds.
{'='*62}""")

    start = driver.current_url
    try:
        WebDriverWait(driver, CAPTCHA_WAIT).until(lambda d: d.current_url != start)
        print(f"  ✔ Login successful  →  {driver.current_url}")
    except TimeoutException:
        print("  [!] URL did not change — continuing anyway …")
    time.sleep(4)


# ─────────────────────────────────────────────────────────────────────────────
#  STEP 4 — SAVE HTML
# ─────────────────────────────────────────────────────────────────────────────
def save_html(driver, path):
    html = driver.page_source
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"\n[STEP 3] HTML saved → {path}  ({os.path.getsize(path):,} bytes)")
    return html


# ─────────────────────────────────────────────────────────────────────────────
#  STEP 5 — PARSE BOOTSTRAP NAV FROM HTML
#
#  DRT uses this exact pattern:
#    <nav class="navbar …">
#      <ul class="navbar-nav …">
#        <li class="nav-item">                          ← top-level (no dropdown)
#          <a class="nav-link" href="…">Home</a>
#        </li>
#        <li class="nav-item dropdown …">              ← has dropdown
#          <a class="nav-link" …>Application in DRT</a>
#          <div class="dropdown-menu …">
#            <a class="dropdown-item" href="…">…</a>   ← sub-items
#          </div>
#        </li>
#      </ul>
#    </nav>
# ─────────────────────────────────────────────────────────────────────────────
def parse_menu(html, base_url):
    """
    Returns list of dicts:
      { main, main_url, sub, sub_url }
    sub / sub_url are '' when a top-level item has no dropdown.
    """
    soup = BeautifulSoup(html, "html.parser")
    rows = []

    def abs_url(href):
        if not href:
            return ""
        href = href.strip()
        if href in ("#", "javascript:void(0)", "javascript:;", ""):
            return ""
        return urljoin(base_url, href)

    # ── Primary strategy: Bootstrap navbar-nav ────────────────────────────
    nav = soup.find("nav", class_=re.compile(r"navbar", re.I))
    if not nav:
        nav = soup  # fall back to whole page

    nav_ul = nav.find("ul", class_=re.compile(r"navbar-nav", re.I))
    if not nav_ul:
        # try any ul with at least 4 li children
        for ul in soup.find_all("ul"):
            if len(ul.find_all("li", recursive=False)) >= 4:
                nav_ul = ul
                break

    if nav_ul:
        top_items = nav_ul.find_all("li", recursive=False)
        print(f"\n[STEP 4] Found {len(top_items)} top-level <li> items in navbar")

        for li in top_items:
            # Main label
            main_a   = li.find("a", class_=re.compile(r"nav-link", re.I))
            if not main_a:
                main_a = li.find("a")
            if not main_a:
                continue

            main_txt = main_a.get_text(strip=True).lstrip("*").strip()
            main_url = abs_url(main_a.get("href", ""))

            # Look for dropdown-menu div inside this li
            dropdown = li.find("div", class_=re.compile(r"dropdown-menu", re.I))

            if dropdown:
                sub_links = dropdown.find_all("a", class_=re.compile(r"dropdown-item", re.I))
                # fallback: any anchor inside dropdown
                if not sub_links:
                    sub_links = dropdown.find_all("a", href=True)

                added = 0
                for a in sub_links:
                    sub_txt = a.get_text(strip=True)
                    sub_url = abs_url(a.get("href", ""))
                    if sub_txt:
                        rows.append({
                            "main":     main_txt,
                            "main_url": main_url,
                            "sub":      sub_txt,
                            "sub_url":  sub_url,
                        })
                        added += 1

                if added:
                    print(f"  ✔ [{main_txt}] → {added} sub-items")
                else:
                    # dropdown div exists but empty (all commented out)
                    rows.append({"main": main_txt, "main_url": main_url,
                                 "sub": "", "sub_url": main_url})
                    print(f"  ✔ [{main_txt}] → dropdown empty (all commented)")
            else:
                # Plain nav-item with no dropdown
                rows.append({"main": main_txt, "main_url": main_url,
                             "sub": "", "sub_url": main_url})
                print(f"  ✔ [{main_txt}] → no dropdown (direct link)")

        return rows

    # ── Fallback: collect all anchors ─────────────────────────────────────
    print("  [!] Could not find navbar-nav, collecting all anchors as fallback")
    for a in soup.find_all("a", href=True):
        txt = a.get_text(strip=True)
        url = abs_url(a["href"])
        if txt and url and len(txt) < 80:
            rows.append({"main": txt, "main_url": url, "sub": "", "sub_url": url})
    return rows


# ─────────────────────────────────────────────────────────────────────────────
#  STEP 6 — SAVE CSV
# ─────────────────────────────────────────────────────────────────────────────
def save_csv(rows, path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Main Menu", "Sub Menu", "URL"])
        for r in rows:
            w.writerow([r["main"], r["sub"], r["sub_url"] or r["main_url"]])
    print(f"\n[STEP 5] Menu CSV saved → {path}  ({len(rows)} rows)")


# ─────────────────────────────────────────────────────────────────────────────
#  STEP 7 — FUNCTIONAL URL TEST
# ─────────────────────────────────────────────────────────────────────────────
def test_url(driver, url):
    r = {
        "url":          url,
        "status":       "NOT_TESTED",
        "load_sec":     "",
        "page_title":   "",
        "has_form":     "",
        "form_count":   "",
        "has_table":    "",
        "has_buttons":  "",
        "input_count":  "",
        "button_labels":"",
        "select_count": "",
        "file_upload":  "",
        "error_on_page":"",
        "page_notes":   "",
    }
    if not url:
        r["status"] = "NO_URL"; return r

    try:
        t0 = time.time()
        driver.get(url)
        time.sleep(2)
        r["load_sec"]   = round(time.time() - t0, 2)
        r["page_title"] = driver.title.strip()[:80]

        soup = BeautifulSoup(driver.page_source, "html.parser")
        body = soup.get_text(separator=" ", strip=True).lower()[:2000]

        # Forms
        forms = soup.find_all("form")
        r["has_form"]   = "YES" if forms else "NO"
        r["form_count"] = len(forms)
        r["input_count"]= sum(len(f.find_all(["input","textarea"])) for f in forms)
        r["select_count"]= len(soup.find_all("select"))

        # File upload
        r["file_upload"] = "YES" if soup.find("input", {"type":"file"}) else "NO"

        # Tables
        r["has_table"] = "YES" if soup.find("table") else "NO"

        # Buttons
        btns = soup.find_all(
            lambda t: t.name == "button" or
                      (t.name == "input" and t.get("type","").lower()
                       in ("submit","button","reset"))
        )
        r["has_buttons"]   = "YES" if btns else "NO"
        r["button_labels"] = " | ".join(
            (b.get_text(strip=True) or b.get("value","") or b.get("name",""))[:25]
            for b in btns[:6]
        )

        # Error detection
        err_kw = ["error","invalid","access denied","not found",
                  "unauthorized","exception","500","403","404"]
        found  = [k for k in err_kw if k in body]
        r["error_on_page"] = ", ".join(found) if found else ""

        # Notes
        notes = []
        if forms:         notes.append(f"{len(forms)} form(s)")
        if soup.find("table"): notes.append("data table")
        if soup.find("select"): notes.append(f"{r['select_count']} dropdown(s)")
        if r["file_upload"]=="YES": notes.append("file upload")
        if soup.find(attrs={"class": re.compile(r"alert|success|danger", re.I)}):
            notes.append("alert box")
        r["page_notes"] = " | ".join(notes)
        r["status"]     = "OK"

    except TimeoutException:
        r["status"] = "TIMEOUT"
    except WebDriverException as e:
        r["status"]     = "ERROR"
        r["error_on_page"] = str(e)[:100]

    return r


# ─────────────────────────────────────────────────────────────────────────────
#  STEP 8 — EXCEL REPORT
# ─────────────────────────────────────────────────────────────────────────────
def write_excel(menu_rows, test_results, cfg):
    wb = Workbook()

    # ── Sheet 1 : Menu Hierarchy ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Menu Hierarchy"

    h1 = ["#", "Main Menu", "Sub Menu", "URL", "Type"]
    w1 = [5, 28, 32, 65, 12]
    for ci, (h, w) in enumerate(zip(h1, w1), 1):
        c = ws1.cell(1, ci, h)
        c.fill = _fill(C_HEADER_BLUE); c.font = _font(True, "FFFFFFFF", 11)
        c.alignment = _align("center"); c.border = _border()
        ws1.column_dimensions[c.column_letter].width = w
    ws1.row_dimensions[1].height = 22

    for ri, row in enumerate(menu_rows, 2):
        is_sub  = bool(row["sub"])
        mtype   = "Sub-item" if is_sub else "Top-level"
        url     = row["sub_url"] or row["main_url"]
        vals    = [ri - 1, row["main"], row["sub"], url, mtype]
        bg      = C_SUB_ROW_EVEN if ri % 2 == 0 else C_SUB_ROW_ODD
        row_bg  = bg if is_sub else C_MAIN_ROW

        for ci, val in enumerate(vals, 1):
            c = ws1.cell(ri, ci, val)
            c.fill = _fill(row_bg); c.border = _border()
            if ci == 4 and val:
                c.font = Font(color="FF0277BD", size=10, underline="single")
            elif ci == 2:
                c.font = _font(bold=not is_sub, size=10)
            else:
                c.font = _font(size=10)
            c.alignment = _align()

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:E{len(menu_rows)+1}"

    # ── Sheet 2 : URL Test Results ────────────────────────────────────────────
    if test_results:
        ws2 = wb.create_sheet("URL Test Results")
        h2 = ["#", "Main Menu", "Sub Menu", "URL", "Status", "Load(s)",
              "Page Title", "Form?", "Forms", "Table?", "Buttons?",
              "Inputs", "Selects", "File Upload?", "Error on Page", "Notes"]
        w2 = [4, 22, 24, 52, 9, 7, 30, 7, 6, 7, 8, 7, 7, 10, 25, 32]

        for ci, (h, w) in enumerate(zip(h2, w2), 1):
            c = ws2.cell(1, ci, h)
            c.fill = _fill(C_HEADER_CYAN); c.font = _font(True, "FFFFFFFF", 11)
            c.alignment = _align("center"); c.border = _border()
            ws2.column_dimensions[c.column_letter].width = w
        ws2.row_dimensions[1].height = 22

        for ri, (mr, tr) in enumerate(zip(menu_rows, test_results), 2):
            st = tr["status"]
            if   st == "OK":      sfill, sfont = _fill(C_OK),   _font(True, C_OK_TEXT)
            elif st == "TIMEOUT": sfill, sfont = _fill(C_WARN),  _font(True, C_WARN_TEXT)
            else:                 sfill, sfont = _fill(C_ERR),   _font(True, C_ERR_TEXT)

            bg   = C_SUB_ROW_EVEN if ri % 2 == 0 else C_WHITE
            vals = [ri - 1, mr["main"], mr["sub"], tr["url"],
                    st, tr["load_sec"], tr["page_title"],
                    tr["has_form"], tr["form_count"], tr["has_table"],
                    tr["has_buttons"], tr["input_count"], tr["select_count"],
                    tr["file_upload"], tr["error_on_page"], tr["page_notes"]]

            for ci, val in enumerate(vals, 1):
                c = ws2.cell(ri, ci, val)
                if ci == 5:
                    c.fill = sfill; c.font = sfont
                else:
                    c.fill = _fill(bg); c.font = _font()
                c.alignment = _align(); c.border = _border()

        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = f"A1:P{len(test_results)+1}"

    # ── Sheet 3 : Summary ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 22

    total_mains = len({r["main"] for r in menu_rows})
    total_subs  = sum(1 for r in menu_rows if r["sub"])
    top_only    = sum(1 for r in menu_rows if not r["sub"])

    rows_s = [
        ("Report Generated",    datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Site URL",            cfg["url"]),
        ("Username",            cfg["username"]),
        ("",                    ""),
        ("── MENU STRUCTURE ──",""),
        ("Total Rows",          len(menu_rows)),
        ("Distinct Main Menus", total_mains),
        ("Sub-menu Items",      total_subs),
        ("Top-level Only",      top_only),
    ]
    if test_results:
        ok  = sum(1 for t in test_results if t["status"] == "OK")
        to  = sum(1 for t in test_results if t["status"] == "TIMEOUT")
        err = sum(1 for t in test_results if t["status"] not in ("OK","TIMEOUT","NOT_TESTED","NO_URL"))
        rows_s += [
            ("",""),
            ("── URL TESTS ──",""),
            ("URLs Tested",        len(test_results)),
            ("✅  OK",            ok),
            ("⚠️   Timeout",       to),
            ("❌  Error",          err),
        ]

    for ri, (lbl, val) in enumerate(rows_s, 1):
        ws3.cell(ri, 1, lbl).font = _font(bold="──" in lbl, size=10)
        ws3.cell(ri, 2, str(val)).font = _font(size=10)
        ws3.cell(ri, 1).alignment = ws3.cell(ri, 2).alignment = _align()

    wb.save(cfg["excel_file"])
    print(f"[✔] Excel report saved → {cfg['excel_file']}")


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    cfg    = get_user_config()
    driver = init_driver()
    menu_rows    = []
    test_results = []

    try:
        # Login
        login(driver, cfg)
        current_url = driver.current_url

        # Save HTML
        html = save_html(driver, cfg["html_file"])

        # Parse menu from HTML
        menu_rows = parse_menu(html, current_url)
        print(f"\n  Total menu rows parsed: {len(menu_rows)}")

        # Print summary
        print("\n  ── Menu Structure ───────────────────────────────────")
        current_main = None
        for r in menu_rows:
            if r["main"] != current_main:
                print(f"  📁 {r['main']}")
                current_main = r["main"]
            if r["sub"]:
                print(f"     └─ {r['sub']}")
        print("  ─────────────────────────────────────────────────────")

        # Save CSV
        save_csv(menu_rows, cfg["csv_file"])

        # URL Testing
        print()
        do_test = input("[STEP 6] Run functional URL test on all menu items? (y/n): ").strip().lower()

        if do_test == "y":
            total = len(menu_rows)
            print(f"\n  Testing {total} URLs …\n")
            for idx, row in enumerate(menu_rows):
                url   = row["sub_url"] or row["main_url"]
                label = f"{row['main']} > {row['sub']}" if row["sub"] else row["main"]
                print(f"  [{idx+1:>3}/{total}] {label[:55]:<55}", end=" … ", flush=True)
                result = test_url(driver, url)
                test_results.append(result)
                print(result["status"])
        else:
            print("  Skipping URL tests.")

        # Excel
        write_excel(menu_rows, test_results if do_test == "y" else [], cfg)

        print(f"""
{'='*62}
  DONE!
  Menu CSV   : {cfg['csv_file']}
  Excel      : {cfg['excel_file']}
  HTML dump  : {cfg['html_file']}
{'='*62}""")

    except KeyboardInterrupt:
        print("\n[!] Interrupted by user.")
    finally:
        input("\nPress ENTER to close the browser …")
        driver.quit()


if __name__ == "__main__":
    main()