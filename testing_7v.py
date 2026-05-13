"""
Universal Portal Testing Tool v1.4
- GUI login window
- Expands ALL sidebar dropdowns
- ALWAYS captures every menu item (even broken URLs)
- Screenshots per page
- Full component extraction
- Excel + CSV named after login user
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import tkinter as tk
from tkinter import ttk, messagebox
import time, os, csv, re
from datetime import datetime
from urllib.parse import urlparse

# ── STEP 1: GUI ───────────────────────────────────────────────────────────────
def get_config():
    config = {}

    def submit():
        url  = url_var.get().strip()
        user = user_var.get().strip()
        pwd  = pass_var.get().strip()
        if not url or not user or not pwd:
            messagebox.showerror("Error", "All fields are required")
            return
        config["url"]      = url
        config["username"] = user
        config["password"] = pwd
        ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
        folder = user + "_" + ts
        ss_dir = os.path.join(folder, "screenshots")
        os.makedirs(ss_dir, exist_ok=True)
        config["folder"]      = folder
        config["screenshots"] = ss_dir
        config["html"]        = os.path.join(folder, user + "_portal.html")
        config["csv"]         = os.path.join(folder, user + "_menu.csv")
        config["excel"]       = os.path.join(folder, user + "_test_report.xlsx")
        config["base"]        = urlparse(url).scheme + "://" + urlparse(url).netloc
        root.destroy()

    root = tk.Tk()
    root.title("Portal Testing Tool v1.4")
    root.geometry("540x320")
    root.resizable(False, False)
    tk.Label(root, text="Universal Portal Testing Tool v1.4",
             font=("Arial", 15, "bold"), fg="#003366").pack(pady=14)
    tk.Label(root, text="Output files & folder named after your username",
             font=("Arial", 9), fg="#777").pack()
    frame = ttk.Frame(root, padding=20)
    frame.pack(fill="both", expand=True)
    url_var  = tk.StringVar(value="https://drt.etribunals.gov.in/cis2.0/filing/login")
    user_var = tk.StringVar(value="")
    pass_var = tk.StringVar(value="")
    ttk.Label(frame, text="Portal URL:").grid(row=0, column=0, sticky="w", pady=8)
    ttk.Entry(frame, textvariable=url_var, width=46).grid(row=0, column=1, padx=8)
    ttk.Label(frame, text="Username:").grid(row=1, column=0, sticky="w", pady=8)
    ttk.Entry(frame, textvariable=user_var, width=46).grid(row=1, column=1, padx=8)
    ttk.Label(frame, text="Password:").grid(row=2, column=0, sticky="w", pady=8)
    ttk.Entry(frame, textvariable=pass_var, show="*", width=46).grid(row=2, column=1, padx=8)
    ttk.Button(frame, text="Start Testing", command=submit).grid(
        row=3, column=1, pady=20, sticky="e")
    root.mainloop()
    return config

# ── STEP 2: Browser ───────────────────────────────────────────────────────────
def init_driver():
    opts = webdriver.ChromeOptions()
    opts.add_argument("--start-maximized")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=opts)

# ── STEP 3: Login ─────────────────────────────────────────────────────────────
def login(driver, cfg):
    print("\n[1] Opening login page...")
    driver.get(cfg["url"])
    time.sleep(3)

    for fid in ["user_name", "username"]:
        try:
            driver.find_element(By.ID, fid).send_keys(cfg["username"])
            print("    Username filled"); break
        except: pass

    for fid in ["user_pass", "password"]:
        try:
            driver.find_element(By.ID, fid).send_keys(cfg["password"])
            print("    Password filled"); break
        except: pass

    driver.save_screenshot(os.path.join(cfg["screenshots"], "00_login.png"))

    print("\n" + "="*52)
    print("  Solve CAPTCHA then click LOGIN")
    print("  Waiting 120 seconds...")
    print("="*52 + "\n")

    WebDriverWait(driver, 120).until(lambda d: "login" not in d.current_url)
    time.sleep(3)
    driver.save_screenshot(os.path.join(cfg["screenshots"], "01_dashboard.png"))
    print("[OK] Logged in:", driver.current_url)

# ── STEP 4: Expand sidebar ────────────────────────────────────────────────────
def get_html(driver, cfg):
    print("\n[2] Expanding ALL sidebar menus...")

    # Pass 1 — JS force show all collapse divs
    driver.execute_script("""
        document.querySelectorAll('.collapse').forEach(el => el.classList.add('show'));
        document.querySelectorAll('[aria-expanded]').forEach(el =>
            el.setAttribute('aria-expanded', 'true'));
    """)
    time.sleep(1)

    # Pass 2 — physically click each toggle
    toggles = driver.find_elements(By.CSS_SELECTOR, "[data-bs-toggle='collapse']")
    print("    Found", len(toggles), "collapsible toggles")
    for toggle in toggles:
        try:
            driver.execute_script("arguments[0].scrollIntoView(true);", toggle)
            time.sleep(0.1)
            target = (toggle.get_attribute("data-bs-target") or "").replace("#", "")
            if target:
                try:
                    el = driver.find_element(By.ID, target)
                    if "show" not in (el.get_attribute("class") or ""):
                        driver.execute_script("arguments[0].click();", toggle)
                        time.sleep(0.3)
                except:
                    driver.execute_script("arguments[0].click();", toggle)
                    time.sleep(0.3)
            name = toggle.text.strip()[:40] or target
            print("    [+] Expanded:", name)
        except: pass

    # Pass 3 — final JS sweep
    driver.execute_script("""
        document.querySelectorAll('.collapse').forEach(el => el.classList.add('show'));
    """)
    time.sleep(2)

    driver.save_screenshot(os.path.join(cfg["screenshots"], "02_sidebar_expanded.png"))

    html = driver.page_source
    with open(cfg["html"], "w", encoding="utf-8") as f:
        f.write(html)

    # Verify capture
    soup  = BeautifulSoup(html, "html.parser")
    items = soup.find_all("a", class_="dropdown-item")
    print("    Captured", len(items), "dropdown items in HTML")
    for item in items:
        print("      -", item.get_text(strip=True), "->", item.get("href","")[:60])
    print("    HTML saved:", cfg["html"])
    return html

# ── STEP 5: URL cleaner ───────────────────────────────────────────────────────
def clean_url(href, base):
    """
    Returns (clean_url, issue).
    NEVER returns None — always returns original href as fallback.
    """
    if not href or not href.strip():
        return "", "Empty href"

    href = href.strip()

    # Pure javascript links — no URL possible
    if href.lower().startswith("javascript"):
        return "", "JavaScript link"

    # Broken Angular routerlink with quotes/commas
    # e.g. /cis2.0/scrutinyreport', 'level1
    if re.search(r"['\",]", href):
        # Try to salvage the path before the first quote
        clean = re.split(r"['\",]", href)[0].strip()
        if clean and clean.startswith("/"):
            full = base + clean
            return full, "PORTAL BUG: broken routerlink (partial URL used). Original: " + href[:60]
        return "", "PORTAL BUG: broken routerlink - " + href[:60]

    # Normal absolute URL
    if href.startswith("http"):
        return href, ""

    # Normal relative URL
    if href.startswith("/"):
        return base + href, ""

    # Relative like charts.html
    return href, "Relative URL (may not load correctly)"

# ── STEP 6: Parse menu — captures EVERYTHING ──────────────────────────────────
def parse_menu(html, cfg):
    print("\n[3] Parsing ALL menu items (including broken URLs)...")
    soup = BeautifulSoup(html, "html.parser")
    rows = []

    sidebar = soup.find("ul", id="accordionSidebar")
    if not sidebar:
        sidebar = soup.find("ul", class_=lambda c: c and "navbar-nav" in c)

    if not sidebar:
        print("    [!] No sidebar found - collecting all anchors")
        for a in soup.find_all("a", href=True):
            txt  = a.get_text(strip=True)
            href = a.get("href", "")
            if txt and len(txt) < 80:
                url, issue = clean_url(href, cfg["base"])
                rows.append({"main": txt, "sub": "", "url": url,
                             "issue": issue, "href": href})
        return rows

    for li in sidebar.find_all("li", class_="nav-item"):
        main_a = li.find("a", class_="nav-link")
        if not main_a:
            continue
        main_name = main_a.get_text(strip=True)
        if not main_name:
            continue

        # Find ALL collapse-inner divs (handles nested dropdowns)
        collapses = li.find_all("div", class_="collapse-inner")

        if collapses:
            for collapse in collapses:
                sub_links = collapse.find_all("a", class_="dropdown-item")

                # Fallback: any anchor inside collapse
                if not sub_links:
                    sub_links = collapse.find_all("a", href=True)

                for a in sub_links:
                    sub  = a.get_text(strip=True)
                    if not sub:
                        continue
                    href = a.get("href", "")
                    url, issue = clean_url(href, cfg["base"])

                    print("    FOUND:", main_name, ">", sub)
                    print("           href :", href[:70])
                    if issue:
                        print("           ISSUE:", issue)

                    # ALWAYS add — never skip based on URL validity
                    rows.append({
                        "main":  main_name,
                        "sub":   sub,
                        "url":   url,
                        "issue": issue,
                        "href":  href
                    })
        else:
            # Top-level direct link
            href = main_a.get("href", "")
            url, issue = clean_url(href, cfg["base"])

            # Only add if it has a real URL (skip pure JS links for top-level)
            if url:
                rows.append({
                    "main":  main_name,
                    "sub":   "",
                    "url":   url,
                    "issue": issue,
                    "href":  href
                })
                print("    FOUND:", main_name, "(direct) ->", url[:55])

    print("\n    Total menu items captured:", len(rows))
    return rows

# ── STEP 7: Save CSV ──────────────────────────────────────────────────────────
def save_csv(rows, cfg):
    with open(cfg["csv"], "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Main Menu", "Sub Menu", "URL", "Original Href", "Issue"])
        for r in rows:
            w.writerow([r["main"], r["sub"], r["url"], r["href"], r["issue"]])
    print("\n[4] CSV saved:", cfg["csv"])

# ── STEP 8: Extract page components ──────────────────────────────────────────
# ── STEP 8: Extract page components (ENHANCED) ───────────────────────────────
def extract_components(driver, soup):
    """
    Extracts detailed UI components from the current page.
    Uses both BeautifulSoup (static) and Selenium (rendered Angular).
    """
    comps = {
        "forms":      [],
        "buttons":    [],
        "inputs":     [],
        "tables":     [],
        "dropdowns":  [],
        "labels":     [],
        "headings":   [],
        "page_sections": [],
    }

    # ── Headings / Page Title ─────────────────────────────────────────────
    for tag in ["h1", "h2", "h3", "h4", "h5"]:
        for h in soup.find_all(tag):
            txt = h.get_text(strip=True)
            if txt and len(txt) > 2:
                comps["headings"].append(tag.upper() + ": " + txt[:50])

    # ── Card / Section Titles ─────────────────────────────────────────────
    for card in soup.find_all(class_=re.compile(r"card-header|panel-heading|section-title", re.I)):
        txt = card.get_text(strip=True)
        if txt:
            comps["page_sections"].append(txt[:50])

    # ── Forms ─────────────────────────────────────────────────────────────
    for i, form in enumerate(soup.find_all("form"), 1):
        method = form.get("method", "GET").upper()
        action = form.get("action", "")[:40]
        # Count fields inside this form
        field_count = len(form.find_all(["input", "select", "textarea"]))
        comps["forms"].append(
            "Form" + str(i) + " [" + method + "] fields:" +
            str(field_count) + (" action:" + action if action else "")
        )

    # ── Labels (tells us what each field is for) ──────────────────────────
    seen_labels = set()
    for lbl in soup.find_all("label"):
        txt = lbl.get_text(strip=True)
        if txt and txt not in seen_labels and len(txt) > 1:
            comps["labels"].append(txt[:40])
            seen_labels.add(txt)

    # ── Input Fields (with label context) ────────────────────────────────
    seen_inp = set()
    for inp in soup.find_all("input"):
        itype = inp.get("type", "text").lower()
        if itype in ["hidden", "submit", "button", "reset"]:
            continue
        name        = inp.get("name", "")
        id_         = inp.get("id", "")
        placeholder = inp.get("placeholder", "")
        label_txt   = ""

        # Try to find associated label
        if id_:
            lbl = soup.find("label", {"for": id_})
            if lbl:
                label_txt = lbl.get_text(strip=True)

        display = label_txt or placeholder or name or id_ or itype
        k = itype + ":" + display
        if k not in seen_inp and display:
            comps["inputs"].append(itype.upper() + ": " + display[:35])
            seen_inp.add(k)

    # ── Textareas ─────────────────────────────────────────────────────────
    for ta in soup.find_all("textarea"):
        name = ta.get("name", "") or ta.get("id", "") or ta.get("placeholder", "")
        if name:
            comps["inputs"].append("TEXTAREA: " + name[:35])

    # ── Select Dropdowns ──────────────────────────────────────────────────
    for sel in soup.find_all("select"):
        name = sel.get("name", "") or sel.get("id", "")
        # Find label for this select
        lbl_txt = ""
        if sel.get("id"):
            lbl = soup.find("label", {"for": sel.get("id")})
            if lbl:
                lbl_txt = lbl.get_text(strip=True)
        display = lbl_txt or name
        opts    = [o.get_text(strip=True) for o in sel.find_all("option") if o.get_text(strip=True)][:5]
        comps["dropdowns"].append(
            "SELECT [" + display[:25] + "]: " + " | ".join(opts)
        )

    # ── Buttons ───────────────────────────────────────────────────────────
    seen_btn = set()
    for btn in soup.find_all(["button", "input"]):
        btype = btn.get("type", "").lower()
        if btype in ["submit", "button", "reset"] or btn.name == "button":
            label = (btn.get_text(strip=True) or
                     btn.get("value",       "") or
                     btn.get("name",        "") or btype)
            if label and label not in seen_btn and len(label) > 1:
                prefix = "[" + btype.upper() + "] " if btype else "[BTN] "
                comps["buttons"].append(prefix + label[:35])
                seen_btn.add(label)

    # ── Tables ────────────────────────────────────────────────────────────
    for i, tbl in enumerate(soup.find_all("table"), 1):
        hdrs = [th.get_text(strip=True) for th in tbl.find_all("th")][:6]
        rows_count = len(tbl.find_all("tr")) - 1
        comps["tables"].append(
            "Table" + str(i) +
            " [" + str(rows_count) + " rows] cols: " +
            " | ".join(hdrs) if hdrs else "Table" + str(i)
        )

    return comps


# ── STEP 9: Test every URL (ENHANCED) ────────────────────────────────────────
def test_urls(driver, rows, cfg):
    print("\n[5] Testing", len(rows), "URLs + extracting screen components...")
    results = []

    for i, row in enumerate(rows):
        main  = row["main"]
        sub   = row["sub"]
        url   = row["url"]
        issue = row["issue"]
        href  = row["href"]
        label = (main + " > " + sub) if sub else main

        print("\n  [" + str(i+1).rjust(3) + "/" + str(len(rows)) + "] " + label)
        print("       URL:", (url or href)[:70])

        # No usable URL
        if not url:
            print("       STATUS: PORTAL BUG (no valid URL)")
            results.append({
                "main": main, "sub": sub, "url": href,
                "status": "PORTAL BUG",
                "issue": issue if issue else "Broken URL: " + href,
                "load": "", "title": "Could not resolve URL",
                "headings": "", "sections": "",
                "forms": "", "labels": "", "inputs": "",
                "buttons": "", "tables": "", "dropdowns": "",
                "screenshot": ""
            })
            continue

        try:
            t0 = time.time()
            driver.get(url)

            # Wait for Angular to finish rendering
            time.sleep(2)
            try:
                WebDriverWait(driver, 8).until(
                    lambda d: d.execute_script(
                        "return document.readyState") == "complete"
                )
            except: pass
            time.sleep(1)  # extra wait for Angular data binding

            load  = round(time.time() - t0, 2)
            title = driver.title[:55]

            # Screenshot
            safe    = re.sub(r"[^\w\-]", "_", label)[:45]
            ss_name = str(i+1).zfill(3) + "_" + safe + ".png"
            driver.save_screenshot(os.path.join(cfg["screenshots"], ss_name))

            # Parse with BeautifulSoup after full render
            soup_p = BeautifulSoup(driver.page_source, "html.parser")
            body   = soup_p.get_text(separator=" ", strip=True).lower()
            comps  = extract_components(driver, soup_p)

            # Status
            err_kw = ["access denied", "not found", "unauthorized",
                      "404", "403", "500"]
            errors = [k for k in err_kw if k in body]
            if errors:
                status = "FAIL"
            elif "PORTAL BUG" in issue:
                status = "PASS*"
            else:
                status = "PASS"

            # Print what was found on screen
            if comps["headings"]:
                print("       Headings :", " | ".join(comps["headings"][:3]))
            if comps["forms"]:
                print("       Forms    :", " | ".join(comps["forms"]))
            if comps["inputs"]:
                print("       Fields   :", " | ".join(comps["inputs"][:5]))
            if comps["buttons"]:
                print("       Buttons  :", " | ".join(comps["buttons"][:4]))
            if comps["tables"]:
                print("       Tables   :", " | ".join(comps["tables"][:2]))
            print("       STATUS   :", status, "(" + str(load) + "s)")

            results.append({
                "main":      main,
                "sub":       sub,
                "url":       url,
                "status":    status,
                "issue":     issue,
                "load":      load,
                "title":     title,
                "headings":  " | ".join(comps["headings"])[:100],
                "sections":  " | ".join(comps["page_sections"])[:80],
                "forms":     " | ".join(comps["forms"])[:100],
                "labels":    " | ".join(comps["labels"])[:100],
                "inputs":    " | ".join(comps["inputs"])[:120],
                "buttons":   " | ".join(comps["buttons"])[:100],
                "tables":    " | ".join(comps["tables"])[:100],
                "dropdowns": " | ".join(comps["dropdowns"])[:100],
                "screenshot": ss_name,
            })

        except Exception as e:
            print("       STATUS: ERROR -", str(e)[:60])
            results.append({
                "main": main, "sub": sub, "url": url,
                "status": "ERROR", "issue": str(e)[:80],
                "load": "", "title": "",
                "headings": "", "sections": "", "forms": "",
                "labels": "", "inputs": "", "buttons": "",
                "tables": "", "dropdowns": "", "screenshot": ""
            })

    return results

# ── STEP 9: Test every URL ────────────────────────────────────────────────────
def test_urls(driver, rows, cfg):
    print("\n[5] Testing", len(rows), "URLs...")
    results = []

    for i, row in enumerate(rows):
        main  = row["main"]
        sub   = row["sub"]
        url   = row["url"]
        issue = row["issue"]
        href  = row["href"]
        label = (main + " > " + sub) if sub else main

        print("  [" + str(i+1).rjust(3) + "/" + str(len(rows)) + "]",
              label[:40].ljust(40), end=" ... ", flush=True)

        # No usable URL — record as PORTAL BUG but still include in report
        if not url:
            print("PORTAL BUG (no valid URL)")
            results.append({
                "main":      main,
                "sub":       sub,
                "url":       href,          # show original broken href
                "status":    "PORTAL BUG",
                "issue":     issue if issue else "Broken or empty URL: " + href,
                "load":      "",
                "title":     "Could not resolve URL",
                "forms":     "", "buttons":   "",
                "inputs":    "", "tables":    "",
                "dropdowns": "", "screenshot": ""
            })
            continue

        try:
            t0 = time.time()
            driver.get(url)
            time.sleep(1.5)
            load  = round(time.time() - t0, 2)
            title = driver.title[:55]

            # Screenshot
            safe    = re.sub(r"[^\w\-]", "_", label)[:45]
            ss_name = str(i+1).zfill(3) + "_" + safe + ".png"
            driver.save_screenshot(os.path.join(cfg["screenshots"], ss_name))

            # Parse components
            soup_p = BeautifulSoup(driver.page_source, "html.parser")
            body   = soup_p.get_text(separator=" ", strip=True).lower()
            # comps  = extract_components(soup_p)
            comps  = extract_components(driver, soup_p)

            # Determine status
            err_kw = ["access denied", "not found", "unauthorized",
                      "404", "403", "500"]
            errors = [k for k in err_kw if k in body]

            if errors:
                status = "FAIL"
            elif "PORTAL BUG" in issue:
                status = "PASS*"   # working but URL was partially fixed
            else:
                status = "PASS"

            suffix = " [PORTAL BUG-partial fix]" if "PORTAL BUG" in issue else ""
            print(status + " (" + str(load) + "s)" + suffix)

            results.append({
                "main":      main,
                "sub":       sub,
                "url":       url,
                "status":    status,
                "issue":     issue,
                "load":      load,
                "title":     title,
                "forms":     " | ".join(comps["forms"])[:80],
                "buttons":   " | ".join(comps["buttons"])[:80],
                "inputs":    " | ".join(comps["inputs"])[:80],
                "tables":    " | ".join(comps["tables"])[:80],
                "dropdowns": " | ".join(comps["dropdowns"])[:80],
                "screenshot": ss_name,
            })

        except Exception as e:
            print("ERROR")
            results.append({
                "main":      main,
                "sub":       sub,
                "url":       url,
                "status":    "ERROR",
                "issue":     str(e)[:80],
                "load":      "", "title":     "",
                "forms":     "", "buttons":   "",
                "inputs":    "", "tables":    "",
                "dropdowns": "", "screenshot": ""
            })

    return results

# ── STEP 10: Excel report ─────────────────────────────────────────────────────
def write_excel(rows, results, cfg):
    print("\n[6] Writing Excel report...")

    def fill(c):
        return PatternFill("solid", fgColor=c)
    def fnt(bold=False, color="FF000000", sz=10):
        return Font(bold=bold, color=color, size=sz)
    def brd():
        s = Side(style="thin", color="FFD0D0D0")
        return Border(left=s, right=s, top=s, bottom=s)
    def aln(h="left"):
        return Alignment(horizontal=h, vertical="center", wrap_text=True)

    wb = Workbook()

    # ── Sheet 1: Menu Structure ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Menu Structure"
    s1_cols = ["#", "Level 1 (Main Menu)", "Level 2 (Sub Menu)", "URL", "Issue"]
    s1_wids = [5, 28, 35, 65, 45]
    for ci, (h, w) in enumerate(zip(s1_cols, s1_wids), 1):
        c = ws1.cell(1, ci, h)
        c.fill = fill("FF003366"); c.font = fnt(True, "FFFFFFFF", 11)
        c.alignment = aln("center"); c.border = brd()
        ws1.column_dimensions[c.column_letter].width = w

    for ri, row in enumerate(rows, 2):
        is_sub = bool(row["sub"])
        if row["issue"]:
            bg = "FFFFF3CD"   # yellow for issues
        elif not is_sub:
            bg = "FFE8F4FD"   # blue tint for main
        else:
            bg = "FFFAFAFA" if ri % 2 == 0 else "FFF5F5F5"

        vals = [ri-1, row["main"], row["sub"], row["url"] or row["href"], row["issue"]]
        for ci, val in enumerate(vals, 1):
            c = ws1.cell(ri, ci, val)
            c.fill = fill(bg); c.border = brd(); c.alignment = aln()
            if ci == 4:
                c.font = Font(color="FF0055AA", size=10, underline="single")
            elif ci == 5 and val:
                c.font = fnt(color="FFB71C1C", bold=True)
            else:
                c.font = fnt(bold=not is_sub and ci == 2)

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = "A1:E" + str(len(rows)+1)

    # ── Sheet 2: Test Results ────────────────────────────────────────────────
    # ── Sheet 2: Test Results ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Test Results")
    s2_cols = ["#", "Main Menu", "Sub Menu", "URL", "Result", "Issue",
               "Load(s)", "Page Title", "Page Headings", "Page Sections",
               "Forms", "Field Labels", "Input Fields",
               "Buttons", "Tables", "Dropdowns", "Screenshot"]
    s2_wids = [4, 18, 22, 45, 10, 35, 7, 30, 40, 30,
               25, 40, 45, 35, 40, 35, 30]

    for ci, (h, w) in enumerate(zip(s2_cols, s2_wids), 1):
        c = ws2.cell(1, ci, h)
        c.fill = fill("FF005B96"); c.font = fnt(True, "FFFFFFFF", 11)
        c.alignment = aln("center"); c.border = brd()
        ws2.column_dimensions[c.column_letter].width = w

    pass_c = fail_c = err_c = bug_c = 0
    for ri, r in enumerate(results, 2):
        st = r["status"]
        if   st == "PASS":       sf = fill("FFE8F5E9"); ff = fnt(True, "FF1B5E20")
        elif st == "PASS*":      sf = fill("FFF1F8E9"); ff = fnt(True, "FF33691E")
        elif st == "FAIL":       sf = fill("FFFFF3E0"); ff = fnt(True, "FFE65100")
        elif st == "PORTAL BUG": sf = fill("FFFCE4EC"); ff = fnt(True, "FFC62828")
        else:                    sf = fill("FFFFEBEE"); ff = fnt(True, "FFB71C1C")

        if   st in ["PASS", "PASS*"]: pass_c += 1
        elif st == "FAIL":            fail_c += 1
        elif st == "PORTAL BUG":      bug_c  += 1
        else:                         err_c  += 1

        bg = "FFFAFAFA" if ri % 2 == 0 else "FFFFFFFF"
        vals = [
            ri-1, r["main"], r["sub"],
            r["url"] or r.get("href", ""),
            st, r["issue"], r["load"], r["title"],
            r.get("headings", ""), r.get("sections", ""),
            r["forms"], r.get("labels", ""), r["inputs"],
            r["buttons"], r["tables"], r["dropdowns"],
            r["screenshot"]
        ]
        for ci, val in enumerate(vals, 1):
            c = ws2.cell(ri, ci, val)
            c.fill = sf if ci == 5 else fill(bg)
            c.font = (ff if ci == 5
                      else fnt(color="FFB71C1C", bold=True) if ci == 6 and val
                      else fnt())
            c.alignment = aln(); c.border = brd()

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = "A1:Q" + str(len(results)+1)

    # ── Sheet 3: Summary ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 30

    # summary = [
    #     ("Report Generated",  datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ""),
    #     ("Portal URL",        cfg["url"],      ""),
    #     ("Login User",        cfg["username"], ""),
    #     ("Output Folder",     cfg["folder"],   ""),
    #     ("", "", ""),
    #     ("Total Menu Items",  len(rows),       ""),
    #     ("Total URLs Tested", len(results),    ""),
    #     ("", "", ""),
    #     ("PASS",              pass_c,  "FF1B5E20"),
    #     ("FAIL",              fail_c,  "FFE65100"),
    #     ("PORTAL BUG",        bug_c,   "FFC62828"),
    #     ("ERROR",             err_c,   "FFB71C1C"),
    # ]
    # for ri, (lbl, val, color) in enumerate(summary, 1):
    #     l = ws3.cell(ri, 1, lbl)
    #     v = ws3.cell(ri, 2, str(val))
    #     l.font = fnt(bold=True, sz=11)
    #     # v.font = fnt(color="FF"+color, sz=12, bold=bool(color)) if color else fnt(sz=11)
    #     v.font = fnt(color="FF"+color, sz=12, bold=True) if color else fnt(sz=11)
    summary = [
        ("Report Generated",  datetime.now().strftime("%Y-%m-%d %H:%M:%S"), None),
        ("Portal URL",        cfg["url"],      None),
        ("Login User",        cfg["username"], None),
        ("Output Folder",     cfg["folder"],   None),
        ("", "",              None),
        ("Total Menu Items",  len(rows),       None),
        ("Total URLs Tested", len(results),    None),
        ("", "",              None),
        ("PASS",              pass_c,  "FF1B5E20"),
        ("FAIL",              fail_c,  "FFE65100"),
        ("PORTAL BUG",        bug_c,   "FFC62828"),
        ("ERROR",             err_c,   "FFB71C1C"),
    ]
    for ri, (lbl, val, color) in enumerate(summary, 1):
        l = ws3.cell(ri, 1, lbl)
        v = ws3.cell(ri, 2, str(val))
        l.font = fnt(bold=True, sz=11)
        if color:
            v.font = fnt(color=color, sz=12, bold=True)
        else:
            v.font = fnt(sz=11)
    wb.save(cfg["excel"])
    print("    Saved:", os.path.abspath(cfg["excel"]))
    print("\n    PASS:", pass_c, "| FAIL:", fail_c,
          "| PORTAL BUG:", bug_c, "| ERROR:", err_c)

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    cfg = get_config()
    if not cfg:
        print("Cancelled.")
        return

    print("\n" + "="*52)
    print("  User  :", cfg["username"])
    print("  Folder:", cfg["folder"])
    print("="*52)

    driver = init_driver()
    try:
        login(driver, cfg)
        html    = get_html(driver, cfg)
        rows    = parse_menu(html, cfg)
        save_csv(rows, cfg)
        results = test_urls(driver, rows, cfg)
        write_excel(rows, results, cfg)

        print("\n" + "="*52)
        print("  DONE!")
        print("  Folder:", cfg["folder"])
        print("  CSV   :", cfg["csv"])
        print("  Excel :", cfg["excel"])
        print("="*52)

    except Exception as e:
        print("\n[ERROR]", e)
        import traceback; traceback.print_exc()
    finally:
        input("\nPress ENTER to close browser...")
        driver.quit()

if __name__ == "__main__":
    main()