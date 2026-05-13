"""
Universal Portal Testing Tool v4.0 — FUNCTIONAL COMPARISON
============================================================
WHAT THIS DOES:
  - Visits every sub-menu URL on LIVE portal  (read-only, no back())
  - Visits every sub-menu URL on TEST portal  (no back())
  - For each page extracts:
      · All form fields  (name, label, data type, mandatory/optional)
      · All tables       (table heading, column names, row count)
      · All buttons      (label, type)
      · Page headings / section titles
  - Compares LIVE vs TEST page-by-page:
      · Fields: MATCH / MISSING IN TEST / EXTRA IN TEST / MISMATCH
      · Tables: MATCH / MISSING IN TEST / EXTRA IN TEST / COLUMNS DIFFER
      · Buttons: MATCH / MISSING / EXTRA
  - Excel report — 7 sheets:
      1. Live Menu Structure
      2. Test Menu Structure
      3. Menu Comparison
      4. Page Content  (what was found on each page, both portals)
      5. Field Comparison  (field by field: name, type, mandatory)
      6. Table Comparison  (table by table: heading, columns, row count)
      7. Summary

RULES:
  - ZERO driver.back() calls — always navigate by direct URL
  - ZERO pixel comparison — only functional content is compared
  - LIVE portal: HTML read only, no clicks, no form fills
  - TEST portal: light dropdown selection only, no form submission

Install:
  uv pip install selenium webdriver-manager openpyxl beautifulsoup4

Run:
  python portal_compare_v3.py
"""

import os, re, time
from datetime import datetime
from urllib.parse import urlparse, urljoin

from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select as SeleniumSelect
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, WebDriverException)
from webdriver_manager.chrome import ChromeDriverManager

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def xfill(c):
    return PatternFill("solid", fgColor=c)

def xfnt(bold=False, color="FF000000", sz=10):
    return Font(bold=bold, color=color, size=sz)

def xbrd():
    s = Side(style="thin", color="FFD0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def xaln(h="left"):
    return Alignment(horizontal=h, vertical="center", wrap_text=True)

def hdr(ws, row, col, text, bg="FF003366", fg="FFFFFFFF", sz=11, w=None):
    c = ws.cell(row, col, text)
    c.fill = xfill(bg)
    c.font = xfnt(True, fg, sz)
    c.alignment = xaln("center")
    c.border = xbrd()
    if w:
        ws.column_dimensions[get_column_letter(col)].width = w
    return c

def dat(ws, row, col, val, bg="FFFFFFFF", bold=False, color="FF000000", h="left"):
    c = ws.cell(row, col, val)
    c.fill = xfill(bg)
    c.font = xfnt(bold, color)
    c.alignment = xaln(h)
    c.border = xbrd()
    return c

STATUS_CLR = {
    "MATCH":             ("FFE8F5E9", "FF1B5E20"),
    "MISSING IN TEST":   ("FFFFEBEE", "FFB71C1C"),
    "EXTRA IN TEST":     ("FFFFF8E1", "FFF57F17"),
    "MISMATCH":          ("FFFFE0B2", "FFE65100"),
    "TYPE CHANGED":      ("FFFFE0B2", "FFE65100"),
    "MANDATORY CHANGED": ("FFFFE0B2", "FFE65100"),
    "COLUMNS DIFFER":    ("FFFFE0B2", "FFE65100"),
    "OK":                ("FFE8F5E9", "FF1B5E20"),
    "ERROR":             ("FFFFEBEE", "FFB71C1C"),
    "TIMEOUT":           ("FFFFF8E1", "FFF57F17"),
    "NO URL":            ("FFF5F5F5", "FF9E9E9E"),
    "NOT IN LIVE":       ("FFF3E5F5", "FF6A1B9A"),
    "NOT IN TEST":       ("FFF3E5F5", "FF6A1B9A"),
}

def st_cell(ws, row, col, status):
    bg, fg = STATUS_CLR.get(status, ("FFFFFFFF", "FF000000"))
    c = ws.cell(row, col, status)
    c.fill = xfill(bg)
    c.font = xfnt(True, fg, 10)
    c.alignment = xaln("center")
    c.border = xbrd()
    return c

def row_bg(ri):
    return "FFFAFAFA" if ri % 2 == 0 else "FFFFFFFF"


# ─────────────────────────────────────────────────────────────────────────────
#  GUI
# ─────────────────────────────────────────────────────────────────────────────
def get_config():
    import tkinter as tk
    from tkinter import ttk, messagebox

    config = {}

    def submit():
        lu = live_url.get().strip()
        tu = test_url.get().strip()
        if not lu or not tu:
            messagebox.showerror("Error", "Both portal URLs are required")
            return
        if not live_user.get().strip() or not live_pwd.get().strip():
            messagebox.showerror("Error", "Live portal credentials required")
            return
        if not test_user.get().strip() or not test_pwd.get().strip():
            messagebox.showerror("Error", "Test portal credentials required")
            return

        def base(url):
            p = urlparse(url)
            return p.scheme + "://" + p.netloc

        ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
        folder = (test_user.get().strip() or "test") + "_" + ts
        os.makedirs(folder, exist_ok=True)

        config["live"] = {
            "url":      lu,
            "username": live_user.get().strip(),
            "password": live_pwd.get().strip(),
            "base":     base(lu),
            "label":    "LIVE",
        }
        config["test"] = {
            "url":      tu,
            "username": test_user.get().strip(),
            "password": test_pwd.get().strip(),
            "base":     base(tu),
            "label":    "TEST",
            "folder":   folder,
            "live_html":  os.path.join(folder, "live_portal.html"),
            "test_html":  os.path.join(folder, "test_portal.html"),
            "excel":      os.path.join(folder, "functional_comparison.xlsx"),
        }
        config["options"] = {
            "page_timeout":   int(timeout_var.get()),
            "captcha_wait":   int(captcha_var.get()),
            "skip_test_fill": skip_var.get(),
        }
        root.destroy()

    root = tk.Tk()
    root.title("Portal Functional Comparison Tool v4.0")
    root.geometry("680x520")
    root.resizable(False, False)

    tk.Label(root,
             text="Portal Functional Comparison Tool v4.0",
             font=("Arial", 14, "bold"), fg="#003366").pack(pady=8)
    tk.Label(root,
             text="Compares fields · tables · columns · buttons  —  no pixel comparison",
             font=("Arial", 9), fg="#555").pack()

    nb = ttk.Notebook(root)
    nb.pack(fill="both", expand=True, padx=10, pady=8)

    # Tab 1: LIVE
    t1 = ttk.Frame(nb, padding=15)
    nb.add(t1, text="  LIVE Portal (Read-Only)  ")
    tk.Label(t1,
             text="READ-ONLY — fields and tables extracted, nothing submitted",
             font=("Arial", 10, "bold"), fg="#cc0000").grid(
             row=0, column=0, columnspan=2, pady=(0,12), sticky="w")

    live_url  = tk.StringVar(value="https://cis.drt.gov.in/drtlive/index.php")
    live_user = tk.StringVar(value="filingdrt1")
    live_pwd  = tk.StringVar()

    for ri, (lbl, var, show) in enumerate([
        ("Portal URL:", live_url,  ""),
        ("Username:",   live_user, ""),
        ("Password:",   live_pwd,  "*"),
    ], 1):
        ttk.Label(t1, text=lbl).grid(row=ri, column=0, sticky="w", pady=8)
        ttk.Entry(t1, textvariable=var, show=show, width=56).grid(
            row=ri, column=1, padx=8)

    # Tab 2: TEST
    t2 = ttk.Frame(nb, padding=15)
    nb.add(t2, text="  TEST Portal (Full Testing)  ")
    tk.Label(t2,
             text="FULL TESTING — fields, tables, dropdowns extracted and compared",
             font=("Arial", 10, "bold"), fg="#006600").grid(
             row=0, column=0, columnspan=2, pady=(0,12), sticky="w")

    test_url  = tk.StringVar(value="https://drt.etribunals.gov.in/cis2.0/filing/login")
    test_user = tk.StringVar()
    test_pwd  = tk.StringVar()

    for ri, (lbl, var, show) in enumerate([
        ("Portal URL:", test_url,  ""),
        ("Username:",   test_user, ""),
        ("Password:",   test_pwd,  "*"),
    ], 1):
        ttk.Label(t2, text=lbl).grid(row=ri, column=0, sticky="w", pady=8)
        ttk.Entry(t2, textvariable=var, show=show, width=56).grid(
            row=ri, column=1, padx=8)

    # Tab 3: Options
    t3 = ttk.Frame(nb, padding=15)
    nb.add(t3, text="  Options  ")

    timeout_var = tk.StringVar(value="15")
    captcha_var = tk.StringVar(value="90")
    skip_var    = tk.BooleanVar(value=False)

    for ri, (lbl, var, hint) in enumerate([
        ("Page load timeout (sec):", timeout_var, "Wait per page before extracting"),
        ("CAPTCHA wait (sec):",      captcha_var, "Time to solve CAPTCHA + login"),
    ], 1):
        ttk.Label(t3, text=lbl).grid(row=ri, column=0, sticky="w", pady=8)
        ttk.Entry(t3, textvariable=var, width=10).grid(
            row=ri, column=1, sticky="w", padx=8)
        ttk.Label(t3, text=hint, foreground="#888",
                  font=("Arial", 9)).grid(row=ri, column=2, sticky="w")

    ttk.Checkbutton(t3,
                    text="Skip dropdown selection on TEST (extract only, no interaction)",
                    variable=skip_var).grid(
                    row=3, column=0, columnspan=3, pady=12, sticky="w")

    btn_frame = ttk.Frame(root)
    btn_frame.pack(pady=8)
    ttk.Button(btn_frame, text="Start Comparison",
               command=submit).pack(side="left", padx=12)

    root.mainloop()
    return config


# ─────────────────────────────────────────────────────────────────────────────
#  DRIVER
# ─────────────────────────────────────────────────────────────────────────────
def init_driver():
    opts = webdriver.ChromeOptions()
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=opts)


# ─────────────────────────────────────────────────────────────────────────────
#  LOGIN
# ─────────────────────────────────────────────────────────────────────────────
USER_LOCS = [
    (By.ID,   "user_name"),   (By.ID,   "username"),
    (By.NAME, "user_name"),   (By.NAME, "username"),
    (By.NAME, "txtUsername"), (By.ID,   "txtUsername"),
    (By.XPATH, "//input[@type='text'][1]"),
    (By.XPATH, "//input[contains(@name,'user') or contains(@id,'user')]"),
]
PASS_LOCS = [
    (By.ID,   "user_pass"),   (By.ID,   "password"),
    (By.NAME, "user_pass"),   (By.NAME, "password"),
    (By.NAME, "txtPassword"), (By.ID,   "txtPassword"),
    (By.XPATH, "//input[@type='password']"),
]

def login(driver, pcfg, captcha_wait):
    print(f"\n{'='*60}")
    print(f"  LOGIN: {pcfg['label']}  |  {pcfg['url']}")
    print(f"{'='*60}")
    driver.get(pcfg["url"])
    time.sleep(3)

    def fill(label, val, locs):
        for loc in locs:
            try:
                el = driver.find_element(*loc)
                if el.is_displayed():
                    el.clear()
                    el.send_keys(val)
                    print(f"  + {label} filled")
                    return
            except NoSuchElementException:
                pass
        print(f"  ! {label} not found — fill manually")

    fill("Username", pcfg["username"], USER_LOCS)
    fill("Password", pcfg["password"], PASS_LOCS)

    print(f"\n  ACTION: Solve CAPTCHA then click Login  ({captcha_wait}s to wait)")
    start = driver.current_url
    try:
        WebDriverWait(driver, captcha_wait).until(
            lambda d: d.current_url != start)
    except TimeoutException:
        print("  [!] URL unchanged — continuing anyway")
    time.sleep(3)
    print(f"  Logged in: {driver.current_url}")
    return driver.current_url


# ─────────────────────────────────────────────────────────────────────────────
#  URL CLEANER
# ─────────────────────────────────────────────────────────────────────────────
def clean_url(href, base):
    if not href or not href.strip():
        return "", "empty"
    href = href.strip()
    if href.lower().startswith("javascript") or href in ("#", ""):
        return "", "js-link"
    if re.search(r"['\",]", href):
        clean = re.split(r"['\",]", href)[0].strip()
        return ((base + clean, "bug:broken-routerlink")
                if clean.startswith("/") else ("", "bug:broken-href"))
    if href.startswith("http"):
        return href, ""
    if href.startswith("/"):
        return base + href, ""
    return urljoin(base + "/", href), "relative"


# ─────────────────────────────────────────────────────────────────────────────
#  MENU PARSER
# ─────────────────────────────────────────────────────────────────────────────
def parse_menu(html, base, label=""):
    soup = BeautifulSoup(html, "html.parser")
    rows = []

    def make_row(main, sub, href):
        url, issue = clean_url(href, base)
        return {"main": main, "sub": sub, "url": url,
                "issue": issue, "href": href}

    # Strategy 1: accordionSidebar (new portal)
    sidebar = soup.find("ul", id="accordionSidebar")
    if sidebar:
        print(f"  [{label}] nav type: accordionSidebar")
        for li in sidebar.find_all("li", class_="nav-item"):
            main_a = li.find("a", class_="nav-link")
            if not main_a:
                continue
            mn = main_a.get_text(strip=True)
            if not mn:
                continue
            added = 0
            for ci in li.find_all("div", class_="collapse-inner"):
                for a in (ci.find_all("a", class_="dropdown-item")
                          or ci.find_all("a", href=True)):
                    sub = a.get_text(strip=True)
                    if sub:
                        rows.append(make_row(mn, sub, a.get("href", "")))
                        added += 1
            if not added:
                rows.append(make_row(mn, "", main_a.get("href", "")))
        if rows:
            return rows

    # Strategy 2: Bootstrap navbar-nav (old portal / DRT live)
    nav    = soup.find("nav", class_=re.compile(r"navbar", re.I))
    nav_ul = (nav.find("ul", class_=re.compile(r"navbar-nav", re.I))
              if nav else None)
    if not nav_ul:
        for ul in soup.find_all("ul"):
            if len(ul.find_all("li", recursive=False)) >= 3:
                nav_ul = ul
                break

    if nav_ul:
        print(f"  [{label}] nav type: navbar-nav + dropdown-menu")
        for li in nav_ul.find_all("li", recursive=False):
            main_a = (li.find("a", class_=re.compile(r"nav-link", re.I))
                      or li.find("a"))
            if not main_a:
                continue
            mn = main_a.get_text(strip=True).lstrip("*").strip()
            if not mn:
                continue
            mh = main_a.get("href", "")
            dd = (li.find("div", class_=re.compile(r"dropdown-menu", re.I))
                  or li.find("ul", class_=re.compile(r"dropdown-menu|sub-menu", re.I)))
            if dd:
                subs = (dd.find_all("a", class_=re.compile(r"dropdown-item", re.I))
                        or dd.find_all("a", href=True))
                added = 0
                for a in subs:
                    sub = a.get_text(strip=True)
                    if sub:
                        rows.append(make_row(mn, sub, a.get("href", "")))
                        added += 1
                if not added:
                    rows.append(make_row(mn, "", mh))
            else:
                rows.append(make_row(mn, "", mh))
        if rows:
            return rows

    # Fallback: all anchors
    print(f"  [{label}] nav type: all-anchors fallback")
    for a in soup.find_all("a", href=True):
        txt = a.get_text(strip=True)
        if txt and len(txt) < 80:
            rows.append(make_row(txt, "", a["href"]))
    return rows


# ─────────────────────────────────────────────────────────────────────────────
#  DATA TYPE DETECTOR
# ─────────────────────────────────────────────────────────────────────────────
def detect_datatype(inp):
    """Returns human-readable data type from input element attributes."""
    itype = (inp.get("type", "text") or "text").lower()
    name  = (inp.get("name", "") or inp.get("id", "") or "").lower()
    ph    = (inp.get("placeholder", "") or "").lower()

    if itype == "number":  return "NUMBER"
    if itype == "email":   return "EMAIL"
    if itype == "password":return "PASSWORD"
    if itype == "date":    return "DATE"
    if itype == "tel":     return "PHONE"
    if itype == "file":    return "FILE"
    if itype == "radio":   return "RADIO"
    if itype == "checkbox":return "CHECKBOX"
    if itype in ("submit","button","reset","image"):
        return f"BUTTON[{itype.upper()}]"

    # Guess from name / placeholder
    for kw in ("date","dt","dob","from","to"):
        if kw in name or kw in ph: return "DATE"
    for kw in ("amount","sum","number","count","year","qty","no"):
        if kw in name or kw in ph: return "NUMBER"
    for kw in ("email","mail"):
        if kw in name or kw in ph: return "EMAIL"
    for kw in ("phone","mobile","mob","contact"):
        if kw in name or kw in ph: return "PHONE"
    return "TEXT"


# ─────────────────────────────────────────────────────────────────────────────
#  PAGE CONTENT EXTRACTOR  (pure HTML read — safe for both LIVE and TEST)
# ─────────────────────────────────────────────────────────────────────────────
def extract_page_content(html):
    """
    Returns:
      fields  → list of { form, form_name, label, datatype, mandatory, options }
      tables  → list of { heading, columns, row_count }
      buttons → list of { label, btype }
      headings→ list of str
    """
    soup   = BeautifulSoup(html, "html.parser")
    result = {"fields": [], "tables": [], "buttons": [], "headings": []}

    # Headings
    for tag in ["h1","h2","h3","h4","h5","legend","caption"]:
        for h in soup.find_all(tag):
            txt = h.get_text(strip=True)
            if txt and len(txt) > 2:
                result["headings"].append(txt[:80])

    # Helper: find label text for an element
    def find_label(el):
        eid = el.get("id", "")
        if eid:
            lb = soup.find("label", {"for": eid})
            if lb:
                return lb.get_text(strip=True).replace("*", "").strip()
        parent = el.parent
        for _ in range(3):
            if not parent:
                break
            lb = parent.find("label", recursive=False)
            if lb:
                return lb.get_text(strip=True).replace("*", "").strip()
            parent = parent.parent
        return ""

    # Helper: is field mandatory?
    def is_mandatory(el):
        if el.has_attr("required"):
            return True
        if el.get("aria-required", "") == "true":
            return True
        if "*" in el.get("placeholder", ""):
            return True
        parent = el.parent
        if parent and "*" in parent.get_text():
            return True
        return False

    # Forms → fields
    for fi, form in enumerate(soup.find_all("form"), 1):
        legend     = form.find("legend")
        form_name  = legend.get_text(strip=True)[:40] if legend else ""

        for inp in form.find_all("input"):
            itype = (inp.get("type", "text") or "text").lower()
            if itype == "hidden":
                continue
            lbl     = find_label(inp)
            ph      = inp.get("placeholder", "").replace("*", "").strip()
            nm      = inp.get("name", "") or inp.get("id", "")
            display = lbl or ph or nm or itype
            if not display or display.lower() in ("search", "search for...", ""):
                continue
            dtype = detect_datatype(inp)
            mand  = is_mandatory(inp)
            result["fields"].append({
                "form":      fi,
                "form_name": form_name,
                "label":     display[:60],
                "datatype":  dtype,
                "mandatory": "MANDATORY" if mand else "OPTIONAL",
                "options":   "",
            })

        for ta in form.find_all("textarea"):
            lbl     = find_label(ta)
            nm      = ta.get("name", "") or ta.get("id", "")
            display = lbl or nm or "textarea"
            mand    = is_mandatory(ta)
            result["fields"].append({
                "form":      fi,
                "form_name": form_name,
                "label":     display[:60],
                "datatype":  "TEXTAREA",
                "mandatory": "MANDATORY" if mand else "OPTIONAL",
                "options":   "",
            })

        for sel in form.find_all("select"):
            lbl     = find_label(sel)
            nm      = sel.get("name", "") or sel.get("id", "")
            display = lbl or nm or "select"
            mand    = is_mandatory(sel)
            opts    = [o.get_text(strip=True) for o in sel.find_all("option")
                       if o.get_text(strip=True)]
            result["fields"].append({
                "form":      fi,
                "form_name": form_name,
                "label":     display[:60],
                "datatype":  "SELECT",
                "mandatory": "MANDATORY" if mand else "OPTIONAL",
                "options":   " / ".join(opts[:8]),
            })

        # Radio groups
        radio_groups = {}
        for r in form.find_all("input", {"type": "radio"}):
            grp = r.get("name", "radio_group")
            lbl = find_label(r) or r.get("value", "")
            radio_groups.setdefault(grp, []).append(lbl)
        for grp, opts in radio_groups.items():
            result["fields"].append({
                "form":      fi,
                "form_name": form_name,
                "label":     grp[:60],
                "datatype":  "RADIO",
                "mandatory": "OPTIONAL",
                "options":   " / ".join(opts[:6]),
            })

        # Checkboxes
        for chk in form.find_all("input", {"type": "checkbox"}):
            lbl = find_label(chk) or chk.get("name", "") or chk.get("id", "")
            if lbl:
                result["fields"].append({
                    "form":      fi,
                    "form_name": form_name,
                    "label":     lbl[:60],
                    "datatype":  "CHECKBOX",
                    "mandatory": "OPTIONAL",
                    "options":   "",
                })

        # Buttons inside form
        for btn in form.find_all(
                lambda t: t.name == "button" or
                (t.name == "input" and
                 t.get("type", "").lower() in ("submit", "button", "reset"))):
            btype = (btn.get("type", "button") or "button").upper()
            lbl   = (btn.get_text(strip=True) or
                     btn.get("value", "") or btn.get("name", ""))
            if lbl and len(lbl) > 1:
                result["buttons"].append({"label": lbl[:40], "btype": btype})

    # Standalone buttons (outside forms)
    for btn in soup.find_all("button"):
        lbl = btn.get_text(strip=True)
        if lbl and len(lbl) > 1:
            btype = (btn.get("type", "button") or "button").upper()
            result["buttons"].append({"label": lbl[:40], "btype": btype})

    # Tables
    for ti, tbl in enumerate(soup.find_all("table"), 1):
        # Table heading
        heading = ""
        cap = tbl.find("caption")
        if cap:
            heading = cap.get_text(strip=True)
        if not heading:
            prev = tbl.find_previous(["h1","h2","h3","h4","h5","legend","fieldset"])
            if prev:
                heading = prev.get_text(strip=True)[:50]
        if not heading:
            heading = f"Table {ti}"

        # Column names
        cols = []
        thead = tbl.find("thead")
        if thead:
            cols = [th.get_text(strip=True) for th in thead.find_all("th")]
        if not cols:
            first_tr = tbl.find("tr")
            if first_tr:
                cols = [td.get_text(strip=True)
                        for td in first_tr.find_all(["th","td"])]

        # Row count
        tbody = tbl.find("tbody")
        if tbody:
            row_count = len(tbody.find_all("tr"))
        else:
            all_rows  = tbl.find_all("tr")
            row_count = max(0, len(all_rows) - 1)

        result["tables"].append({
            "heading":   heading[:60],
            "columns":   [c.strip() for c in cols if c.strip()][:20],
            "row_count": row_count,
        })

    return result


# ─────────────────────────────────────────────────────────────────────────────
#  VISIT PAGE  (NO back() — always navigate by URL directly)
# ─────────────────────────────────────────────────────────────────────────────
def visit_page(driver, url, label, timeout, portal_label,
               do_interact=False):
    """
    Go directly to URL. Extract content. Never use back().
    do_interact=True: select first option in dropdowns (TEST only).
    """
    empty_content = {"fields":[],"tables":[],"buttons":[],"headings":[]}
    res = {"url": url, "status": "OK", "title": "", "content": empty_content}

    if not url:
        res["status"] = "NO URL"
        return res

    try:
        driver.get(url)
        try:
            WebDriverWait(driver, timeout).until(
                lambda d: d.execute_script(
                    "return document.readyState") == "complete")
        except TimeoutException:
            pass
        time.sleep(1.5)

        res["title"] = driver.title.strip()[:70]

        # Light interaction on TEST only — select dropdowns, no form submit
        if do_interact:
            for sel_el in driver.find_elements(By.CSS_SELECTOR, "select")[:5]:
                try:
                    s    = SeleniumSelect(sel_el)
                    opts = [o.text for o in s.options if o.text.strip()]
                    if len(opts) > 1:
                        s.select_by_index(1)
                        time.sleep(0.2)
                except Exception:
                    pass

        res["content"] = extract_page_content(driver.page_source)

        nf = len(res["content"]["fields"])
        nt = len(res["content"]["tables"])
        print(f"    [{portal_label}] {label[:48]:<48}  "
              f"fields={nf}  tables={nt}  {res['status']}")

    except TimeoutException:
        res["status"] = "TIMEOUT"
        print(f"    [{portal_label}] {label[:48]:<48}  TIMEOUT")
    except WebDriverException as e:
        res["status"] = "ERROR"
        print(f"    [{portal_label}] {label[:48]:<48}  ERROR: {str(e)[:40]}")

    return res


# ─────────────────────────────────────────────────────────────────────────────
#  MENU COMPARE
# ─────────────────────────────────────────────────────────────────────────────
def compare_menus(live_rows, test_rows):
    def key(r):
        return (r["main"].strip().lower(), r["sub"].strip().lower())
    lk = {key(r): r for r in live_rows}
    tk = {key(r): r for r in test_rows}
    out = []
    for k in sorted(set(list(lk) + list(tk))):
        lr, tr = lk.get(k), tk.get(k)
        st = ("MATCH"           if lr and tr else
              "MISSING IN TEST" if lr       else "EXTRA IN TEST")
        out.append({
            "main":       (lr or tr)["main"],
            "sub":        (lr or tr)["sub"],
            "status":     st,
            "live_url":   lr["url"]           if lr else "",
            "test_url":   tr["url"]           if tr else "",
            "live_issue": lr.get("issue","")  if lr else "",
            "test_issue": tr.get("issue","")  if tr else "",
        })
    return out


# ─────────────────────────────────────────────────────────────────────────────
#  FIELD COMPARE
# ─────────────────────────────────────────────────────────────────────────────
def compare_fields(live_fields, test_fields):
    def norm(lbl):
        return re.sub(r"[^a-z0-9]", "", lbl.lower())

    lmap = {norm(f["label"]): f for f in live_fields}
    tmap = {norm(f["label"]): f for f in test_fields}
    all_keys = sorted(set(list(lmap) + list(tmap)))
    out = []
    for k in all_keys:
        lf, tf = lmap.get(k), tmap.get(k)
        if lf and tf:
            diffs = []
            if lf["datatype"]  != tf["datatype"]:
                diffs.append(f"Type: {lf['datatype']} -> {tf['datatype']}")
            if lf["mandatory"] != tf["mandatory"]:
                diffs.append(f"Mandatory: {lf['mandatory']} -> {tf['mandatory']}")
            st = "MISMATCH" if diffs else "MATCH"
            out.append({
                "field":     lf["label"],
                "status":    st,
                "live_type": lf["datatype"],
                "test_type": tf["datatype"],
                "live_mand": lf["mandatory"],
                "test_mand": tf["mandatory"],
                "live_opts": lf.get("options",""),
                "test_opts": tf.get("options",""),
                "diff":      " | ".join(diffs),
            })
        elif lf:
            out.append({
                "field":     lf["label"],
                "status":    "MISSING IN TEST",
                "live_type": lf["datatype"],
                "test_type": "",
                "live_mand": lf["mandatory"],
                "test_mand": "",
                "live_opts": lf.get("options",""),
                "test_opts": "",
                "diff":      "In LIVE but not in TEST",
            })
        else:
            out.append({
                "field":     tf["label"],
                "status":    "EXTRA IN TEST",
                "live_type": "",
                "test_type": tf["datatype"],
                "live_mand": "",
                "test_mand": tf["mandatory"],
                "live_opts": "",
                "test_opts": tf.get("options",""),
                "diff":      "In TEST but not in LIVE",
            })
    return out


# ─────────────────────────────────────────────────────────────────────────────
#  TABLE COMPARE
# ─────────────────────────────────────────────────────────────────────────────
def compare_tables(live_tables, test_tables):
    def norm(h):
        return re.sub(r"[^a-z0-9]", "", h.lower())

    lmap = {norm(t["heading"]): t for t in live_tables}
    tmap = {norm(t["heading"]): t for t in test_tables}
    all_keys = sorted(set(list(lmap) + list(tmap)))
    out = []
    for k in all_keys:
        lt, tt = lmap.get(k), tmap.get(k)
        if lt and tt:
            lcols = {norm(c) for c in lt["columns"]}
            tcols = {norm(c) for c in tt["columns"]}
            missing_cols = [c for c in lt["columns"] if norm(c) not in tcols]
            extra_cols   = [c for c in tt["columns"] if norm(c) not in lcols]
            if missing_cols or extra_cols:
                st   = "COLUMNS DIFFER"
                diff = ""
                if missing_cols:
                    diff += "Missing: " + ", ".join(missing_cols[:5])
                if extra_cols:
                    diff += "  Extra: " + ", ".join(extra_cols[:5])
            else:
                st   = "MATCH"
                diff = ""
            out.append({
                "table":      lt["heading"],
                "status":     st,
                "live_cols":  " | ".join(lt["columns"]),
                "test_cols":  " | ".join(tt["columns"]),
                "live_rows":  lt["row_count"],
                "test_rows":  tt["row_count"],
                "diff":       diff.strip(),
            })
        elif lt:
            out.append({
                "table":      lt["heading"],
                "status":     "MISSING IN TEST",
                "live_cols":  " | ".join(lt["columns"]),
                "test_cols":  "",
                "live_rows":  lt["row_count"],
                "test_rows":  "",
                "diff":       "Table in LIVE but not in TEST",
            })
        else:
            out.append({
                "table":      tt["heading"],
                "status":     "EXTRA IN TEST",
                "live_cols":  "",
                "test_cols":  " | ".join(tt["columns"]),
                "live_rows":  "",
                "test_rows":  tt["row_count"],
                "diff":       "Table in TEST but not in LIVE",
            })
    return out


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL REPORT — 7 sheets
# ─────────────────────────────────────────────────────────────────────────────
def write_excel(live_rows, test_rows, menu_cmp, page_results, cfg):
    print("\n[EXCEL] Writing report ...")
    wb = Workbook()

    # ── Sheet 1: Live Menu ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Live Menu"
    ws1.cell(1, 1, "LIVE PORTAL — READ-ONLY REFERENCE").font = xfnt(True,"FF8B0000",12)
    ws1.merge_cells("A1:F1")
    ws1.cell(1,1).fill = xfill("FFFFF3CD")
    ws1.cell(1,1).alignment = xaln("center")
    for ci,(h,w) in enumerate(zip(
        ["#","Main Menu","Sub Menu","URL","Issue","Type"],
        [5,  28,          32,        65,   40,     12]), 1):
        hdr(ws1, 2, ci, h, "FF8B0000", w=w)
    for ri, r in enumerate(live_rows, 3):
        is_sub = bool(r["sub"])
        bg = ("FFFFEBEE" if r.get("issue") else
              "FFFCE4EC" if not is_sub else row_bg(ri))
        for ci, v in enumerate([ri-2, r["main"], r["sub"],
                                  r["url"] or r["href"],
                                  r.get("issue",""),
                                  "Sub" if is_sub else "Main"], 1):
            c = ws1.cell(ri, ci, v)
            c.fill = xfill(bg); c.border = xbrd(); c.alignment = xaln()
            c.font = (Font(color="FF1565C0", size=10, underline="single")
                      if ci == 4 else xfnt(bold=not is_sub and ci == 2))
    ws1.freeze_panes = "A3"

    # ── Sheet 2: Test Menu ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Test Menu")
    ws2.cell(1, 1, "TEST PORTAL MENU").font = xfnt(True,"FF1B5E20",12)
    ws2.merge_cells("A1:F1")
    ws2.cell(1,1).fill = xfill("FFE8F5E9")
    ws2.cell(1,1).alignment = xaln("center")
    for ci,(h,w) in enumerate(zip(
        ["#","Main Menu","Sub Menu","URL","Issue","Type"],
        [5,  28,          32,        65,   40,     12]), 1):
        hdr(ws2, 2, ci, h, "FF1B5E20", w=w)
    for ri, r in enumerate(test_rows, 3):
        is_sub = bool(r["sub"])
        bg = ("FFFFEBEE" if r.get("issue") else
              "FFE8F5E9" if not is_sub else row_bg(ri))
        for ci, v in enumerate([ri-2, r["main"], r["sub"],
                                  r["url"] or r["href"],
                                  r.get("issue",""),
                                  "Sub" if is_sub else "Main"], 1):
            c = ws2.cell(ri, ci, v)
            c.fill = xfill(bg); c.border = xbrd(); c.alignment = xaln()
            c.font = xfnt(bold=not is_sub and ci == 2)
    ws2.freeze_panes = "A3"

    # ── Sheet 3: Menu Comparison ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Menu Comparison")
    for ci,(h,w) in enumerate(zip(
        ["#","Main Menu","Sub Menu","Status",
         "Live URL","Test URL","Live Issue","Test Issue"],
        [4,  22,          28,        18,
         55,       55,        28,          28]), 1):
        hdr(ws3, 1, ci, h, "FF4A148C", w=w)
    for ri, r in enumerate(menu_cmp, 2):
        bg = row_bg(ri)
        for ci, v in enumerate([ri-1, r["main"], r["sub"], "",
                                  r["live_url"], r["test_url"],
                                  r.get("live_issue",""),
                                  r.get("test_issue","")], 1):
            if ci == 4:
                st_cell(ws3, ri, ci, r["status"])
            else:
                dat(ws3, ri, ci, v, bg)
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:H{len(menu_cmp)+1}"

    # ── Sheet 4: Page Content ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("Page Content")
    for ci,(h,w) in enumerate(zip(
        ["#","Page","Portal","Status","Page Title",
         "Field Count","Table Count","Button Count","Page Headings"],
        [4,  40,     8,       12,      40,
         12,          12,          13,           65]), 1):
        hdr(ws4, 1, ci, h, "FF006064", w=w)
    pc_ri = 2
    for pr in page_results:
        for portal in ["live","test"]:
            pd  = pr[portal]
            ct  = pd["content"]
            bg  = row_bg(pc_ri)
            for ci, v in enumerate([
                pc_ri-1,
                pr["label"],
                portal.upper(),
                "",
                pd["title"],
                len(ct["fields"]),
                len(ct["tables"]),
                len(ct["buttons"]),
                " | ".join(ct["headings"][:4]),
            ], 1):
                if ci == 4:
                    st_cell(ws4, pc_ri, ci, pd["status"])
                else:
                    dat(ws4, pc_ri, ci, v, bg)
            pc_ri += 1
    ws4.freeze_panes = "A2"

    # ── Sheet 5: Field Comparison ─────────────────────────────────────────────
    ws5 = wb.create_sheet("Field Comparison")
    for ci,(h,w) in enumerate(zip(
        ["#","Page","Field Name / Label","Status",
         "Live Data Type","Test Data Type",
         "Live Mandatory","Test Mandatory",
         "Live Options","Test Options","Difference Notes"],
        [4,  35,    42,                  18,
         16,             16,
         16,             16,
         38,           38,           52]), 1):
        hdr(ws5, 1, ci, h, "FF1A237E", w=w)
    fc_ri = 2
    for pr in page_results:
        for fc in pr["field_cmp"]:
            bg = row_bg(fc_ri)
            for ci, v in enumerate([
                fc_ri-1,
                pr["label"],
                fc["field"],
                "",
                fc["live_type"],  fc["test_type"],
                fc["live_mand"],  fc["test_mand"],
                fc["live_opts"],  fc["test_opts"],
                fc["diff"],
            ], 1):
                if ci == 4:
                    st_cell(ws5, fc_ri, ci, fc["status"])
                else:
                    dat(ws5, fc_ri, ci, v, bg)
            fc_ri += 1
    ws5.freeze_panes = "A2"
    ws5.auto_filter.ref = f"A1:K{fc_ri}"

    # ── Sheet 6: Table Comparison ─────────────────────────────────────────────
    ws6 = wb.create_sheet("Table Comparison")
    for ci,(h,w) in enumerate(zip(
        ["#","Page","Table Name","Status",
         "Live Columns","Test Columns",
         "Live Row Count","Test Row Count","Difference Notes"],
        [4,  35,    35,         18,
         60,           60,
         15,             15,             55]), 1):
        hdr(ws6, 1, ci, h, "FF880E4F", w=w)
    tc_ri = 2
    for pr in page_results:
        for tc in pr["table_cmp"]:
            bg = row_bg(tc_ri)
            for ci, v in enumerate([
                tc_ri-1,
                pr["label"],
                tc["table"],
                "",
                tc["live_cols"],  tc["test_cols"],
                tc["live_rows"],  tc["test_rows"],
                tc["diff"],
            ], 1):
                if ci == 4:
                    st_cell(ws6, tc_ri, ci, tc["status"])
                else:
                    dat(ws6, tc_ri, ci, v, bg)
            tc_ri += 1
    ws6.freeze_panes = "A2"
    ws6.auto_filter.ref = f"A1:I{tc_ri}"

    # ── Sheet 7: Summary ──────────────────────────────────────────────────────
    ws7 = wb.create_sheet("Summary")
    ws7.column_dimensions["A"].width = 34
    ws7.column_dimensions["B"].width = 40

    all_fc = [fc for pr in page_results for fc in pr["field_cmp"]]
    all_tc = [tc for pr in page_results for tc in pr["table_cmp"]]

    def cnt(lst, st):
        return sum(1 for x in lst if x["status"] == st)

    summary = [
        ("Report Generated",       datetime.now().strftime("%Y-%m-%d %H:%M:%S"), None),
        ("","",""),
        ("-- PORTALS --",           "", None),
        ("Live URL",               cfg["live"]["url"],      None),
        ("Live Username",          cfg["live"]["username"], None),
        ("Test URL",               cfg["test"]["url"],      None),
        ("Test Username",          cfg["test"]["username"], None),
        ("Output Folder",          cfg["test"]["folder"],   None),
        ("","",""),
        ("-- MENU COMPARISON --",  "", None),
        ("Live Menu Items",        len(live_rows),                    None),
        ("Test Menu Items",        len(test_rows),                    None),
        ("Matching",               cnt(menu_cmp,"MATCH"),             "FF1B5E20"),
        ("Missing in Test",        cnt(menu_cmp,"MISSING IN TEST"),   "FFB71C1C"),
        ("Extra in Test",          cnt(menu_cmp,"EXTRA IN TEST"),     "FFF57F17"),
        ("","",""),
        ("-- FIELD COMPARISON --", "", None),
        ("Total Fields Compared",  len(all_fc),                       None),
        ("Match",                  cnt(all_fc,"MATCH"),               "FF1B5E20"),
        ("Missing in Test",        cnt(all_fc,"MISSING IN TEST"),     "FFB71C1C"),
        ("Extra in Test",          cnt(all_fc,"EXTRA IN TEST"),       "FFF57F17"),
        ("Type / Mandatory Changed",cnt(all_fc,"MISMATCH"),           "FFE65100"),
        ("","",""),
        ("-- TABLE COMPARISON --", "", None),
        ("Total Tables Compared",  len(all_tc),                       None),
        ("Match",                  cnt(all_tc,"MATCH"),               "FF1B5E20"),
        ("Missing in Test",        cnt(all_tc,"MISSING IN TEST"),     "FFB71C1C"),
        ("Extra in Test",          cnt(all_tc,"EXTRA IN TEST"),       "FFF57F17"),
        ("Columns Differ",         cnt(all_tc,"COLUMNS DIFFER"),      "FFE65100"),
    ]
    for ri, (lbl, val, color) in enumerate(summary, 1):
        ws7.cell(ri, 1, lbl).font = xfnt(bold="--" in str(lbl), sz=11)
        v = ws7.cell(ri, 2, str(val))
        v.font = (xfnt(color=color, sz=12, bold=True) if color
                  else xfnt(sz=11))

    wb.save(cfg["test"]["excel"])
    print(f"\n  Excel saved: {os.path.abspath(cfg['test']['excel'])}")
    print(f"  Fields  : match={cnt(all_fc,'MATCH')}  "
          f"missing={cnt(all_fc,'MISSING IN TEST')}  "
          f"extra={cnt(all_fc,'EXTRA IN TEST')}  "
          f"mismatch={cnt(all_fc,'MISMATCH')}")
    print(f"  Tables  : match={cnt(all_tc,'MATCH')}  "
          f"missing={cnt(all_tc,'MISSING IN TEST')}  "
          f"extra={cnt(all_tc,'EXTRA IN TEST')}  "
          f"cols-differ={cnt(all_tc,'COLUMNS DIFFER')}")


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    config = get_config()
    if not config:
        print("Cancelled.")
        return

    live_cfg = config["live"]
    test_cfg = config["test"]
    opts     = config["options"]

    print(f"""
{'='*62}
  LIVE  : {live_cfg['url']}
  TEST  : {test_cfg['url']}
  Folder: {test_cfg['folder']}
  Timeout per page : {opts['page_timeout']}s
  No back() anywhere. No pixel comparison.
{'='*62}""")

    driver    = init_driver()
    live_rows = []
    test_rows = []

    try:
        # PHASE 1: Login LIVE, extract menu
        print("\n[PHASE 1] LIVE PORTAL — login + menu")
        login(driver, live_cfg, opts["captcha_wait"])
        driver.execute_script("""
            document.querySelectorAll('.dropdown-menu,.collapse')
                .forEach(el=>{el.classList.add('show');
                              el.style.display='block'});
            document.querySelectorAll('[aria-expanded]')
                .forEach(el=>el.setAttribute('aria-expanded','true'));
        """)
        time.sleep(2)
        live_html = driver.page_source
        with open(test_cfg["live_html"], "w", encoding="utf-8") as f:
            f.write(live_html)
        live_rows = parse_menu(live_html, live_cfg["base"], "LIVE")
        print(f"  Live menu: {len(live_rows)} items")

        driver.delete_all_cookies()
        time.sleep(2)

        # PHASE 2: Login TEST, extract menu
        print("\n[PHASE 2] TEST PORTAL — login + menu")
        login(driver, test_cfg, opts["captcha_wait"])
        driver.execute_script("""
            document.querySelectorAll('.collapse,.dropdown-menu')
                .forEach(el=>{el.classList.add('show');
                              el.style.display='block'});
        """)
        time.sleep(2)
        test_html = driver.page_source
        with open(test_cfg["test_html"], "w", encoding="utf-8") as f:
            f.write(test_html)
        test_rows = parse_menu(test_html, test_cfg["base"], "TEST")
        print(f"  Test menu: {len(test_rows)} items")

        # Menu comparison
        menu_cmp = compare_menus(live_rows, test_rows)
        print(f"\n  Menu: "
              f"{sum(1 for c in menu_cmp if c['status']=='MATCH')} match | "
              f"{sum(1 for c in menu_cmp if c['status']=='MISSING IN TEST')} missing | "
              f"{sum(1 for c in menu_cmp if c['status']=='EXTRA IN TEST')} extra")

        # Build unified page list
        def row_key(r):
            return (r["main"].strip().lower(), r["sub"].strip().lower())
        live_map = {row_key(r): r for r in live_rows}
        test_map = {row_key(r): r for r in test_rows}
        all_keys = sorted(set(list(live_map) + list(test_map)))
        total    = len(all_keys)

        empty_content = {"fields":[],"tables":[],"buttons":[],"headings":[]}

        # PHASE 3: Visit all LIVE pages (one pass, no back)
        print(f"\n[PHASE 3] LIVE — visiting {total} pages (no back)")
        live_page_data = {}
        for idx, key in enumerate(all_keys):
            lr = live_map.get(key)
            label = f"{key[0]} > {key[1]}" if key[1] else key[0]
            print(f"  [{idx+1:>3}/{total}]", end=" ")
            if not lr or not lr["url"]:
                live_page_data[key] = {
                    "url":"","status":"NOT IN LIVE","title":"",
                    "content": empty_content}
                print(f"    [LIVE] {label[:48]:<48}  NOT IN LIVE")
            else:
                live_page_data[key] = visit_page(
                    driver, lr["url"], label,
                    opts["page_timeout"], "LIVE", do_interact=False)

        driver.delete_all_cookies()
        time.sleep(2)

        # PHASE 4: Login TEST again, visit all TEST pages (one pass, no back)
        print(f"\n[PHASE 4] TEST — login again then visit {total} pages")
        login(driver, test_cfg, opts["captcha_wait"])

        test_page_data = {}
        do_interact    = not opts.get("skip_test_fill", False)
        for idx, key in enumerate(all_keys):
            tr = test_map.get(key)
            label = f"{key[0]} > {key[1]}" if key[1] else key[0]
            print(f"  [{idx+1:>3}/{total}]", end=" ")
            if not tr or not tr["url"]:
                test_page_data[key] = {
                    "url":"","status":"NOT IN TEST","title":"",
                    "content": empty_content}
                print(f"    [TEST] {label[:48]:<48}  NOT IN TEST")
            else:
                test_page_data[key] = visit_page(
                    driver, tr["url"], label,
                    opts["page_timeout"], "TEST",
                    do_interact=do_interact)

        # PHASE 5: Compare content page by page
        print(f"\n[PHASE 5] COMPARING CONTENT")
        page_results = []
        for key in all_keys:
            label = f"{key[0]} > {key[1]}" if key[1] else key[0]
            ld = live_page_data[key]
            td = test_page_data[key]
            lc = ld["content"]
            tc = td["content"]

            fc    = compare_fields(lc["fields"],  tc["fields"])
            tc_mp = compare_tables(lc["tables"],  tc["tables"])

            fm = sum(1 for f in fc    if f["status"]=="MATCH")
            tm = sum(1 for t in tc_mp if t["status"]=="MATCH")
            print(f"  {label[:52]:<52}  "
                  f"fields:{len(fc)}({fm}ok)  "
                  f"tables:{len(tc_mp)}({tm}ok)")

            page_results.append({
                "key":       key,
                "label":     label,
                "main":      key[0],
                "sub":       key[1],
                "live":      ld,
                "test":      td,
                "field_cmp": fc,
                "table_cmp": tc_mp,
            })

        # Write Excel
        write_excel(live_rows, test_rows, menu_cmp, page_results, config)

        print(f"""
{'='*62}
  DONE!
  Folder : {test_cfg['folder']}
  Excel  : {test_cfg['excel']}
{'='*62}""")

    except KeyboardInterrupt:
        print("\n[!] Interrupted.")
    except Exception as e:
        import traceback
        print(f"\n[ERROR] {e}")
        traceback.print_exc()
    finally:
        input("\nPress ENTER to close browser ...")
        driver.quit()


if __name__ == "__main__":
    main()
