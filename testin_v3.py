"""
Universal Portal Testing Tool v1.0
- Login with CAPTCHA support
- Extract sidebar menu (Level 1 + Level 2)
- Save to CSV
- Hit every URL and check status
- Create Excel test report
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import time, os, csv
from datetime import datetime

URL      = "https://drt.etribunals.gov.in/cis2.0/filing/login"
USERNAME = "filingdrt1"
PASSWORD = "Admin@123"
BASE     = "https://drt.etribunals.gov.in"
CSV_FILE  = "menu.csv"
EXCEL_FILE = "test_report.xlsx"
HTML_FILE  = "portal.html"

# ── STEP 1: Launch browser and login ─────────────────────────────────────────
def login():
    opts = webdriver.ChromeOptions()
    opts.add_argument("--start-maximized")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=opts)

    print("[1] Opening login page...")
    driver.get(URL)
    time.sleep(3)

    # Fill credentials
    try:
        driver.find_element(By.ID, "user_name").send_keys(USERNAME)
        print("    Username filled")
    except:
        print("    Could not fill username - do it manually")

    try:
        driver.find_element(By.ID, "user_pass").send_keys(PASSWORD)
        print("    Password filled")
    except:
        print("    Could not fill password - do it manually")

    print("\n" + "="*50)
    print("  Solve CAPTCHA and click LOGIN button")
    print("  Waiting 120 seconds...")
    print("="*50 + "\n")

    # Wait for page to change after login
    WebDriverWait(driver, 120).until(
        lambda d: "login" not in d.current_url
    )
    print("[OK] Logged in:", driver.current_url)
    time.sleep(3)
    return driver

# ── STEP 2: Expand sidebar and save HTML ─────────────────────────────────────
def get_html(driver):
    print("\n[2] Expanding sidebar menus...")

    # Force expand all Bootstrap collapse items
    driver.execute_script("""
        document.querySelectorAll('.collapse').forEach(el => {
            el.classList.add('show');
        });
        document.querySelectorAll('[aria-expanded]').forEach(el => {
            el.setAttribute('aria-expanded', 'true');
        });
    """)
    time.sleep(2)

    html = driver.page_source
    with open(HTML_FILE, "w", encoding="utf-8") as f:
        f.write(html)
    print("    HTML saved:", HTML_FILE)
    return html

# ── STEP 3: Parse menu from HTML ─────────────────────────────────────────────
def parse_menu(html):
    print("\n[3] Parsing menu items...")
    soup = BeautifulSoup(html, "html.parser")
    rows = []

    sidebar = soup.find("ul", id="accordionSidebar")
    if not sidebar:
        print("    [!] accordionSidebar not found, trying generic nav")
        sidebar = soup.find("ul", class_=lambda c: c and "navbar-nav" in c)

    if sidebar:
        for li in sidebar.find_all("li", class_="nav-item"):
            main_a = li.find("a", class_="nav-link")
            if not main_a:
                continue
            main_name = main_a.get_text(strip=True)

            collapse = li.find("div", class_="collapse-inner")
            if collapse:
                for a in collapse.find_all("a", class_="dropdown-item"):
                    sub  = a.get_text(strip=True)
                    href = a.get("href", "")
                    url  = (BASE + href) if href.startswith("/") else href
                    rows.append([main_name, sub, url])
                    print("    +", main_name, ">", sub)
            else:
                href = main_a.get("href", "")
                if href and "javascript" not in href:
                    url = (BASE + href) if href.startswith("/") else href
                    rows.append([main_name, "", url])
                    print("    +", main_name, "(direct)")

    print("    Total:", len(rows), "items")
    return rows

# ── STEP 4: Save CSV ──────────────────────────────────────────────────────────
def save_csv(rows):
    with open(CSV_FILE, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Main Menu", "Sub Menu", "URL"])
        w.writerows(rows)
    print("\n[4] CSV saved:", CSV_FILE)

# ── STEP 5: Test each URL ─────────────────────────────────────────────────────
def test_urls(driver, rows):
    print("\n[5] Testing", len(rows), "URLs...")
    results = []
    for i, (main, sub, url) in enumerate(rows):
        label = (main + " > " + sub) if sub else main
        print("  [" + str(i+1).rjust(3) + "/" + str(len(rows)) + "]",
              label[:45].ljust(45), end=" ... ", flush=True)

        if not url:
            print("NO URL")
            results.append([main, sub, url, "NO URL", "", "", "", ""])
            continue

        try:
            t0 = time.time()
            driver.get(url)
            time.sleep(1.2)
            load = round(time.time() - t0, 2)
            title = driver.title[:50]

            soup  = BeautifulSoup(driver.page_source, "html.parser")
            body  = soup.get_text(separator=" ", strip=True).lower()

            forms   = len(soup.find_all("form"))
            tables  = len(soup.find_all("table"))

            err_kw  = ["access denied", "not found", "unauthorized", "404", "403", "500", "error"]
            errors  = [k for k in err_kw if k in body]
            status  = "FAIL" if errors else "PASS"

            print(status)
            results.append([main, sub, url, status, load, title,
                            str(forms) + " form(s)", str(tables) + " table(s)"])
        except Exception as e:
            print("ERROR")
            results.append([main, sub, url, "ERROR", "", "", str(e)[:60], ""])

    return results

# ── STEP 6: Write Excel report ────────────────────────────────────────────────
def write_excel(rows, results):
    print("\n[6] Writing Excel report...")

    def fill(c): return PatternFill("solid", fgColor=c)
    def fnt(bold=False, color="FF000000", sz=10): return Font(bold=bold, color=color, size=sz)
    def brd():
        s = Side(style="thin", color="FFD0D0D0")
        return Border(left=s, right=s, top=s, bottom=s)
    def aln(h="left"): return Alignment(horizontal=h, vertical="center", wrap_text=True)

    wb = Workbook()

    # Sheet 1 — Menu Structure
    ws1 = wb.active
    ws1.title = "Menu Structure"
    hdrs = ["#", "Level 1 (Main Menu)", "Level 2 (Sub Menu)", "URL"]
    wids = [5, 28, 35, 70]
    for ci, (h, w) in enumerate(zip(hdrs, wids), 1):
        c = ws1.cell(1, ci, h)
        c.fill = fill("FF003366"); c.font = fnt(True, "FFFFFFFF", 11)
        c.alignment = aln("center"); c.border = brd()
        ws1.column_dimensions[c.column_letter].width = w

    for ri, (main, sub, url) in enumerate(rows, 2):
        is_sub = bool(sub)
        bg = "FFE8F4FD" if not is_sub else ("FFFAFAFA" if ri%2==0 else "FFF5F5F5")
        vals = [ri-1, main, sub, url]
        for ci, val in enumerate(vals, 1):
            c = ws1.cell(ri, ci, val)
            c.fill = fill(bg); c.border = brd(); c.alignment = aln()
            c.font = Font(color="FF0055AA", size=10, underline="single") if ci==4 else fnt(bold=not is_sub and ci==2)
    ws1.freeze_panes = "A2"

    # Sheet 2 — Test Results
    ws2 = wb.create_sheet("Test Results")
    hdrs2 = ["#", "Main Menu", "Sub Menu", "URL", "Result", "Load(s)", "Page Title", "Forms", "Tables"]
    wids2 = [4, 22, 30, 55, 9, 8, 38, 10, 10]
    for ci, (h, w) in enumerate(zip(hdrs2, wids2), 1):
        c = ws2.cell(1, ci, h)
        c.fill = fill("FF005B96"); c.font = fnt(True, "FFFFFFFF", 11)
        c.alignment = aln("center"); c.border = brd()
        ws2.column_dimensions[c.column_letter].width = w

    pass_count = fail_count = error_count = 0
    for ri, res in enumerate(results, 2):
        st = res[3] if len(res) > 3 else ""
        if   st == "PASS":  bg_st = fill("FFE8F5E9"); ft_st = fnt(True, "FF1B5E20")
        elif st == "FAIL":  bg_st = fill("FFFFF3E0"); ft_st = fnt(True, "FFE65100")
        else:               bg_st = fill("FFFFEBEE"); ft_st = fnt(True, "FFB71C1C")
        if st == "PASS":  pass_count  += 1
        elif st == "FAIL":  fail_count  += 1
        else:               error_count += 1

        row_bg = "FFFAFAFA" if ri%2==0 else "FFFFFFFF"
        # Pad to 9 columns
        while len(res) < 9: res.append("")
        for ci, val in enumerate(res[:9], 1):
            c = ws2.cell(ri, ci, val if ci != 1 else ri-1)
            c.fill = bg_st if ci == 5 else fill(row_bg)
            c.font = ft_st if ci == 5 else fnt()
            c.alignment = aln(); c.border = brd()
    ws2.freeze_panes = "A2"

    # Sheet 3 — Summary
    ws3 = wb.create_sheet("Summary")
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 22
    summary = [
        ("Report Date",        datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Portal URL",         URL),
        ("Username",           USERNAME),
        ("", ""),
        ("Total Menu Items",   len(rows)),
        ("Total URLs Tested",  len(results)),
        ("", ""),
        ("✅ PASS",            pass_count),
        ("⚠️  FAIL",           fail_count),
        ("❌ ERROR",           error_count),
    ]
    for ri, (lbl, val) in enumerate(summary, 1):
        ws3.cell(ri, 1, lbl).font = fnt(bold=True, sz=11)
        ws3.cell(ri, 2, str(val)).font = fnt(sz=11)

    wb.save(EXCEL_FILE)
    print("    Excel saved:", os.path.abspath(EXCEL_FILE))
    print("\n    PASS:", pass_count, " | FAIL:", fail_count, " | ERROR:", error_count)

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    driver = login()
    try:
        html   = get_html(driver)
        rows   = parse_menu(html)
        save_csv(rows)
        results = test_urls(driver, rows)
        write_excel(rows, results)
        print("\n" + "="*50)
        print("  DONE!")
        print("  CSV  :", CSV_FILE)
        print("  Excel:", EXCEL_FILE)
        print("="*50)
    finally:
        input("\nPress ENTER to close browser...")
        driver.quit()

if __name__ == "__main__":
    main()