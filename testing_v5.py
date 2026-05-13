"""
Universal Portal Testing Tool v1.2
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import tkinter as tk
from tkinter import ttk, messagebox
import time, os, csv, re
from datetime import datetime
from urllib.parse import urlparse

def get_config():
    config = {}
    def submit():
        url  = url_var.get().strip()
        user = user_var.get().strip()
        pwd  = pass_var.get().strip()
        if not url or not user or not pwd:
            messagebox.showerror("Error", "All fields are required")
            return
        config["url"]      = url
        config["username"] = user
        config["password"] = pwd
        ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
        folder = user + "_" + ts
        ss_dir = os.path.join(folder, "screenshots")
        os.makedirs(ss_dir, exist_ok=True)
        config["folder"]      = folder
        config["screenshots"] = ss_dir
        config["html"]        = os.path.join(folder, user + "_portal.html")
        config["csv"]         = os.path.join(folder, user + "_menu.csv")
        config["excel"]       = os.path.join(folder, user + "_test_report.xlsx")
        config["base"]        = urlparse(url).scheme + "://" + urlparse(url).netloc
        root.destroy()

    root = tk.Tk()
    root.title("Portal Testing Tool v1.2")
    root.geometry("540x320")
    root.resizable(False, False)
    tk.Label(root, text="Universal Portal Testing Tool v1.2",
             font=("Arial", 15, "bold"), fg="#003366").pack(pady=14)
    tk.Label(root, text="All output files named after your username",
             font=("Arial", 9), fg="#777").pack()
    frame = ttk.Frame(root, padding=20)
    frame.pack(fill="both", expand=True)
    url_var  = tk.StringVar(value="https://drt.etribunals.gov.in/cis2.0/filing/login")
    user_var = tk.StringVar(value="")
    pass_var = tk.StringVar(value="")
    ttk.Label(frame, text="Portal URL:").grid(row=0, column=0, sticky="w", pady=8)
    ttk.Entry(frame, textvariable=url_var, width=46).grid(row=0, column=1, padx=8)
    ttk.Label(frame, text="Username:").grid(row=1, column=0, sticky="w", pady=8)
    ttk.Entry(frame, textvariable=user_var, width=46).grid(row=1, column=1, padx=8)
    ttk.Label(frame, text="Password:").grid(row=2, column=0, sticky="w", pady=8)
    ttk.Entry(frame, textvariable=pass_var, show="*", width=46).grid(row=2, column=1, padx=8)
    ttk.Button(frame, text="▶  Start Testing", command=submit).grid(row=3, column=1, pady=20, sticky="e")
    root.mainloop()
    return config

def init_driver():
    opts = webdriver.ChromeOptions()
    opts.add_argument("--start-maximized")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)

def login(driver, cfg):
    print("\n[1] Opening login page...")
    driver.get(cfg["url"])
    time.sleep(3)
    for fid in ["user_name", "username"]:
        try:
            driver.find_element(By.ID, fid).send_keys(cfg["username"])
            print("    Username filled"); break
        except: pass
    for fid in ["user_pass", "password"]:
        try:
            driver.find_element(By.ID, fid).send_keys(cfg["password"])
            print("    Password filled"); break
        except: pass
    driver.save_screenshot(os.path.join(cfg["screenshots"], "00_login.png"))
    print("\n" + "="*52)
    print("  Solve CAPTCHA then click LOGIN")
    print("  Waiting 120 seconds...")
    print("="*52 + "\n")
    WebDriverWait(driver, 120).until(lambda d: "login" not in d.current_url)
    time.sleep(3)
    driver.save_screenshot(os.path.join(cfg["screenshots"], "01_dashboard.png"))
    print("[OK] Logged in:", driver.current_url)

def get_html(driver, cfg):
    print("\n[2] Expanding ALL sidebar menus...")

    # Pass 1 — JS force show all collapse divs
    driver.execute_script("""
        document.querySelectorAll('.collapse').forEach(el => el.classList.add('show'));
        document.querySelectorAll('[aria-expanded]').forEach(el => el.setAttribute('aria-expanded','true'));
    """)
    time.sleep(1)

    # Pass 2 — click each toggle that controls a collapsed section
    toggles = driver.find_elements(By.CSS_SELECTOR, "[data-bs-toggle='collapse']")
    print("    Found", len(toggles), "collapsible toggles")
    for toggle in toggles:
        try:
            driver.execute_script("arguments[0].scrollIntoView(true);", toggle)
            time.sleep(0.1)
            target = (toggle.get_attribute("data-bs-target") or "").replace("#", "")
            if target:
                try:
                    el = driver.find_element(By.ID, target)
                    classes = el.get_attribute("class") or ""
                    if "show" not in classes:
                        driver.execute_script("arguments[0].click();", toggle)
                        time.sleep(0.3)
                except:
                    driver.execute_script("arguments[0].click();", toggle)
                    time.sleep(0.3)
            name = toggle.text.strip()[:35] or target
            print("    [+] Expanded:", name)
        except:
            pass

    # Pass 3 — final JS to ensure all visible
    driver.execute_script("""
        document.querySelectorAll('.collapse').forEach(el => el.classList.add('show'));
        document.querySelectorAll('[aria-expanded="false"]').forEach(el => {
            try { el.click(); } catch(e) {}
        });
    """)
    time.sleep(2)

    # Take screenshot with all menus open
    driver.save_screenshot(os.path.join(cfg["screenshots"], "02_sidebar_expanded.png"))

    html = driver.page_source
    with open(cfg["html"], "w", encoding="utf-8") as f:
        f.write(html)

    # Verify capture
    soup  = BeautifulSoup(html, "html.parser")
    items = soup.find_all("a", class_="dropdown-item")
    print("    Captured", len(items), "dropdown items")
    for item in items:
        print("      -", item.get_text(strip=True))
    print("    HTML saved:", cfg["html"])
    return html

def is_valid_url(href):
    if not href:                      return False
    if "javascript" in href.lower():  return False
    if re.search(r"['\",]", href):    return False
    return True

def parse_menu(html, cfg):
    print("\n[3] Parsing menu...")
    soup = BeautifulSoup(html, "html.parser")
    rows = []
    sidebar = soup.find("ul", id="accordionSidebar")
    if not sidebar:
        sidebar = soup.find("ul", class_=lambda c: c and "navbar-nav" in c)
    if sidebar:
        for li in sidebar.find_all("li", class_="nav-item"):
            main_a = li.find("a", class_="nav-link")
            if not main_a: continue
            main_name = main_a.get_text(strip=True)
            if not main_name: continue
            collapse = li.find("div", class_="collapse-inner")
            if collapse:
                for a in collapse.find_all("a", class_="dropdown-item"):
                    sub  = a.get_text(strip=True)
                    href = a.get("href", "").strip()
                    if not is_valid_url(href): continue
                    url = (cfg["base"] + href) if href.startswith("/") else href
                    rows.append([main_name, sub, url])
                    print("    +", main_name, ">", sub)
            else:
                href = main_a.get("href", "").strip()
                if is_valid_url(href):
                    url = (cfg["base"] + href) if href.startswith("/") else href
                    rows.append([main_name, "", url])
                    print("    +", main_name, "(direct)")
    print("    Total:", len(rows), "items")
    return rows

def save_csv(rows, cfg):
    with open(cfg["csv"], "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Main Menu", "Sub Menu", "URL"])
        w.writerows(rows)
    print("\n[4] CSV saved:", cfg["csv"])

def extract_components(soup):
    comps = {"forms": [], "buttons": [], "inputs": [], "tables": [], "dropdowns": []}
    for i, form in enumerate(soup.find_all("form"), 1):
        comps["forms"].append("Form" + str(i) + "[" + form.get("method","GET").upper() + "]")
    seen_btns = set()
    for btn in soup.find_all(["button", "input"]):
        btype = btn.get("type","").lower()
        if btype in ["submit","button","reset"] or btn.name == "button":
            label = btn.get_text(strip=True) or btn.get("value","") or btype
            if label and label not in seen_btns:
                comps["buttons"].append(label[:35])
                seen_btns.add(label)
    seen_inp = set()
    for inp in soup.find_all("input"):
        itype = inp.get("type","text").lower()
        if itype in ["hidden","submit","button","reset"]: continue
        name = inp.get("placeholder","") or inp.get("name","") or inp.get("id","")
        key  = itype + ":" + name
        if key not in seen_inp:
            comps["inputs"].append(key[:35])
            seen_inp.add(key)
    for i, tbl in enumerate(soup.find_all("table"), 1):
        hdrs = [th.get_text(strip=True) for th in tbl.find_all("th")][:4]
        comps["tables"].append("T" + str(i) + ":[" + ",".join(hdrs) + "]")
    for sel in soup.find_all("select"):
        name = sel.get("name","") or sel.get("id","")
        opts = [o.get_text(strip=True) for o in sel.find_all("option")][:3]
        comps["dropdowns"].append(name + ":[" + ",".join(opts) + "]")
    return comps

def test_urls(driver, rows, cfg):
    print("\n[5] Testing", len(rows), "URLs...")
    results = []
    for i, (main, sub, url) in enumerate(rows):
        label = (main + " > " + sub) if sub else main
        print("  [" + str(i+1).rjust(3) + "/" + str(len(rows)) + "]",
              label[:42].ljust(42), end=" ... ", flush=True)
        if not url:
            print("NO URL")
            results.append({"main":main,"sub":sub,"url":url,"status":"NO URL",
                "load":"","title":"","forms":"","buttons":"","inputs":"",
                "tables":"","dropdowns":"","screenshot":""})
            continue
        try:
            t0 = time.time()
            driver.get(url)
            time.sleep(1.5)
            load  = round(time.time() - t0, 2)
            title = driver.title[:55]
            safe  = re.sub(r"[^\w\-]", "_", label)[:45]
            ss_name = str(i+1).zfill(3) + "_" + safe + ".png"
            driver.save_screenshot(os.path.join(cfg["screenshots"], ss_name))
            soup  = BeautifulSoup(driver.page_source, "html.parser")
            body  = soup.get_text(separator=" ", strip=True).lower()
            comps = extract_components(soup)
            errors = [k for k in ["access denied","not found","unauthorized","404","403","500"] if k in body]
            status = "FAIL" if errors else "PASS"
            print(status + " (" + str(load) + "s)")
            results.append({
                "main": main, "sub": sub, "url": url,
                "status": status, "load": load, "title": title,
                "forms":     " | ".join(comps["forms"])[:80],
                "buttons":   " | ".join(comps["buttons"])[:80],
                "inputs":    " | ".join(comps["inputs"])[:80],
                "tables":    " | ".join(comps["tables"])[:80],
                "dropdowns": " | ".join(comps["dropdowns"])[:80],
                "screenshot": ss_name,
            })
        except Exception as e:
            print("ERROR")
            results.append({"main":main,"sub":sub,"url":url,"status":"ERROR",
                "load":"","title":str(e)[:60],"forms":"","buttons":"",
                "inputs":"","tables":"","dropdowns":"","screenshot":""})
    return results

def write_excel(rows, results, cfg):
    print("\n[6] Writing Excel report...")
    def fill(c):  return PatternFill("solid", fgColor=c)
    def fnt(bold=False, color="FF000000", sz=10): return Font(bold=bold, color=color, size=sz)
    def brd():
        s = Side(style="thin", color="FFD0D0D0")
        return Border(left=s, right=s, top=s, bottom=s)
    def aln(h="left"): return Alignment(horizontal=h, vertical="center", wrap_text=True)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Menu Structure"
    for ci, (h, w) in enumerate(zip(["#","Level 1 (Main Menu)","Level 2 (Sub Menu)","URL"],[5,28,35,70]),1):
        c = ws1.cell(1, ci, h)
        c.fill = fill("FF003366"); c.font = fnt(True,"FFFFFFFF",11)
        c.alignment = aln("center"); c.border = brd()
        ws1.column_dimensions[c.column_letter].width = w
    for ri, (main, sub, url) in enumerate(rows, 2):
        is_sub = bool(sub)
        bg = "FFE8F4FD" if not is_sub else ("FFFAFAFA" if ri%2==0 else "FFF5F5F5")
        for ci, val in enumerate([ri-1, main, sub, url], 1):
            c = ws1.cell(ri, ci, val)
            c.fill = fill(bg); c.border = brd(); c.alignment = aln()
            c.font = Font(color="FF0055AA",size=10,underline="single") if ci==4 else fnt(bold=not is_sub and ci==2)
    ws1.freeze_panes = "A2"

    ws2 = wb.create_sheet("Test Results")
    cols = ["#","Main Menu","Sub Menu","URL","Result","Load(s)","Page Title","Forms","Buttons","Input Fields","Tables","Dropdowns","Screenshot"]
    wids = [4,20,25,50,9,7,32,25,30,30,30,30,32]
    for ci, (h, w) in enumerate(zip(cols, wids), 1):
        c = ws2.cell(1, ci, h)
        c.fill = fill("FF005B96"); c.font = fnt(True,"FFFFFFFF",11)
        c.alignment = aln("center"); c.border = brd()
        ws2.column_dimensions[c.column_letter].width = w

    pass_c = fail_c = err_c = 0
    for ri, r in enumerate(results, 2):
        st = r["status"]
        if   st=="PASS": sf=fill("FFE8F5E9"); ff=fnt(True,"FF1B5E20")
        elif st=="FAIL": sf=fill("FFFFF3E0"); ff=fnt(True,"FFE65100")
        else:            sf=fill("FFFFEBEE"); ff=fnt(True,"FFB71C1C")
        if st=="PASS": pass_c+=1
        elif st=="FAIL": fail_c+=1
        else: err_c+=1
        bg = "FFFAFAFA" if ri%2==0 else "FFFFFFFF"
        vals = [ri-1,r["main"],r["sub"],r["url"],st,r["load"],r["title"],
                r["forms"],r["buttons"],r["inputs"],r["tables"],r["dropdowns"],r["screenshot"]]
        for ci, val in enumerate(vals, 1):
            c = ws2.cell(ri, ci, val)
            c.fill = sf if ci==5 else fill(bg)
            c.font = ff if ci==5 else fnt()
            c.alignment = aln(); c.border = brd()
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("Summary")
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 30
    for ri, (lbl, val) in enumerate([
        ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Portal URL",       cfg["url"]),
        ("Login User",       cfg["username"]),
        ("Output Folder",    cfg["folder"]),
        ("",""),
        ("Total Menu Items", len(rows)),
        ("URLs Tested",      len(results)),
        ("",""),
        ("✅  PASS",         pass_c),
        ("⚠️   FAIL",        fail_c),
        ("❌  ERROR",        err_c),
    ], 1):
        ws3.cell(ri,1,lbl).font = fnt(True,sz=11)
        ws3.cell(ri,2,str(val)).font = fnt(sz=11)

    wb.save(cfg["excel"])
    print("    Saved:", os.path.abspath(cfg["excel"]))
    print("\n    PASS:", pass_c, "| FAIL:", fail_c, "| ERROR:", err_c)

def main():
    cfg = get_config()
    if not cfg:
        print("Cancelled."); return

    print("\n" + "="*52)
    print("  User  :", cfg["username"])
    print("  Folder:", cfg["folder"])
    print("="*52)

    driver = init_driver()
    try:
        login(driver, cfg)
        html    = get_html(driver, cfg)
        rows    = parse_menu(html, cfg)
        save_csv(rows, cfg)
        results = test_urls(driver, rows, cfg)
        write_excel(rows, results, cfg)
        print("\n" + "="*52)
        print("  DONE!")
        print("  Folder:", cfg["folder"])
        print("  CSV   :", cfg["csv"])
        print("  Excel :", cfg["excel"])
        print("="*52)
    except Exception as e:
        print("\n[ERROR]", e)
        import traceback; traceback.print_exc()
    finally:
        input("\nPress ENTER to close browser...")
        driver.quit()

if __name__ == "__main__":
    main()